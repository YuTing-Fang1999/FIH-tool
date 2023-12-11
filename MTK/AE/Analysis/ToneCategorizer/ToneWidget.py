from PyQt5 import QtCore
from PyQt5.QtWidgets import QWidget, QApplication, QFileDialog, QMessageBox
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QUrl
from PyQt5.QtGui import QDesktopServices
from .ToneUI import Ui_Form
from .explain1 import Ui_Form_explain1
from .explain2 import Ui_Form_explain2
from time import sleep
from myPackage.ParentWidget import ParentWidget
from PyQt5.QtWebEngineWidgets import QWebEngineView
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from urllib.request import urlopen
import requests
import configparser
import numpy as np
import subprocess
import sys
from .mtkclassifyTONEanalysis_SX3 import main as mtk_main
import tkinter as tk
from tkinter import filedialog
import os
import subprocess
from openpyxl import load_workbook
import urllib.request
from urllib.parse import urljoin


class SolverThread(QThread):
    update_status_bar_signal = pyqtSignal(str)
    failed_signal = pyqtSignal(str)
    finish_signal = pyqtSignal()
    data = None

    def __init__(self):
        super().__init__()

    def run(self):
        try:
            # . . . (要執行的程式)
            mtk_main(self.data['your_path'], self.data['file_path'],
                     self.data['LV_region'], self.data['DR_region'])
            self.finish_signal.emit()
        except Exception as error:
            print(error)
            self.update_status_bar_signal.emit("Failed...")
            self.failed_signal.emit("Failed...\n"+str(error))


class MyWidget(ParentWidget):
    def __init__(self):
        super().__init__()  # in python3, super(Class, self).xxx = super().xxx
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.solver_thread = SolverThread()
        self.controller()

    def controller(self):
        self.ui.btn_Browse_DataFolder.clicked.connect(
            self.load_data_path_dataFolder)
        self.ui.btn_Browse_Tonecpp.clicked.connect(
            self.load_data_path_Tonecpp)
        self.ui.btn_download_1.clicked.connect(self.do_download1)
        self.ui.btn_explain_1.clicked.connect(self.do_explain1)
        self.ui.btn_download_2.clicked.connect(self.do_download2)
        self.ui.btn_explain_2.clicked.connect(self.do_explain2)
        self.ui.btn_Reset.clicked.connect(self.do_reset)
        self.ui.btn_Category.clicked.connect(self.do_category)
        self.ui.btn_OpenFolder.clicked.connect(self.do_openfolder)
        self.ui.btn_Update.clicked.connect(self.do_update)
        # self.solver_thread.update_status_bar_signal.connect(self.update_status_bar)
        self.solver_thread.failed_signal.connect(self.failed)
        self.solver_thread.finish_signal.connect(self.solver_finish)

    def load_data_path_dataFolder(self):
        your_path = QFileDialog.getExistingDirectory(
            self, "選擇Data Path", self.get_path("./"))

        if your_path == '':
            return
        # self.solver_thread.dir_path = filepath
        filefolder = '/'.join(your_path.split('/')[:-1])
        self.set_path("./", filefolder)

        self.ui.lineEdit_DataFolder.setText(your_path)
        return your_path

    def load_data_path_Tonecpp(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("cpp files", "*.cpp")])
        self.ui.lineEdit_Tonecpp.setText(file_path)
        return file_path

    # need test
    def do_download1(self):
        url = 'http://10.57.55.72/Camera_ImageQuality/Tuning/5.Tool_工具'
        # QDesktopServices.openUrl(QUrl(url)) # 有些未知情況會不能跑
        subprocess.run(["explorer", url])

    # change code after download1 is done
    def do_download2(self):
        url = 'http://10.57.55.72/Camera_ImageQuality/Tuning/5.Tool_工具/#NAME'
        # Direct to the url
        # QDesktopServices.openUrl(QUrl(url)) # 有些未知情況會不能跑
        subprocess.run(["explorer", url])

        # Save url to directory
        '''
        save_dir = self.ui.lineEdit_DataFolder.text() # Save to assigned folder
        save_dir = os.path.join(os.path.expanduser("~"), "Downloads") # Save to download folder
        print(save_dir)
        save_path, _ = QFileDialog.getSaveFileName(
            self, 'Save File', os.path.join(save_dir, 'downloaded_file'), 'All Files (*)')
        urllib.request.urlretrieve(url, save_path)
        '''

    def do_explain1(self):
        # tst. pic ISP module Exif
        self.ui = QWidget()
        self.ui_explain1 = Ui_Form_explain1()
        self.ui.setFixedSize(1330, 600)
        self.ui_explain1.setupUi(self.ui)
        self.ui.show()
        pass

    def do_explain2(self):
        # Set XLOOKUPs to Excel
        self.ui = QWidget()
        self.ui_explain2 = Ui_Form_explain2()
        self.ui.setFixedSize(795, 645)
        self.ui_explain2.setupUi(self.ui)
        self.ui.show()
        pass

    def do_reset(self):
        _translate = QtCore.QCoreApplication.translate
        self.ui.lineEdit_LV1.setText("1")
        self.ui.lineEdit_LV2.setText("25")
        self.ui.lineEdit_LV3.setText("55")
        self.ui.lineEdit_LV4.setText("85")
        self.ui.lineEdit_LV5.setText("115")
        self.ui.lineEdit_LV6.setText("135")
        self.ui.lineEdit_LV8.clear()
        self.ui.lineEdit_LV9.clear()
        self.ui.lineEdit_LV10.clear()

        self.ui.lineEdit_DR1.setText("1")
        self.ui.lineEdit_DR2.setText("50")
        self.ui.lineEdit_DR3.setText("250")
        self.ui.lineEdit_DR4.setText("450")
        self.ui.lineEdit_DR5.setText("650")
        self.ui.lineEdit_DR6.clear()
        self.ui.lineEdit_DR8.clear()
        self.ui.lineEdit_DR9.clear()
        self.ui.lineEdit_DR10.clear()

        self.ui.lineEdit_LV1_2.setText(_translate("Form", "25"))
        self.ui.lineEdit_LV2_2.setText(_translate("Form", "55"))
        self.ui.lineEdit_LV3_2.setText(_translate("Form", "85"))
        self.ui.lineEdit_LV4_2.setText(_translate("Form", "115"))
        self.ui.lineEdit_LV5_2.setText(_translate("Form", "135"))
        self.ui.lineEdit_LV6_2.setText(_translate("Form", "180"))
        self.ui.lineEdit_LV7_2.setText(_translate("Form", ""))
        self.ui.lineEdit_LV8_2.setText(_translate("Form", ""))
        self.ui.lineEdit_LV9_2.setText(_translate("Form", ""))
        self.ui.lineEdit_LV10_2.setText(_translate("Form", ""))

        self.ui.lineEdit_DR1_2.setText(_translate("Form", "50"))
        self.ui.lineEdit_DR2_2.setText(_translate("Form", "250"))
        self.ui.lineEdit_DR3_2.setText(_translate("Form", "450"))
        self.ui.lineEdit_DR4_2.setText(_translate("Form", "650"))
        self.ui.lineEdit_DR5_2.setText(_translate("Form", "1000"))
        self.ui.lineEdit_DR6_2.setText(_translate("Form", ""))
        self.ui.lineEdit_DR7_2.setText(_translate("Form", ""))
        self.ui.lineEdit_DR8_2.setText(_translate("Form", ""))
        self.ui.lineEdit_DR9_2.setText(_translate("Form", ""))
        self.ui.lineEdit_DR10_2.setText(_translate("Form", ""))

    # signal

    def do_category(self):
        error = False

        lineEdit_LV1_2 = self.ui.lineEdit_LV1_2.text()
        lineEdit_LV2_2 = self.ui.lineEdit_LV2_2.text()
        lineEdit_LV3_2 = self.ui.lineEdit_LV3_2.text()
        lineEdit_LV4_2 = self.ui.lineEdit_LV4_2.text()
        lineEdit_LV5_2 = self.ui.lineEdit_LV5_2.text()
        lineEdit_LV6_2 = self.ui.lineEdit_LV6_2.text()
        lineEdit_LV7_2 = self.ui.lineEdit_LV7_2.text()
        lineEdit_LV8_2 = self.ui.lineEdit_LV8_2.text()
        lineEdit_LV9_2 = self.ui.lineEdit_LV9_2.text()
        lineEdit_LV10_2 = self.ui.lineEdit_LV10_2.text()

        lineEdit_DR1_2 = self.ui.lineEdit_DR1_2.text()
        lineEdit_DR2_2 = self.ui.lineEdit_DR2_2.text()
        lineEdit_DR3_2 = self.ui.lineEdit_DR3_2.text()
        lineEdit_DR4_2 = self.ui.lineEdit_DR4_2.text()
        lineEdit_DR5_2 = self.ui.lineEdit_DR5_2.text()
        lineEdit_DR6_2 = self.ui.lineEdit_DR6_2.text()
        lineEdit_DR7_2 = self.ui.lineEdit_DR7_2.text()
        lineEdit_DR8_2 = self.ui.lineEdit_DR8_2.text()
        lineEdit_DR9_2 = self.ui.lineEdit_DR9_2.text()
        lineEdit_DR10_2 = self.ui.lineEdit_DR10_2.text()

        lv_vector_temp = [lineEdit_LV1_2, lineEdit_LV2_2, lineEdit_LV3_2, lineEdit_LV4_2, lineEdit_LV5_2,
                          lineEdit_LV6_2, lineEdit_LV7_2, lineEdit_LV8_2, lineEdit_LV9_2, lineEdit_LV10_2]
        dr_vector_temp = [lineEdit_DR1_2, lineEdit_DR2_2, lineEdit_DR3_2, lineEdit_DR4_2, lineEdit_DR5_2,
                          lineEdit_DR6_2, lineEdit_DR7_2, lineEdit_DR8_2, lineEdit_DR9_2, lineEdit_DR10_2]

        # array with values
        lv_vector = []
        dr_vector = []
        flag = 0
        for lv in lv_vector_temp:
            if (flag == 1):
                if (lv != ""):  # once appear "", later ones all should be ""
                    error = True
                    break
                else:
                    lv_vector.append("")
            else:
                if (lv == ""):
                    flag = 1
                    lv_vector.append("")
                else:
                    lv = int(lv)
                    lv_vector.append(lv)

        flag = 0
        for dr in dr_vector_temp:
            if (flag == 1):
                if (dr != ""):
                    error = True
                    break
                else:
                    dr_vector.append("")
            else:
                if (dr == ""):
                    flag = 1
                    dr_vector.append("")
                else:
                    dr = int(dr)
                    dr_vector.append(dr)

        # check values are in assending order and in the range
        i = 0
        for lv in lv_vector:
            if (lv != ""):
                if (i == 0):
                    if (lv >= 1):
                        temp = lv
                        i += 1
                    else:
                        error = True
                        break
                else:
                    if (lv < temp):
                        error = True
                        break
                    else:
                        if (lv > 180):
                            error = True
                            break
                        else:
                            temp = lv
                            i += 1

        i = 0
        for dr in dr_vector:
            if (dr != ""):
                if (i == 0):
                    if (dr >= 1):
                        temp = dr
                        i += 1
                    else:
                        error = True
                        break
                else:
                    if (dr < temp):
                        error = True
                        break
                    else:
                        if (dr > 1000):
                            error = True
                            break
                        else:
                            temp = dr
                            i += 1

        # error procession
        if (error):
            # show the error message
            QMessageBox.about(
                self,  "ERROR", "Range settings are wrong please check.")
        else:
            # # set new data
            # # LV new data
            # self.ui.lineEdit_LV2.setText(str(lv_vector[0]))
            # self.ui.lineEdit_LV3.setText(str(lv_vector[1]))
            # self.ui.lineEdit_LV4.setText(str(lv_vector[2]))
            # self.ui.lineEdit_LV5.setText(str(lv_vector[3]))
            # self.ui.lineEdit_LV6.setText(str(lv_vector[4]))
            # self.ui.lineEdit_LV7.setText(str(lv_vector[5]))
            # self.ui.lineEdit_LV8.setText(str(lv_vector[6]))
            # self.ui.lineEdit_LV9.setText(str(lv_vector[7]))
            # self.ui.lineEdit_LV10.setText(str(lv_vector[8]))
            # if (str(lv_vector[0]) != ""):
            #     if (lv_vector[0] == 180):
            #         self.ui.lineEdit_LV2.setText("")

            # if (str(lv_vector[1]) != ""):
            #     if (lv_vector[1] == 180):
            #         self.ui.lineEdit_LV3.setText("")

            # if (str(lv_vector[2]) != ""):
            #     if (lv_vector[2] == 180):
            #         self.ui.lineEdit_LV4.setText("")

            # if (str(lv_vector[3]) != ""):
            #     if (lv_vector[3] == 180):
            #         self.ui.lineEdit_LV5.setText("")

            # if (str(lv_vector[4]) != ""):
            #     if (lv_vector[4] == 180):
            #         self.ui.lineEdit_LV6.setText("")

            # if (str(lv_vector[5]) != ""):
            #     if (lv_vector[5] == 180):
            #         self.ui.lineEdit_LV7.setText("")

            # if (str(lv_vector[6]) != ""):
            #     if (lv_vector[6] == 180):
            #         self.ui.lineEdit_LV8.setText("")

            # if (str(lv_vector[7]) != ""):
            #     if (lv_vector[7] == 180):
            #         self.ui.lineEdit_LV9.setText("")

            # if (str(lv_vector[8]) != ""):
            #     if (lv_vector[8] == 180):
            #         self.ui.lineEdit_LV10.setText("")

            # # DR new data
            # self.ui.lineEdit_DR2.setText(str(dr_vector[0]))
            # self.ui.lineEdit_DR3.setText(str(dr_vector[1]))
            # self.ui.lineEdit_DR4.setText(str(dr_vector[2]))
            # self.ui.lineEdit_DR5.setText(str(dr_vector[3]))
            # self.ui.lineEdit_DR6.setText(str(dr_vector[4]))
            # self.ui.lineEdit_DR7.setText(str(dr_vector[5]))
            # self.ui.lineEdit_DR8.setText(str(dr_vector[6]))
            # self.ui.lineEdit_DR9.setText(str(dr_vector[7]))
            # self.ui.lineEdit_DR10.setText(str(dr_vector[8]))
            # if (str(dr_vector[0]) != ""):
            #     if (dr_vector[0] == 1000):
            #         self.ui.lineEdit_DR2.setText("")

            # if (str(dr_vector[1]) != ""):
            #     if (dr_vector[1] == 1000):
            #         self.ui.lineEdit_DR3.setText("")

            # if (str(dr_vector[2]) != ""):
            #     if (dr_vector[2] == 1000):
            #         self.ui.lineEdit_DR4.setText("")

            # if (str(dr_vector[3]) != ""):
            #     if (dr_vector[3] == 1000):
            #         self.ui.lineEdit_DR5.setText("")

            # if (str(dr_vector[4]) != ""):
            #     if (dr_vector[4] == 1000):
            #         self.ui.lineEdit_DR6.setText("")

            # if (str(dr_vector[5]) != ""):
            #     if (dr_vector[5] == 1000):
            #         self.ui.lineEdit_DR7.setText("")

            # if (str(dr_vector[6]) != ""):
            #     if (dr_vector[6] == 1000):
            #         self.ui.lineEdit_DR8.setText("")

            # if (str(dr_vector[7]) != ""):
            #     if (dr_vector[7] == 1000):
            #         self.ui.lineEdit_DR9.setText("")

            # if (str(dr_vector[8]) != ""):
            #     if (dr_vector[8] == 1000):
            #         self.ui.lineEdit_DR10.setText("")

            ######## do the function ########
            #
            LV_region = []
            DR_region = []
            for i in range(10):
                if str(lv_vector[i]) == "":
                    for j in range(0, i):
                        LV_region.append(lv_vector[j])
                    break

            for i in range(10):
                if str(dr_vector[i]) == "":
                    for j in range(0, i):
                        DR_region.append(dr_vector[j])
                    break

            # # 发送信号给子进程，通知它重新加载参数
            # subprocess_obj.send_signal(subprocess.SIGUSR1)

            # 启动子进程
            your_path = self.ui.lineEdit_DataFolder.text()
            file_path = self.ui.lineEdit_Tonecpp.text()
            if (your_path != '' and file_path.endswith('TONE.cpp')):
                data = {
                    "your_path": your_path,
                    "file_path": file_path,
                    "LV_region": LV_region,
                    "DR_region": DR_region
                }
                self.solver_thread.data = data
                self.solver_thread.start()
                # mtk_main(your_path,file_path, LV_region, DR_region)
                # input()
                sys.exit()
            elif (your_path == ''):
                # show the error message
                QMessageBox.about(
                    self,  "ERROR", "Choose Data Folder first.")
            elif (not (file_path.endswith('TONE.cpp'))):
                # show the error message
                QMessageBox.about(
                    self,  "ERROR", "Choose an TONE.cpp file first.")
            pass

    # Win/Linux switch
    def do_openfolder(self):
        # open folder
        your_path = self.ui.lineEdit_DataFolder.text()

        # For Windows
        os.startfile(your_path)

        # For Linux
        # opener = "open" if sys.platform == "darwin" else "xdg-open"
        # subprocess.call([opener, your_path])

    def do_update(self):
        # Load Excel file
        def process_xlsm(file_path):
            workbook = load_workbook(file_path, keep_vba=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        if isinstance(cell.value, str) and "_xlfn.XLOOKUP" in cell.value:
                            cell.value = cell.value.replace(
                                "_xlfn.XLOOKUP", "XLOOKUPs")

            modified_file_path = os.path.join(os.path.dirname(
                file_path), "modified_" + os.path.basename(file_path))
            workbook.save(modified_file_path)

            print(
                f"{os.path.basename(file_path)} ----> modified_{os.path.basename(file_path)}")

        folder_path = self.ui.lineEdit_DataFolder.text()
        if (folder_path == ''):
            # show the error message
            QMessageBox.about(
                self,  "ERROR", "Choose Data Folder first.")
        else:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    if file.endswith(".xlsm"):
                        process_xlsm(os.path.join(root, file))

    # def update_status_bar(self, text):
    #     self.statusBar.showMessage(text, 8000)

    def failed(self, text="Failed"):
        self.set_all_enable(True)
        QMessageBox.about(self, "Failed", text)

    def solver_finish(self):
        self.set_all_enable(True)
        self.statusBar.hide()


if __name__ == "__main__":
    import sys
    root = tk.Tk()
    root.withdraw()
    app = QApplication(sys.argv)
    Form = MyWidget()
    Form.setFixedSize(590, 740)
    Form.show()
    sys.exit(app.exec_())
