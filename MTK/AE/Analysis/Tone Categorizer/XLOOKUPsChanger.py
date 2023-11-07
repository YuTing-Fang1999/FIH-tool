import os
from openpyxl import load_workbook

def process_xlsm(file_path):
    workbook = load_workbook(file_path, keep_vba=True)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if isinstance(cell.value, str) and "_xlfn.XLOOKUP" in cell.value:
                    cell.value = cell.value.replace("_xlfn.XLOOKUP", "XLOOKUPs")

    modified_file_path = os.path.join(os.path.dirname(file_path), "modified_" + os.path.basename(file_path))
    workbook.save(modified_file_path)

    print(f"{os.path.basename(file_path)} ----> modified_{os.path.basename(file_path)}")

folder_path = input("Computed Excel folder path : ")

for root, dirs, files in os.walk(folder_path):
    for file in files:
        if file.endswith(".xlsm"):
            process_xlsm(os.path.join(root, file))