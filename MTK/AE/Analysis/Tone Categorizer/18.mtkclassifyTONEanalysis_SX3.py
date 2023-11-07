import os
import cv2
import openpyxl
import time
import re
import tkinter as tk
import numpy as np
from tkinter import filedialog
import shutil
from pathlib import Path

def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]

def file_filter(f):
    if f[-5:] in ['.exif'] or f[-4:] in ['.txt']:
        return True
    else:
        return False

def file_filter_jpg(f):
    if f[-4:] in ['.jpg', '.JPG']:
        return True
    else:
        return False
    
def classify(LVnum,LV,DR,DR_region):
    Path(yourPath+"/"+LVnum).mkdir(parents=True, exist_ok=True)
    if DR < DR_region[0]:
        final_path = yourPath+"/"+LVnum+f"/DR1_{DR_region[0]}down"
        Path(final_path).mkdir(parents=True, exist_ok=True)
    for numDR in range(0,np.size(DR_region)-1,1):
        if DR >= DR_region[numDR] and DR < DR_region[numDR+1]:
            final_path = yourPath+"/"+LVnum+f"/DR{numDR+2}_{DR_region[numDR]}_{DR_region[numDR+1]}"
            Path(final_path).mkdir(parents=True, exist_ok=True)
    if DR > DR_region[np.size(DR_region)-1]:
        final_path = yourPath+"/"+LVnum+f"/DR{np.size(DR_region)+1}_{DR_region[np.size(DR_region)-1]}up"
        Path(final_path).mkdir(parents=True, exist_ok=True)
    return final_path

def create_xls(file_path):
    fn = 'mtkTONEanalysis_SX3.xlsm'
    wb = openpyxl.load_workbook(fn, read_only=False, keep_vba=True)
    wb.active = 0
    ws = wb.active
    
    f = open(file_path, "r")
    
    darkTBL = []
    brightTBL = []
    
    for line in f:
        if "// base strength LV" in line:
            baseStrengthLV = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "// dark strength base ratio" in line:
            DbaseRatio = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "// bright strength base ratio" in line:
            BbaseRatio = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "LVTarget" in line:
            LVTarget = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "MaxLceGain" in line:
            MaxLceGain = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "MaxFinalTarget" in line:
            MaxFinalTarget = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "Dark    Strength" in line:
            line2, line3, line4, line5, line6, line7, line8, line9, line10, line11, line12, line13 = [next(f) for _ in range(12)]
            if darkTBL == []:
                darkTBL.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                darkTBL.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                darkTBL.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                darkTBL.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                darkTBL.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                darkTBL.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                darkTBL.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                darkTBL.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                darkTBL.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
                darkTBL.append(re.sub("[^0-9-,]","", line12).split(",")[0:-1])
                pre_slash, post_slash = line13.split('//')
                darkTBL.append(re.sub("[^0-9-,]","", pre_slash).split(","))
        if "Bright    Strength" in line:
            line2, line3, line4, line5, line6, line7, line8, line9, line10, line11, line12, line13 = [next(f) for _ in range(12)]
            if brightTBL == []:
                brightTBL.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                brightTBL.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                brightTBL.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                brightTBL.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                brightTBL.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                brightTBL.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                brightTBL.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                brightTBL.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                brightTBL.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
                brightTBL.append(re.sub("[^0-9-,]","", line12).split(",")[0:-1])
                pre_slash, post_slash = line13.split('//')
                brightTBL.append(re.sub("[^0-9-,]","", pre_slash).split(","))
        
    for j, item in enumerate(baseStrengthLV, start=0):
        ws.cell(column=4+j, row=15).value = int(item)
    for j, item in enumerate(DbaseRatio, start=0):
        ws.cell(column=4+j, row=16).value = int(item)
    for j, item in enumerate(BbaseRatio, start=0):
        ws.cell(column=4+j, row=17).value = int(item)
    for j, item in enumerate(LVTarget, start=0):
        ws.cell(column=18+j, row=3).value = int(item)
    for j, item in enumerate(MaxLceGain, start=0):
        ws.cell(column=18+j, row=4).value = int(item)
    for j, item in enumerate(MaxFinalTarget, start=0):
        ws.cell(column=18+j, row=5).value = int(item)
        
    for j, item in enumerate(darkTBL, start=0):
        for i in range(0,np.size(item)):
            ws.cell(column=18+i, row=7+j).value = str(item[i] + " / " + brightTBL[j][i])
            
    sheet = wb[wb.sheetnames[0]]
    target = wb.copy_worksheet(sheet)
    target.title = "default"
    
    print("TONE.cpp is ok!")
    return wb

def classify_xls(yourPath):
    allFileList = os.listdir(yourPath)
    allFileList_exif = list(filter(file_filter, allFileList))
    allFileList_exif.sort(key=natural_keys)
    allFileList_jpg = list(filter(file_filter_jpg, allFileList))
    allFileList_jpg.sort(key=natural_keys)
    
    for i, item in enumerate(allFileList_exif, start=0):
        path_name = yourPath + "/" + item
        exifFile = open(path_name, "r")
        base = os.path.splitext(item)[0]
        num = base.split("_")[0]
        
        if i == 0:
            wb = create_xls(file_path)
            wb.active = 0
            ws = wb.active
        
        print(base)
        
        checkP50 = 0
        checkO50 = 0
        checkLV = 0
        
        ws.cell(column=2, row=6).value = file_path.split("/")[-3]
        ws.cell(column=2, row=23+(5*(i%20))).value = int(num)
        for line in exifFile:
            if "SW_LCE_P0" in line:
                ws.cell(column=15, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "SW_LCE_P1" in line:
                ws.cell(column=16, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "SW_LCE_P50" in line:
                if checkP50 == 0:
                    ws.cell(column=17, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[2:])
                    checkP50 = 1
            if "SW_LCE_P250" in line:
                ws.cell(column=18, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "SW_LCE_P500" in line:
                ws.cell(column=19, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "SW_LCE_P750" in line:
                ws.cell(column=20, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "SW_LCE_P950" in line:
                ws.cell(column=21, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "SW_LCE_P999" in line:
                ws.cell(column=22, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "SW_LCE_O0" in line:
                ws.cell(column=15, row=24+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "SW_LCE_O1" in line:
                ws.cell(column=16, row=24+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[1:])
            if "SW_LCE_O50" in line:
                if checkO50 == 0:
                    ws.cell(column=17, row=24+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[2:])
                    checkO50 = 1
            if "SW_LCE_O250" in line:
                ws.cell(column=18, row=24+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "SW_LCE_O500" in line:
                ws.cell(column=19, row=24+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "SW_LCE_O750" in line:
                ws.cell(column=20, row=24+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "SW_LCE_O950" in line:
                ws.cell(column=21, row=24+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "SW_LCE_O999" in line:
                ws.cell(column=22, row=24+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line)[3:])
                
            if "SW_LCE_LumaProb" in line:
                ws.cell(column=24, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
            if "SW_LCE_FlatProb" in line:
                ws.cell(column=25, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
            if "SW_LCE_LVProb" in line:
                ws.cell(column=44, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
                
            if "SW_LCE_LumaTarget" in line:
                ws.cell(column=28, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
            if "SW_LCE_FlatTarget" in line:
                ws.cell(column=29, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
            if "SW_LCE_LumaFlatTarget" in line:
                ws.cell(column=30, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
            if "SW_LCE_FinalTarget" in line:
                ws.cell(column=32, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
                
            if "SW_LCE_LV" in line:
                if checkLV == 0:
                    ws.cell(column=33, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
                    checkLV = 1
            if "SW_LCE_CurrDR" in line:
                ws.cell(column=36, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
                
            if "SW_LCE_FinalDStrength" in line:
                ws.cell(column=41, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
            if "SW_LCE_FinalBStrength" in line:
                ws.cell(column=42, row=23+(5*(i%20))).value = int(re.sub("[^0-9-,]","", line))
        
        for j, item in enumerate(allFileList_jpg, start=0):
            path_name_jpg = yourPath + "/" + str(item)
            base2 = os.path.splitext(str(item))[0]
            if str(item) == base or base2 == base[0:-8]:
                img = cv2.imread(path_name_jpg)
                height, width = img.shape[0], img.shape[1]
                
                save_img = openpyxl.drawing.image.Image(path_name_jpg)
                save_img.height = 166
                save_img.width = 166 * width / height
                anchor_name = "F" + str(23+(5*(i%20)))
                save_img.anchor = anchor_name
                ws.add_image(save_img)
                
                if refer == 1:
                    if j % 2 == 0:
                        path_name_jpg2 = yourPath + "/" + allFileList_jpg[j+1]
                    else:
                        path_name_jpg2 = yourPath + "/" + allFileList_jpg[j-1]
                    
                    save_img2 = openpyxl.drawing.image.Image(path_name_jpg2)
                    save_img2.height = 166
                    save_img2.width = 166 * width / height
                    anchor_name = "C" + str(23+(5*(i%20)))
                    save_img2.anchor = anchor_name
                    ws.add_image(save_img2)
                    
        if i == len(allFileList_exif)-1:
            LVregion = yourPath.split("/")[1].split("_")[0]
            DRregion = yourPath.split("/")[2].split("_")[0]
            file = "mtkTONEanalysis_SX3_" + str(localtime[0]) + "_" + str(localtime[1]) + "_" + str(localtime[2]) + "_" + clock + "_" + LVregion + "_" + DRregion + ".xlsm"
            wb.active = 0
            wb.save(file)

print("mtkclassifyTONEanalysis is runing...")

root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(filetypes=[("cpp files", "*.cpp")])
print(file_path)

refer = input("Have reference or not (0: no, 1: yes): ")
refer = int(refer)

localtime = time.localtime()
clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])

yourPath = "Exif"
allFileList = os.listdir(yourPath)
allFileList_exif = list(filter(file_filter, allFileList))
allFileList_exif.sort(key=natural_keys)
allFileList_jpg = list(filter(file_filter_jpg, allFileList))
allFileList_jpg.sort(key=natural_keys)

LV_region = [25,55,85,115,135,155]
DR_region = [50,250,450,650,850]
LV_all = []
destination_set = []
for numLV in range(0,np.size(LV_region)+1):
    locals()['LV'+str(numLV+1)] = []

for i, item in enumerate(allFileList_exif, start=0):
    path_name = yourPath + "/" + item
    with open(path_name, 'r' ) as exifFile :
        base = os.path.splitext(item)[0]
        
        checkLV = 0
        
        for line in exifFile:
            if "SW_LCE_LV" in line:
                if checkLV == 0:
                    LV = int(re.sub("[^0-9-,]","", line))
                    checkLV = 1
            if "SW_LCE_CurrDR" in line:
                DR = int(re.sub("[^0-9-,]","", line))
    
    for j, item_jpg in enumerate(allFileList_jpg, start=0):
        path_name_jpg = yourPath + "/" + item_jpg
        if base == item_jpg:
            if j % 2 == 0 and refer == 1:
                path_name_jpg2 = yourPath + "/" + allFileList_jpg[j+1]
            else:
                path_name_jpg2 = yourPath + "/" + allFileList_jpg[j-1]
            
            LV_all.append([base.split("_")[0],LV,DR])
            
            if LV < LV_region[0]:
                locals()['LV1'].append([base.split("_")[0],LV,DR])
                destination = classify(f"LV1_{LV_region[0]}down",LV,DR,DR_region)
                break
            for numLV in range(0,np.size(LV_region)-1,1):
                if LV >= LV_region[numLV] and LV < LV_region[numLV+1]:
                    locals()['LV'+str(numLV+2)].append([base.split("_")[0],LV,DR])
                    destination = classify(f"LV{numLV+2}_{LV_region[numLV]}_{LV_region[numLV+1]}",LV,DR,DR_region)
                    break
            if LV > LV_region[np.size(LV_region)-1]:
                locals()['LV'+str(np.size(LV_region)+1)].append([base.split("_")[0],LV,DR])
                destination = classify(f"LV{np.size(LV_region)+1}_{LV_region[np.size(LV_region)-1]}up",LV,DR,DR_region)
                break
            exifFile.close()
            shutil.copy(path_name,destination)
            shutil.copy(path_name_jpg,destination)
            if refer == 1:
                shutil.copy(path_name_jpg2,destination)
            destination_set.append(destination)

destination_set = [x for i, x in enumerate(destination_set) if x not in destination_set[:i]]

for i, item in enumerate(destination_set, start=0):
    classify_xls(item)

print("mtkclassifyTONEanalysis is ok!")
os.system("pause")