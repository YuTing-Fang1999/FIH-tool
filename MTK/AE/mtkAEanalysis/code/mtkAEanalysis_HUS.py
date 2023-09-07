import os
import cv2
import openpyxl
import numpy as np
import time
import matplotlib.pyplot as plt
import re
import tkinter as tk
from tkinter import filedialog
from PIL import Image
from PyQt5.QtWidgets import QWidget
from PyQt5.QtCore import pyqtSignal

import matplotlib
matplotlib.use('agg')

def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]

def file_filter(f):
    if f[-5:] in ['.exif'] or f[-4:] in ['.txt']:
        return True
    else:
        return False

def file_filter_jpg(f):
    if f[-4:] in ['.jpg', '.JPG']:
        return True
    else:
        return False

def create_xls(file_path):
    fn = 'MTK/AE/mtkAEanalysis/code/mtkAEanalysis_HUS.xlsm'
    wb = openpyxl.load_workbook(fn, read_only=False, keep_vba=True)
    wb.active = 0
    ws = wb.active
    
    f = open(file_path, "r")
    
    HS_MixW = []
    MT_THD = []
    HS_BT_BV = []
    HS_BT_EVD = []
    HS_MixW_BV = []
    HS_MixW_EVD = []
    HS_MixW_Mid = []
    HS_MT_BV = []
    EVD_B2M = []
    fd_bv = []
    fbt_bv = []
    fbt_dr = []
    flt_bv = []
    flt_dr = []
    fbt_ns_bv = []
    fbt_ns_dr = []
    flt_ns_bv = []
    flt_ns_dr = []
    fdy_min_bv = []
    fdy_min_dr = []
    fbt_fd_th_tbl = []
    face_loc_lut = []
    face_loc_wet = []
    face_bv_lut = []
    face_bv_wet = []
    
    ws.cell(column=2, row=6).value = file_path.split("/")[-5]
    for line in f:
        if "mt_midrt_hl" in line:
            MidL = re.sub("[^0-9-,]","", line).split(",")[0]
            MidH = re.sub("[^0-9-,]","", line).split(",")[1]
        if "mt_lut_bv_size" in line:
            mt_lut_bv_size = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "mt_thd_bv" in line:
            mt_thd_bv = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "mt_thd_based" in line:
            MT_base = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "mt_thd_exp" in line:
            MT_exp = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "bt_bv_sz" in line:
            HS_BT_BV.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "bt_bvrt_lut" in line:
            HS_BT_BV.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "bt_evd_sz" in line:
            HS_BT_EVD.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "bt_b2d_rt_lut" in line:
            HS_BT_EVD.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "bt_thd_lut" in line:
            HS_EV_THD = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "mixwet_bv_sz" in line:
            HS_MixW_BV.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "mixwet_bv_rt" in line:
            HS_MixW_BV.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "mixwet_evd_sz" in line:
            HS_MixW_EVD.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "mixwet_b2d_rt" in line:
            HS_MixW_EVD.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "mixwet_midrt" in line:
            HS_MixW_Mid.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "mix_wet_tbl" in line:
            HS_MixW = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "mt_bv_sz" in line:
            HS_MT_BV.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "mt_bvrt_lut" in line:
            HS_MT_BV.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "mt_evd_sz" in line:
            EVD_B2M.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "mt_b2d_rt_lut" in line:
            EVD_B2M.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "mt_thd_lut" in line:
            MT_THD = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        
        if "fd_dr_ra_x" in line:
            fd_dr_ra_x = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fd_dr_ra_y" in line:
            fd_dr_ra_y = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "flt_fdsz_ra_x" in line:
            flt_fdsz_ra_x = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "flt_fdsz_ra_y" in line:
            flt_fdsz_ra_y = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fbt_bv" in line:
            fbt_bv.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "fbt_dr" in line:
            fbt_dr.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "flt_bv" in line:
            flt_bv.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "flt_dr" in line:
            flt_dr.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "fbt_ns_bv" in line:
            fbt_ns_bv.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "fbt_ns_dr" in line:
            fbt_ns_dr.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "flt_ns_bv" in line:
            flt_ns_bv.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "flt_ns_dr" in line:
            flt_ns_dr.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "en_fd_locsz_bv" in line:
            en_fd_locsz_bv = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fd_sz_smal2bg" in line:
            ws.cell(column=59, row=17).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
            ws.cell(column=60, row=17).value = int(re.sub("[^0-9-,]","", line).split(",")[1])
        if "fd_pb_smal2bg" in line:
            ws.cell(column=59, row=18).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
            ws.cell(column=60, row=18).value = int(re.sub("[^0-9-,]","", line).split(",")[1])
        if "fd_loc_near2far" in line:
            ws.cell(column=59, row=19).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
            ws.cell(column=60, row=19).value = int(re.sub("[^0-9-,]","", line).split(",")[1])
        if "fd_pb_near2far" in line:
            ws.cell(column=59, row=20).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
            ws.cell(column=60, row=20).value = int(re.sub("[^0-9-,]","", line).split(",")[1])
        if "fd_tbl_sz" in line:
            fd_bv.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "fd_bv" in line:
            fd_bv.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "fd_pb_smalsz" in line:
            fd_pb_smalsz = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fd_pb_bgsz" in line:
            fd_pb_bgsz = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fd_pb_locnear" in line:
            fd_pb_locnear = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fd_pb_locfar" in line:
            fd_pb_locfar = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fdy_min_bv" in line:
            fdy_min_bv.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "fdy_min_dr" in line:
            fdy_min_dr.append(re.sub("[^0-9-,]","", line).split(",")[0:-1])
        if "fbt_fd_th_tbl" in line:
            fbt_fd_th_tbl = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "flt_fd_th_tbl" in line:
            flt_fd_th_tbl = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fbt_ns_fd_th_tbl" in line:
            fbt_ns_fd_th_tbl = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "flt_ns_fd_th_tbl" in line:
            flt_ns_fd_th_tbl = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fdy_min_th" in line:
            fdy_min_th = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        
        if "ns_bv_cfg_tbll" in line:
            ns_bv_cfg_tbll = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "ns_bv_cfg_bv" in line:
            ns_bv_cfg_bv = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "ns_bv_cfg_btthd" in line:
            ns_bv_cfg_btthd = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "ns_bv_cfg_nsthd" in line:
            ns_bv_cfg_nsthd = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        
        if "ns_flat_rt_x1" in line:
            ws.cell(column=70, row=23).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
        if "ns_flat_rt_y1" in line:
            ws.cell(column=72, row=23).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
        if "ns_flat_rt_x2" in line:
            ws.cell(column=70, row=24).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
        if "ns_flat_rt_y2" in line:
            ws.cell(column=72, row=24).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
            
        if "face_loc_sz" in line:
            face_loc_sz = re.sub("[^0-9-,]","", line).split(",")[0]
        if "face_loc_lut" in line:
            face_loc_lut = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "face_loc_wet" in line:
            face_loc_wet = re.sub("[^0-9-,]","", line).split(",")[0:-1]
            
        if "face_bv_lut_sz" in line:
            face_bv_lut_sz = re.sub("[^0-9-,]","", line).split(",")[0]
        if "face_bv_lut" in line:
            face_bv_lut = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "face_bv_wet" in line:
            face_bv_wet = re.sub("[^0-9-,]","", line).split(",")[0:-1]
    
    ws.cell(column=3, row=16).value = int(MidL)
    ws.cell(column=4, row=16).value = int(MidH)
    
    for j in range(0,int(np.array(mt_lut_bv_size[0]))):
        ws.cell(column=j+3, row=17).value = int(mt_thd_bv[j])
        ws.cell(column=j+3, row=18).value = int(MT_base[j])
        ws.cell(column=j+3, row=19).value = int(MT_exp[j])
        
    for j in range(0,int(np.array(HS_BT_BV[0]))):
        ws.cell(column=j+13, row=17).value = int(HS_BT_BV[1][j])
        for i in range(0,int(np.array(HS_BT_EVD[0]))):
            ws.cell(column=i+13, row=j+22).value = int(HS_EV_THD[j*int(np.array(HS_BT_EVD[0]))+i])
    for j in range(0,int(np.array(HS_BT_EVD[0]))):
        ws.cell(column=j+13, row=18).value = int(HS_BT_EVD[1][j])
        
    for j in range(0,int(np.array(HS_MT_BV[0]))):
        ws.cell(column=j+13, row=37).value = int(HS_MT_BV[1][j])
        for i in range(0,int(np.array(EVD_B2M[0]))):
            ws.cell(column=i+13, row=j+41).value = int(MT_THD[j*int(np.array(EVD_B2M[0]))+i])
    for j in range(0,int(np.array(EVD_B2M[0]))):
        ws.cell(column=j+13, row=38).value = int(EVD_B2M[1][j])
    
    for j in range(0,int(np.array(HS_MixW_BV[0]))):
        if j <= 1:
            ws.cell(column=23+j, row=17).value = int(HS_MixW_BV[1][j])
        else:
            ws.cell(column=23+j*2, row=17).value = int(HS_MixW_BV[1][j])
    
    for j in range(0,int(np.array(HS_MixW_EVD[0]))):
        if j <= 1:
            ws.cell(column=j+23, row=18).value = int(HS_MixW_EVD[1][j])
        else:
            ws.cell(column=27+(j-2)*3, row=18).value = int(HS_MixW_EVD[1][j])
    
    for j in range(0,int(np.array(HS_MixW_Mid[0]))):
        if j <= 1:
            ws.cell(column=j+23, row=19).value = int(HS_MixW_Mid[1][j])
        else:
            ws.cell(column=27+(j-2)*3, row=19).value = int(HS_MixW_Mid[1][j])
    
    one = int(np.array(HS_MixW_EVD[0]))*int(np.array(HS_MixW_Mid[0]))*4
    two = int(np.array(HS_MixW_Mid[0]))*4
    for z in range(0,int(np.array(HS_MixW_BV[0]))):
        for j in range(0,int(np.array(HS_MixW_EVD[0]))):
            for i in range(0,int(np.array(HS_MixW_Mid[0]))):
                for y in range(0,4):
                    ws.cell(column=(i*4)+y+24, row=(z*5)+j+23).value = int(HS_MixW[z*one+j*two+i*4+y])
    
    for j in range(0,np.size(fd_dr_ra_x)):
        if j < 9:
            ws.cell(column=j+47, row=16).value = int(fd_dr_ra_x[j])
            ws.cell(column=j+47, row=17).value = int(fd_dr_ra_y[j])
        else:
            ws.cell(column=(j-9)+47, row=18).value = int(fd_dr_ra_x[j])
            ws.cell(column=(j-9)+47, row=19).value = int(fd_dr_ra_y[j])
    
    for j in range(0,np.size(flt_fdsz_ra_x)):
        if j < 9:
            ws.cell(column=j+47, row=22).value = int(flt_fdsz_ra_x[j])
            ws.cell(column=j+47, row=23).value = int(flt_fdsz_ra_y[j])
        else:
            ws.cell(column=(j-9)+47, row=24).value = int(flt_fdsz_ra_x[j])
            ws.cell(column=(j-9)+47, row=25).value = int(flt_fdsz_ra_y[j])
            
    for j in range(0,int(np.array(flt_bv[0]))):
        ws.cell(column=j+46, row=29).value = int(flt_bv[1][j])
    for j in range(0,int(np.array(fbt_bv[0]))):
        ws.cell(column=j+58, row=23).value = int(fbt_bv[1][j])
    for j in range(0,int(np.array(flt_dr[0]))):
        ws.cell(column=j+46, row=30).value = int(flt_dr[1][j])
    for j in range(0,int(np.array(fbt_dr[0]))):
        ws.cell(column=j+58, row=24).value = int(fbt_dr[1][j])
    for j in range(0,int(np.array(flt_ns_bv[0]))):
        ws.cell(column=j+46, row=47).value = int(flt_ns_bv[1][j])
    for j in range(0,int(np.array(fbt_ns_bv[0]))):
        ws.cell(column=j+58, row=41).value = int(fbt_ns_bv[1][j])
    for j in range(0,int(np.array(flt_ns_dr[0]))):
        ws.cell(column=j+46, row=48).value = int(flt_ns_dr[1][j])
    for j in range(0,int(np.array(fbt_ns_dr[0]))):
        ws.cell(column=j+58, row=42).value = int(fbt_ns_dr[1][j])
    for j in range(0,int(np.array(fdy_min_bv[0]))):
        ws.cell(column=j+46, row=64).value = int(fdy_min_bv[1][j])
    for j in range(0,int(np.array(fdy_min_dr[0]))):
        ws.cell(column=j+46, row=65).value = int(fdy_min_dr[1][j])
        
    ws.cell(column=60, row=16).value = int(en_fd_locsz_bv[0])
    
    if int(np.array(fd_bv[0])) > 6:
        for j in range(0,6):
            ws.cell(column=j+62, row=16).value = int(fd_bv[1][j])
            ws.cell(column=j+62, row=17).value = int(fd_pb_smalsz[j])
            ws.cell(column=j+62, row=18).value = int(fd_pb_bgsz[j])
            ws.cell(column=j+62, row=19).value = int(fd_pb_locnear[j])
            ws.cell(column=j+62, row=20).value = int(fd_pb_locfar[j])
    else:
        for j in range(0,int(np.array(fd_bv[0]))):
            ws.cell(column=j+62, row=16).value = int(fd_bv[1][j])
            ws.cell(column=j+62, row=17).value = int(fd_pb_smalsz[j])
            ws.cell(column=j+62, row=18).value = int(fd_pb_bgsz[j])
            ws.cell(column=j+62, row=19).value = int(fd_pb_locnear[j])
            ws.cell(column=j+62, row=20).value = int(fd_pb_locfar[j])
            
    for j in range(0,int(np.array(fbt_bv[0]))):
        for i in range(0,int(np.array(fbt_dr[0]))):
            ws.cell(column=i+58, row=j+28).value = int(fbt_fd_th_tbl[j*int(np.array(fbt_dr[0]))+i])
            
    for j in range(0,int(np.array(fbt_ns_bv[0]))):
        for i in range(0,int(np.array(fbt_ns_dr[0]))):
            ws.cell(column=i+58, row=j+45).value = int(fbt_ns_fd_th_tbl[j*int(np.array(fbt_ns_dr[0]))+i])
            
    for j in range(0,int(np.array(flt_bv[0]))):
        for i in range(0,int(np.array(flt_dr[0]))):
            ws.cell(column=i+46, row=j+34).value = int(flt_fd_th_tbl[j*int(np.array(flt_dr[0]))+i])
            
    for j in range(0,int(np.array(flt_ns_bv[0]))):
        for i in range(0,int(np.array(flt_ns_dr[0]))):
            ws.cell(column=i+46, row=j+51).value = int(flt_ns_fd_th_tbl[j*int(np.array(flt_ns_dr[0]))+i])
            
    for j in range(0,int(np.array(fdy_min_bv[0]))):
        for i in range(0,int(np.array(fdy_min_dr[0]))):
            ws.cell(column=i+46, row=j+68).value = int(fdy_min_th[j*int(np.array(fdy_min_dr[0]))+i])
    
    for j in range(0, int(np.array(ns_bv_cfg_tbll[0]))):
        ws.cell(column=j+70, row=18).value = int(ns_bv_cfg_bv[j])
        ws.cell(column=j+70, row=19).value = int(ns_bv_cfg_btthd[j])
        ws.cell(column=j+70, row=20).value = int(ns_bv_cfg_nsthd[j])
        
    for j in range(0,int(np.array(face_loc_sz[0]))):
        ws.cell(column=j+70, row=34).value = int(face_loc_lut[j])
        ws.cell(column=j+70, row=35).value = int(face_loc_wet[j])
        
    for j in range(0,int(np.array(face_bv_lut_sz[0]))):
        ws.cell(column=j+70, row=36).value = int(face_bv_lut[j])
        ws.cell(column=j+70, row=37).value = int(face_bv_wet[j])
    
    print("AE.cpp is ok!")
    return wb

##############################
class HUS(QWidget):
    update_progress_bar_signal = pyqtSignal(int)
    set_progress_bar_value_signal = pyqtSignal(int)
    def __init__(self) -> None:
        super().__init__()
    def run(self, exif_path, code_path):
##############################
        print("mtkAEanalysis is runing...")

        # root = tk.Tk()
        # root.withdraw()
        # code_path = filedialog.askopenfilename()
        # print(code_path)

        # refer = input("Have reference or not (0: no, 1: yes): ")
        # refer = int(refer)

        localtime = time.localtime()
        clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])

        # exif_path = "Exif"
        allFileList = os.listdir(exif_path)
        allFileList_exif = np.sort(allFileList,axis=0)
        allFileList_exif = list(filter(file_filter, allFileList_exif))
        allFileList_exif.sort(key=natural_keys)
        allFileList_jpg = np.sort(allFileList,axis=0)
        allFileList_jpg = list(filter(file_filter_jpg, allFileList_jpg))
        allFileList_jpg.sort(key=natural_keys)

        ##############################
        self.set_progress_bar_value_signal.emit(np.size(allFileList_exif))
        ##############################
        for i in range(0,(np.size(allFileList_exif))):
            ##############################
            self.update_progress_bar_signal.emit(i+1) 
            ##############################
            path_name = exif_path + "/" + allFileList_exif[i]
            exifFile = open(path_name, "r")
            file_name = os.path.basename(path_name)
            base = os.path.splitext(file_name)[0]
            baseTag = base.split(".")[0]
            
            if i % 10 == 0:
                startNum = base.split("_")[0]
                wb = create_xls(code_path)
            
            sheet = wb[wb.sheetnames[0]]
            target = wb.copy_worksheet(sheet)
            target.title = baseTag
            wb.active = int((i%10)+2)
            ws = wb.active
            
            print(base)
            
            AE_TAG_STABLE = []
            AE_TAG_FACE_STS_SIZE = []
            AE_TAG_CWV = 0
            AE_TAG_NS_NV_LOWBNDTHD = []
            
            for line in exifFile:
                if "AE_TAG_CWV" in line:
                    if AE_TAG_CWV == 0:
                        ws.cell(column=5, row=11).value = int(re.sub("[^0-9-,]","", line))
                        AE_TAG_CWV = 1
                if "AE_TAG_REALBVX1000" in line:
                    ws.cell(column=5, row=12).value = int(re.sub("[^0-9-,]","", line)[4:])
                if "AE_TAG_MTV6_MAINT_MID_INTRATIO" in line:
                    ws.cell(column=5, row=13).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_MTV6_MAINT_Y" in line:
                    ws.cell(column=5, row=14).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_DRV6_B2D" in line:
                    ws.cell(column=15, row=13).value = int(re.sub("[^0-9-,]","", line)[2:])
                if "AE_TAG_HSV6_BT_Final_Y" in line:
                    AE_TAG_HSV6_BT_Final_Y = int(re.sub("[^0-9-,]","", line)[1:])
                    ws.cell(column=15, row=14).value = AE_TAG_HSV6_BT_Final_Y
                if "AE_TAG_HSV6_BT_THD" in line:
                    AE_TAG_HSV6_BT_THD = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_DRV6_CORR_B2M" in line:
                    ws.cell(column=15, row=33).value = int(re.sub("[^0-9-,]","", line)[2:])
                if "AE_TAG_HSV6_MT_Final_Y" in line:
                    AE_TAG_HSV6_MT_Final_Y = int(re.sub("[^0-9-,]","", line)[1:])
                    ws.cell(column=15, row=34).value = AE_TAG_HSV6_MT_Final_Y
                if "AE_TAG_HSV6_MT_THD" in line:
                    AE_TAG_HSV6_MT_THD = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_DRV6_B2D" in line:
                    ws.cell(column=27, row=13).value = int(re.sub("[^0-9-,]","", line)[2:])
                if "AE_TAG_DRV6_MIDRATIO" in line:
                    ws.cell(column=27, row=14).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_HSV6_DT_FINAL_Y" in line:
                    AE_TAG_HSV6_DT_FINAL_Y = int(re.sub("[^0-9-,]","", line)[1:])
                    ws.cell(column=15, row=39).value = AE_TAG_HSV6_DT_FINAL_Y
                if "AE_TAG_HSV6_DT_THD" in line:
                    AE_TAG_HSV6_DT_THD = int(re.sub("[^0-9-,]","", line)[1:])
                    ws.cell(column=19, row=39).value = AE_TAG_HSV6_DT_THD
                    
                if "AE_TAG_AE_TARGET" in line:
                    ws.cell(column=46, row=3).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_STABLE" in line:
                    AE_TAG_STABLE.append(re.sub("[^0-9-,]","", line))
                
                if "AE_TAG_PROB_FACE_HBND" in line:
                    ws.cell(column=48, row=5).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_FACE_LBND" in line:
                    ws.cell(column=48, row=6).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_FACE_FLT_DR" in line:
                    ws.cell(column=48, row=8).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_FACE_FLT_HBND" in line:
                    ws.cell(column=48, row=9).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                
                if "AE_TAG_FACE_LOW_BOUND" in line:
                    ws.cell(column=52, row=3).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FACE_HIGH_BOUND" in line:
                    ws.cell(column=52, row=4).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_LINK_FACE_TH_MIN" in line:
                    ws.cell(column=52, row=7).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_LINK_AE_CWR_STABLE" in line:
                    ws.cell(column=55, row=6).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_LINK_MAX_GAIN" in line:
                    ws.cell(column=55, row=7).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_LINK_FACE_CWR_STABLE" in line:
                    ws.cell(column=55, row=9).value = int(re.sub("[^0-9-,]","", line))
                    
                if "AE_TAG_FLT_FDY" in line:
                    ws.cell(column=48, row=14).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FLT_OE_SYS" in line:
                    ws.cell(column=51, row=11).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FLT_OETH" in line:
                    ws.cell(column=51, row=12).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FLT_FDDR" in line:
                    ws.cell(column=53, row=12).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FACE_STS_SIZE" in line:
                    AE_TAG_FACE_STS_SIZE.append(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FBT_FDY" in line:
                    ws.cell(column=60, row=14).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FBT_OE_SYS" in line:
                    ws.cell(column=63, row=11).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FBT_OETH" in line:
                    ws.cell(column=63, row=12).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FACE_STS_LOC" in line:
                    ws.cell(column=65, row=12).value = int(re.sub("[^0-9-,]","", line))
                    
                if "AE_TAG_NS_STS_BVPROB" in line:
                    ws.cell(column=74, row=3).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_STS_CDFPROB" in line:
                    ws.cell(column=74, row=4).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_PROB_NS_THD" in line:
                    ws.cell(column=74, row=6).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_BT_FLAT" in line:
                    ws.cell(column=71, row=7).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_BT_THD" in line:
                    ws.cell(column=74, row=7).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_DT_THD" in line:
                    ws.cell(column=74, row=8).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_DT_LBND" in line:
                    ws.cell(column=71, row=9).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_DT_HBND" in line:
                    ws.cell(column=74, row=9).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                    
                if "AE_TAG_NS_EVD" in line:
                    ws.cell(column=72, row=13).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_FLATTHD" in line:
                    ws.cell(column=72, row=14).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_NORMAL_Y" in line:
                    ws.cell(column=76, row=13).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_BT_Y" in line:
                    ws.cell(column=80, row=13).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_DT_Y" in line:
                    ws.cell(column=84, row=13).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_DT_THD" in line:
                    ws.cell(column=82, row=12).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_DT_LIMIT" in line:
                    ws.cell(column=84, row=12).value = int(re.sub("[^0-9-,]","", line))
                
                if "AE_TAG_NS_SKYENABLE" in line:
                    ws.cell(column=77, row=22).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_SKYBV_X1" in line:
                    ws.cell(column=74, row=23).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_NS_SKYBV_Y1" in line:
                    ws.cell(column=77, row=23).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_NS_SKYBV_X2" in line:
                    ws.cell(column=74, row=24).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_NS_SKYBV_Y2" in line:
                    ws.cell(column=77, row=24).value = int(re.sub("[^0-9-,]","", line)[1:])
                
            ws.cell(column=48, row=3).value = int(AE_TAG_STABLE[2])
            ws.cell(column=53, row=13).value = int(AE_TAG_FACE_STS_SIZE[0])
            
            for j in range(0,(np.size(allFileList_jpg))):
                path_name_jpg = exif_path + "/" + allFileList_jpg[j]
                file_name_jpg = os.path.basename(path_name_jpg)
                base2 = os.path.splitext(file_name_jpg)[0]
                if file_name_jpg == base or base2 == base[0:-8]:
                    img = cv2.imdecode( np.fromfile( file = path_name_jpg, dtype = np.uint8 ), cv2.IMREAD_COLOR )
                    height, width = img.shape[0], img.shape[1]
                    
                    save_img = openpyxl.drawing.image.Image(path_name_jpg)
                    if height > width and save_img.height < save_img.width:
                        rotate_name = exif_path + "/" + os.path.splitext(file_name_jpg)[0] + "_rotate.png"
                        img_rotate = Image.open(path_name_jpg)
                        img_rotate = img_rotate.rotate(90, expand = True)
                        img_rotate.save(rotate_name)
                        save_img = openpyxl.drawing.image.Image(rotate_name)
                        save_img.height = 176
                        save_img.width = 176 * width / height 
                        save_img.anchor = 'Y2'
                        ws.add_image(save_img)
                    elif height > width and save_img.height > save_img.width:
                        save_img.height = 176
                        save_img.width = 176 * width / height
                        save_img.anchor = 'Y2'
                        ws.add_image(save_img)
                    elif height < width and save_img.height > save_img.width:
                        rotate_name = exif_path + "/" + os.path.splitext(file_name_jpg)[0] + "_rotate.png"
                        img_rotate = Image.open(path_name_jpg)
                        img_rotate = img_rotate.rotate(90, expand = True)
                        img_rotate.save(rotate_name)
                        save_img = openpyxl.drawing.image.Image(rotate_name)
                        save_img.height = 176
                        save_img.width = 176 * width / height 
                        save_img.anchor = 'Y2'
                        ws.add_image(save_img)
                    else:
                        save_img.height = 176
                        save_img.width = 176 * width / height
                        save_img.anchor = 'Y2'
                        ws.add_image(save_img)
                        
                    refer = 0
                    if j+1<np.size(allFileList_jpg) and allFileList_jpg[j+1].split("/")[-1].split("_")[0] == path_name.split("/")[-1].split("_")[0]: 
                        path_name_jpg2 = exif_path + "/" + allFileList_jpg[j+1]
                        refer = 1
                    if j-1>=0 and allFileList_jpg[j-1].split("/")[-1].split("_")[0] == path_name.split("/")[-1].split("_")[0]: 
                        path_name_jpg2 = exif_path + "/" + allFileList_jpg[j-1]
                        refer = 1
                    
                    # Have reference
                    if refer == 1:
                        if j % 2 == 0:
                            path_name_jpg2 = exif_path + "/" + allFileList_jpg[j+1]
                        else:
                            path_name_jpg2 = exif_path + "/" + allFileList_jpg[j-1]
                        file_name_jpg2 = os.path.basename(path_name_jpg2)
                        img2 = cv2.imdecode( np.fromfile( file = path_name_jpg2, dtype = np.uint8 ), cv2.IMREAD_COLOR )
                        height2, width2 = img2.shape[0], img2.shape[1]
                        
                        save_img2 = openpyxl.drawing.image.Image(path_name_jpg2)
                        if height > width and save_img2.height < save_img2.width:
                            rotate_name = exif_path + "/" + os.path.splitext(file_name_jpg2)[0] + "_rotate.png"
                            img2_rotate = Image.open(path_name_jpg2)
                            img2_rotate = img2_rotate.rotate(270, expand = True)
                            img2_rotate.save(rotate_name)
                            save_img2 = openpyxl.drawing.image.Image(rotate_name)
                            save_img2.height = 176
                            save_img2.width = 176 * width2 / height2 
                            save_img2.anchor = 'AD2'
                            ws.add_image(save_img2)
                        elif height > width and save_img2.height > save_img2.width:
                            save_img2.height = 176
                            save_img2.width = 176 * width2 / height2
                            save_img2.anchor = 'AD2'
                            ws.add_image(save_img2)
                        elif height < width and save_img2.height > save_img2.width:
                            rotate_name = exif_path + "/" + os.path.splitext(file_name_jpg2)[0] + "_rotate.png"
                            img2_rotate = Image.open(path_name_jpg2)
                            img2_rotate = img2_rotate.rotate(270, expand = True)
                            img2_rotate.save(rotate_name)
                            save_img2 = openpyxl.drawing.image.Image(rotate_name)
                            save_img2.height = 176
                            save_img2.width = 176 * width2 / height2 
                            save_img2.anchor = 'AH2'
                            ws.add_image(save_img2)
                        else:
                            save_img2.height = 176
                            save_img2.width = 176 * width2 / height2
                            save_img2.anchor = 'AH2'
                            ws.add_image(save_img2)
                            
                        gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)
                        plt.figure(figsize=(2.88,2.2))
                        plt.hist(gray2.ravel(), 255, [0, 255], color='powderblue',alpha=0.75)
                        plt.yticks(alpha=0)
                        # plt.xlim(0,255)
                        # plt.ylim(0,350000)
                        save_name2 = exif_path + "/" + os.path.splitext(file_name_jpg2)[0] + "_histogram.png"
                        plt.savefig(save_name2,dpi=100)
                        plt.close()
                        save_img_hist2 = openpyxl.drawing.image.Image(save_name2)
                        save_img_hist2.anchor = 'F24'
                        ws.add_image(save_img_hist2)
                    
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    plt.figure(figsize=(2.88,2.2))
                    plt.hist(gray.ravel(), 255, [0, 255], color='powderblue',alpha=0.75)
                    plt.axvline(AE_TAG_HSV6_BT_Final_Y/4096*255, color='r',linestyle=':')
                    plt.axvline(AE_TAG_HSV6_BT_THD/4096*255, color='r')
                    plt.axvline(AE_TAG_HSV6_MT_Final_Y/4096*255, color='b',linestyle=':')
                    plt.axvline(AE_TAG_HSV6_MT_THD/4096*255, color='b')
                    plt.axvline(AE_TAG_HSV6_DT_FINAL_Y/4096*255, color='g',linestyle=':')
                    plt.axvline(AE_TAG_HSV6_DT_THD/4096*255, color='g')
                    plt.yticks(alpha=0)
                    # plt.xlim(0,255)
                    # plt.ylim(0,350000)
                    save_name1 = exif_path + "/" + os.path.splitext(file_name_jpg)[0] + "_histogram.png"
                    plt.savefig(save_name1,dpi=100)
                    plt.close()
                    save_img_hist = openpyxl.drawing.image.Image(save_name1)
                    save_img_hist.anchor = 'B24'
                    ws.add_image(save_img_hist)
                    
                    faceCase = "faceCase.png"
                    if os.path.exists(faceCase):
                        face_img = openpyxl.drawing.image.Image(faceCase)
                        face_img_2 = cv2.imdecode( np.fromfile( file = faceCase, dtype = np.uint8 ), cv2.IMREAD_COLOR )
                        height3, width3 = face_img_2.shape[0], face_img_2.shape[1]
                        face_img.height = 176
                        face_img.width = 176 * width3 / height3
                        face_img.anchor = 'BE2'
                        ws.add_image(face_img)
                    else:
                        print("No faceCase.png in path.")
                    
            if i % 20 == 19 or i == len(allFileList_exif)-1:
                endNum = base.split("_")[0]
                file = "mtkAEanalysis_HUS_" + str(localtime[0]) + "_" + str(localtime[1]) + "_" + str(localtime[2]) + "_" + clock + "_" + startNum + "_" + endNum + ".xlsm"
                wb.active = 0
                wb.save(file)
            break

        print("mtkAEanalysis is ok!")
        # os.system("pause")
        ##############################
        return file
        ##############################