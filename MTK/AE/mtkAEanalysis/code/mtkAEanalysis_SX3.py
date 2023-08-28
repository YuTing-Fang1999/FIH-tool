import os
import cv2
import openpyxl
import numpy as np
import time
import matplotlib.pyplot as plt
import re
import tkinter as tk
from tkinter import filedialog
from PIL import Image
from PyQt5.QtWidgets import QWidget
from PyQt5.QtCore import pyqtSignal

import matplotlib
matplotlib.use('agg')

def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]

def file_filter(f):
    if f[-5:] in ['.exif'] or f[-4:] in ['.txt']:
        return True
    else:
        return False

def file_filter_jpg(f):
    if f[-4:] in ['.jpg', '.JPG']:
        return True
    else:
        return False

def create_xls(file_path):
    fn = 'MTK/AE/mtkAEanalysis/code/mtkAEanalysis_SX3.xlsm'
    wb = openpyxl.load_workbook(fn, read_only=False, keep_vba=True)
    wb.active = 0
    ws = wb.active
    
    f = open(file_path, "r")
    
    HS_MixW = []
    MT_THD = []
    fbt_bv = []
    fbt_dr = []
    fbt_ns_bv = []
    fbt_ns_dr = []
    TH_tbl_1 = []
    TH_tbl_2 = []
    TH_tbl_3 = []
    TH_tbl_4 = []
    TH_tbl_5 = []
    TH_tbl_6 = []
    TH_tbl_7 = []
    TH_tbl_8 = []
    min_TH_tbl = []
    HS_EV7_THD = []
    
    ws.cell(column=2, row=6).value = file_path.split("/")[-3]
    for line in f:
        if "TargetMidRatioTbl.i4X1" in line:
            MidL = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "TargetMidRatioTbl.i4X2" in line:
            MidH = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "MainTHD_BV" in line:
            MT_BV = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "MainTHD_based" in line:
            MT_base = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "MainTHD_exp" in line:
            MT_exp = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "i4HS_BVRatio" in line:
            HS_BT_BV = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "u4HS_EVDRatio" in line:
            HS_BT_EVD = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "BV0-THD" in line:
            HS_EV0_THD = re.sub("[^0-9-,]","", next(f).strip()).split(",")[0:-1]
        if "BV1-THD" in line:
            HS_EV1_THD = re.sub("[^0-9-,]","", next(f).strip()).split(",")[0:-1]
        if "BV2-THD" in line:
            HS_EV2_THD = re.sub("[^0-9-,]","", next(f).strip()).split(",")[0:-1]
        if "BV3-THD" in line:
            HS_EV3_THD = re.sub("[^0-9-,]","", next(f).strip()).split(",")[0:-1]
        if "BV4-THD" in line:
            HS_EV4_THD = re.sub("[^0-9-,]","", next(f).strip()).split(",")[0:-1]
        if "BV5-THD" in line:
            HS_EV5_THD = re.sub("[^0-9-,]","", next(f).strip()).split(",")[0:-1]
        if "BV6-THD" in line:
            HS_EV6_THD = re.sub("[^0-9-,]","", next(f).strip()).split(",")[0:-1]
        if "BV7-THD" in line:
            HS_EV7_THD = re.sub("[^0-9-,]","", next(f).strip()).split(",")[0:-1]
        if "i4MixWet_BVRatio" in line:
            HS_MixW_BV = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "u4MixWet_EVDB2DRatio" in line:
            HS_MixW_EVD = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "u4MixWet_MidRatio" in line:
            HS_MixW_Mid = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "LRatio%" in line:
            line2, line3, line4, line5, line6 = [next(f) for _ in range(5)]
            HS_MixW.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
            HS_MixW.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
            HS_MixW.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
            HS_MixW.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
            HS_MixW.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
        if "//MT" in line:
            line2, line3, line4 = [next(f) for _ in range(3)]
            HS_MT_BV = re.sub("[^0-9-,]","", line4).split(",")[0:-1]
        if "EVD B2M Ratio" in line:
            EVD_B2M = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "L_DR_B2M" in line:
            line2, line3, line4, line5, line6, line7, line8, line9 = [next(f) for _ in range(8)]
            MT_THD.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
            MT_THD.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
            MT_THD.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
            MT_THD.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
            MT_THD.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
            MT_THD.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
            MT_THD.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
            MT_THD.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
        
        if "fd_dr_ra_x" in line:
            fd_dr_ra_x = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "fd_dr_ra_y" in line:
            fd_dr_ra_y = re.sub("[^0-9-,]","", line).split(",")[0:-3]
        if "flt_fdsz_ra_x" in line:
            flt_fdsz_ra_x = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "flt_fdsz_ra_y" in line:
            flt_fdsz_ra_y = re.sub("[^0-9-,]","", line).split(",")[0:-3]
        if "//int32_t  fbt_bv" in line:
            if fbt_bv != []:
                flt_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "//int32_t  fbt_dr" in line:
            if fbt_dr != []:
                flt_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "//int32_t  fbt_ns_bv" in line:
            if fbt_ns_bv != []:
                flt_ns_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_ns_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "//int32_t  fbt_ns_dr" in line:
            if fbt_ns_dr != []:
                flt_ns_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_ns_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "en_fd_locsz_bv" in line:
            en_fd_locsz_bv = re.sub(" ","", line).split(",")[0]
        if "fd_sz_smal2bg" in line:
            ws.cell(column=59, row=17).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
            ws.cell(column=60, row=17).value = int(re.sub("[^0-9-,]","", line).split(",")[1])
        if "fd_pb_smal2bg" in line:
            ws.cell(column=59, row=18).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
            ws.cell(column=60, row=18).value = int(re.sub("[^0-9-,]","", line).split(",")[1])
        if "fd_loc_near2far" in line:
            ws.cell(column=59, row=19).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
            ws.cell(column=60, row=19).value = int(re.sub("[^0-9-,]","", line).split(",")[1])
        if "fd_pb_near2far" in line:
            ws.cell(column=59, row=20).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
            ws.cell(column=60, row=20).value = int(re.sub("[^0-9-,]","", line).split(",")[1])
        if "fd_bv[AE_BVDR_MAXSIZE];" in line:
            fd_bv = re.sub("[^0-9-,]","", line).split(",")[0:-3]
        if "fd_pb_smalsz" in line:
            fd_pb_smalsz = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "fd_pb_bgsz" in line:
            fd_pb_bgsz = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "fd_pb_locnear" in line:
            fd_pb_locnear = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "fd_pb_locfar" in line:
            fd_pb_locfar = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "/*int32_t  fdy_min_bv" in line:
            fdy_min_bv = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "/*int32_t  fdy_min_dr" in line:
            fdy_min_dr = re.sub("[^0-9-,]","", line).split(",")[0:-1]
            
        if "//u4_FD_TH: FD background brightness target" in line:
            line2, line3, line4, line5, line6, line7, line8, line9, line10, line11 = [next(f) for _ in range(10)]
            if TH_tbl_1 == []:
                TH_tbl_1.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_1.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_1.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_1.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_1.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_1.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_1.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_1.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_1.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_1.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_1 != [] and TH_tbl_2 == []:
                TH_tbl_2.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_2.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_2.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_2.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_2.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_2.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_2.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_2.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_2.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_2.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_2 != [] and TH_tbl_3 == []:
                TH_tbl_3.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_3.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_3.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_3.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_3.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_3.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_3.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_3.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_3.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_3.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_3 != [] and TH_tbl_4 == []:
                TH_tbl_4.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_4.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_4.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_4.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_4.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_4.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_4.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_4.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_4.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_4.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            
        if "//u4_FD_TH: FD brightness target" in line:
            line2, line3, line4, line5, line6, line7, line8, line9, line10, line11 = [next(f) for _ in range(10)]
            if TH_tbl_5 == []:
                TH_tbl_5.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_5 != [] and TH_tbl_6 == []:
                TH_tbl_6.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_6 != [] and TH_tbl_7 == []:
                TH_tbl_7.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_7 != [] and TH_tbl_8 == []:
                TH_tbl_8.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
                
        if "fdy_min_th" in line:
            line2, line3, line4, line5, line6, line7, line8, line9, line10, line11 = [next(f) for _ in range(10)]
            min_TH_tbl.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
            min_TH_tbl.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
            min_TH_tbl.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
            min_TH_tbl.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
            min_TH_tbl.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
            min_TH_tbl.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
            min_TH_tbl.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
            min_TH_tbl.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
            min_TH_tbl.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
            min_TH_tbl.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            
        if "ns_bv_cfg_bv" in line:
            ns_bv_cfg_bv = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "ns_bv_cfg_btthd" in line:
            ns_bv_cfg_btthd = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        if "ns_bv_cfg_nsthd" in line:
            ns_bv_cfg_nsthd = re.sub("[^0-9-,]","", line).split(",")[0:-1]
        
        if "ns_flat_rt_x1" in line:
            ws.cell(column=70, row=23).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
        if "ns_flat_rt_y1" in line:
            ws.cell(column=72, row=23).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
        if "ns_flat_rt_x2" in line:
            ws.cell(column=70, row=24).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
        if "ns_flat_rt_y2" in line:
            ws.cell(column=72, row=24).value = int(re.sub("[^0-9-,]","", line).split(",")[0])
    
    ws.cell(column=3, row=16).value = int(MidL[0])
    ws.cell(column=4, row=16).value = int(MidH[0])
    
    for j in range(0,np.size(MT_BV)):
        ws.cell(column=j+3, row=17).value = int(MT_BV[j])
        ws.cell(column=j+3, row=18).value = int(MT_base[j])
        ws.cell(column=j+3, row=19).value = int(MT_exp[j])
        
    for j in range(0,np.size(HS_BT_BV)):
        ws.cell(column=j+13, row=17).value = int(HS_BT_BV[j])
    for j in range(0,np.size(HS_BT_EVD)):
        ws.cell(column=j+13, row=18).value = int(HS_BT_EVD[j])
        ws.cell(column=j+13, row=22).value = int(HS_EV0_THD[j])
        ws.cell(column=j+13, row=23).value = int(HS_EV1_THD[j])
        ws.cell(column=j+13, row=24).value = int(HS_EV2_THD[j])
        ws.cell(column=j+13, row=25).value = int(HS_EV3_THD[j])
        ws.cell(column=j+13, row=26).value = int(HS_EV4_THD[j])
        ws.cell(column=j+13, row=27).value = int(HS_EV5_THD[j])
        ws.cell(column=j+13, row=28).value = int(HS_EV6_THD[j])
        if HS_EV7_THD != []:
            ws.cell(column=j+13, row=29).value = int(HS_EV7_THD[j])
        
    for j in range(0,np.size(HS_MT_BV)):
        ws.cell(column=j+13, row=37).value = int(HS_MT_BV[j])
    for j in range(0,np.size(EVD_B2M)):
        ws.cell(column=j+13, row=38).value = int(EVD_B2M[j])
        
    for j in range(0,np.size(EVD_B2M)):
        for i in range(0,np.size(HS_MT_BV)):
            ws.cell(column=j+13, row=41+i).value = int(MT_THD[i][j])
        
    ws.cell(column=23, row=17).value = int(HS_MixW_BV[0])
    ws.cell(column=24, row=17).value = int(HS_MixW_BV[1])
    ws.cell(column=27, row=17).value = int(HS_MixW_BV[2])
    
    for j in range(0,np.size(HS_MixW_EVD)):
        if j <= 1:
            ws.cell(column=j+23, row=18).value = int(HS_MixW_EVD[j])
        else:
            ws.cell(column=27+(j-2)*3, row=18).value = int(HS_MixW_EVD[j])
    
    for j in range(0,np.size(HS_MixW_Mid)):
        if j <= 1:
            ws.cell(column=j+23, row=19).value = int(HS_MixW_Mid[j])
        else:
            ws.cell(column=27+(j-2)*3, row=19).value = int(HS_MixW_Mid[j])
            
    for j in range(0,np.size(HS_MixW_EVD)*3):
        for i in range(0,np.size(HS_MixW_Mid)*4):
            ws.cell(column=i+24, row=j+23).value = int(HS_MixW[j][i])
    
    for j in range(0,np.size(fd_dr_ra_x)):
        if j < 9:
            ws.cell(column=j+47, row=16).value = int(fd_dr_ra_x[j])
            ws.cell(column=j+47, row=17).value = int(fd_dr_ra_y[j])
        else:
            ws.cell(column=(j-9)+47, row=18).value = int(fd_dr_ra_x[j])
            ws.cell(column=(j-9)+47, row=19).value = int(fd_dr_ra_y[j])
    
    for j in range(0,np.size(flt_fdsz_ra_x)):
        if j < 9:
            ws.cell(column=j+47, row=22).value = int(flt_fdsz_ra_x[j])
            ws.cell(column=j+47, row=23).value = int(flt_fdsz_ra_y[j])
        else:
            ws.cell(column=(j-9)+47, row=24).value = int(flt_fdsz_ra_x[j])
            ws.cell(column=(j-9)+47, row=25).value = int(flt_fdsz_ra_y[j])
            
    for j in range(0,np.size(flt_bv)):
        ws.cell(column=j+46, row=29).value = int(flt_bv[j])
    for j in range(0,np.size(fbt_bv)):
        ws.cell(column=j+58, row=23).value = int(fbt_dr[j])
    for j in range(0,np.size(flt_dr)):
        ws.cell(column=j+46, row=30).value = int(flt_dr[j])
    for j in range(0,np.size(fbt_dr)):
        ws.cell(column=j+58, row=24).value = int(fbt_dr[j])
    for j in range(0,np.size(flt_ns_bv)):
        ws.cell(column=j+46, row=47).value = int(flt_ns_bv[j])
    for j in range(0,np.size(fbt_ns_bv)):
        ws.cell(column=j+58, row=41).value = int(fbt_ns_bv[j])
    for j in range(0,np.size(flt_ns_dr)):
        ws.cell(column=j+46, row=48).value = int(flt_ns_dr[j])
    for j in range(0,np.size(fbt_ns_dr)):
        ws.cell(column=j+58, row=42).value = int(fbt_ns_dr[j])
    for j in range(0,np.size(fdy_min_bv)):
        ws.cell(column=j+46, row=64).value = int(fdy_min_bv[j])
    for j in range(0,np.size(fdy_min_dr)):
        ws.cell(column=j+46, row=65).value = int(fdy_min_dr[j])
        
    if en_fd_locsz_bv == "true":
        ws.cell(column=60, row=16).value = 1
    else:
        ws.cell(column=60, row=16).value = 0
    
    if np.size(fd_bv) > 6:
        for j in range(0,6):
            ws.cell(column=j+62, row=16).value = int(fd_bv[j])
            ws.cell(column=j+62, row=17).value = int(fd_pb_smalsz[j])
            ws.cell(column=j+62, row=18).value = int(fd_pb_bgsz[j])
            ws.cell(column=j+62, row=19).value = int(fd_pb_locnear[j])
            ws.cell(column=j+62, row=20).value = int(fd_pb_locfar[j])
    else:
        for j in range(0,np.size(fd_bv)):
            ws.cell(column=j+62, row=16).value = int(fd_bv[j])
            ws.cell(column=j+62, row=17).value = int(fd_pb_smalsz[j])
            ws.cell(column=j+62, row=18).value = int(fd_pb_bgsz[j])
            ws.cell(column=j+62, row=19).value = int(fd_pb_locnear[j])
            ws.cell(column=j+62, row=20).value = int(fd_pb_locfar[j])
    
    for j in range(0,np.size(fbt_dr)):
        for i in range(0,np.size(fbt_bv)):
            ws.cell(column=j+58, row=28+i).value = int(TH_tbl_1[i][j])
    for j in range(0,np.size(fbt_ns_dr)):
        for i in range(0,np.size(fbt_ns_bv)):
            ws.cell(column=j+58, row=45+i).value = int(TH_tbl_3[i][j])
    for j in range(0,np.size(flt_dr)):
        for i in range(0,np.size(flt_bv)):
            ws.cell(column=j+46, row=34+i).value = int(TH_tbl_5[i][j])
    for j in range(0,np.size(flt_ns_dr)):
        for i in range(0,np.size(flt_ns_bv)):
            ws.cell(column=j+46, row=51+i).value = int(TH_tbl_7[i][j])
    for j in range(0,np.size(fdy_min_dr)):
        for i in range(0,np.size(fdy_min_bv)):
            ws.cell(column=j+46, row=68+i).value = int(min_TH_tbl[i][j])
    
    for j in range(0, np.size(ns_bv_cfg_bv)):
        ws.cell(column=j+70, row=18).value = int(ns_bv_cfg_bv[j])
        ws.cell(column=j+70, row=19).value = int(ns_bv_cfg_btthd[j])
        ws.cell(column=j+70, row=20).value = int(ns_bv_cfg_nsthd[j])
    
    print("AE.cpp is ok!")
    return wb

##############################
class SX3(QWidget):
    
    update_progress_bar_signal = pyqtSignal(int)
    set_progress_bar_value_signal = pyqtSignal(int)
    def __init__(self) -> None:
        super().__init__()
    def run(self, exif_path, code_path):
##############################
        print("mtkAEanalysis is runing...")

        # root = tk.Tk()
        # root.withdraw()
        # code_path = filedialog.askopenfilename()
        # print(code_path)

        # refer = input("Have reference or not (0: no, 1: yes): ")
        # refer = int(refer)

        localtime = time.localtime()
        clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])

        # exif_path = "Exif"
        allFileList = os.listdir(exif_path)
        allFileList_exif = np.sort(allFileList,axis=0)
        allFileList_exif = list(filter(file_filter, allFileList_exif))
        allFileList_exif.sort(key=natural_keys)
        allFileList_jpg = np.sort(allFileList,axis=0)
        allFileList_jpg = list(filter(file_filter_jpg, allFileList_jpg))
        allFileList_jpg.sort(key=natural_keys)
        
        ##############################
        self.set_progress_bar_value_signal.emit(np.size(allFileList_exif))
        ##############################
        for i in range(0,(np.size(allFileList_exif))):
            ##############################
            self.update_progress_bar_signal.emit(i+1) 
            ##############################
            path_name = exif_path + "/" + allFileList_exif[i]
            exifFile = open(path_name, "r")
            file_name = os.path.basename(path_name)
            base = os.path.splitext(file_name)[0]
            baseTag = base.split(".")[0]
            
            if i % 20 == 0:
                startNum = base.split("_")[0]
                wb = create_xls(code_path)
            
            sheet = wb[wb.sheetnames[0]]
            target = wb.copy_worksheet(sheet)
            target.title = baseTag
            wb.active = int((i%20)+2)
            ws = wb.active
            
            print(base)
            
            AE_TAG_STABLE = []
            AE_TAG_FACE_STS_SIZE = []
            AE_TAG_CWV = 0
            AE_TAG_NS_DT_THD = []
            AE_TAG_NS_DT_LIMIT = []
            
            for line in exifFile:
                if "AE_TAG_CWV" in line:
                    if AE_TAG_CWV == 0:
                        ws.cell(column=5, row=11).value = int(re.sub("[^0-9-,]","", line))
                        AE_TAG_CWV = 1
                if "AE_TAG_REALBVX1000" in line:
                    ws.cell(column=5, row=12).value = int(re.sub("[^0-9-,]","", line)[4:])
                if "AE_TAG_MTV6_MAINT_MID_INTRATIO" in line:
                    ws.cell(column=5, row=13).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_MTV6_MAINT_Y" in line:
                    ws.cell(column=5, row=14).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_HS_EVD" in line:
                    ws.cell(column=15, row=13).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_HSV6_BT_Final_Y" in line:
                    AE_TAG_HSV6_BT_Final_Y = int(re.sub("[^0-9-,]","", line)[1:])
                    ws.cell(column=15, row=14).value = AE_TAG_HSV6_BT_Final_Y
                if "AE_TAG_HSV6_BT_THD" in line:
                    AE_TAG_HSV6_BT_THD = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_DRV6_CORR_B2M" in line:
                    ws.cell(column=15, row=33).value = int(re.sub("[^0-9-,]","", line)[2:])
                if "AE_TAG_HSV6_MT_Final_Y" in line:
                    AE_TAG_HSV6_MT_Final_Y = int(re.sub("[^0-9-,]","", line)[1:])
                    ws.cell(column=15, row=34).value = AE_TAG_HSV6_MT_Final_Y
                if "AE_TAG_HSV6_MT_THD" in line:
                    AE_TAG_HSV6_MT_THD = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_DRV6_B2D" in line:
                    ws.cell(column=27, row=13).value = int(re.sub("[^0-9-,]","", line)[2:])
                if "AE_TAG_DRV6_MIDRATIO" in line:
                    ws.cell(column=27, row=14).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_HSV6_DT_FINAL_Y" in line:
                    AE_TAG_HSV6_DT_FINAL_Y = int(re.sub("[^0-9-,]","", line)[1:])
                    ws.cell(column=15, row=39).value = AE_TAG_HSV6_DT_FINAL_Y
                if "AE_TAG_HSV6_DT_THD" in line:
                    AE_TAG_HSV6_DT_THD = int(re.sub("[^0-9-,]","", line)[1:])
                    ws.cell(column=19, row=39).value = AE_TAG_HSV6_DT_THD
                    
                if "AE_TAG_AE_TARGET" in line:
                    ws.cell(column=46, row=3).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_STABLE" in line:
                    AE_TAG_STABLE.append(re.sub("[^0-9-,]","", line))
                
                if "AE_TAG_PROB_FACE_HBND" in line:
                    ws.cell(column=48, row=5).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_FACE_LBND" in line:
                    ws.cell(column=48, row=6).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_FACE_FLT_DR" in line:
                    ws.cell(column=48, row=8).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_FACE_FLT_HBND" in line:
                    ws.cell(column=48, row=9).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                
                if "AE_TAG_FACE_LOW_BOUND" in line:
                    ws.cell(column=52, row=3).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FACE_HIGH_BOUND" in line:
                    ws.cell(column=52, row=4).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_LINK_FACE_TH_MIN" in line:
                    ws.cell(column=52, row=7).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_LINK_AE_CWR_STABLE" in line:
                    ws.cell(column=55, row=6).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_LINK_MAX_GAIN" in line:
                    ws.cell(column=55, row=7).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_LINK_FACE_CWR_STABLE" in line:
                    ws.cell(column=55, row=9).value = int(re.sub("[^0-9-,]","", line))
                    
                if "AE_TAG_FLT_FDY" in line:
                    ws.cell(column=48, row=14).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FLT_OE_SYS" in line:
                    ws.cell(column=51, row=11).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FLT_OETH" in line:
                    ws.cell(column=51, row=12).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FLT_FDDR" in line:
                    ws.cell(column=53, row=12).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FACE_STS_SIZE" in line:
                    AE_TAG_FACE_STS_SIZE.append(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FBT_FDY" in line:
                    ws.cell(column=60, row=14).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FBT_OE_SYS" in line:
                    ws.cell(column=63, row=11).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FBT_OETH" in line:
                    ws.cell(column=63, row=12).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_FACE_STS_LOC" in line:
                    ws.cell(column=65, row=12).value = int(re.sub("[^0-9-,]","", line))
                    
                if "AE_TAG_NS_STS_BVPROB" in line:
                    ws.cell(column=74, row=3).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_STS_CDFPROB" in line:
                    ws.cell(column=74, row=4).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_PROB_NS_THD" in line:
                    ws.cell(column=74, row=6).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_BT_FLAT" in line:
                    ws.cell(column=71, row=7).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_BT_THD" in line:
                    ws.cell(column=74, row=7).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_DT_THD" in line:
                    ws.cell(column=74, row=8).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_DT_LBND" in line:
                    ws.cell(column=71, row=9).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                if "AE_TAG_PROB_NS_DT_HBND" in line:
                    ws.cell(column=74, row=9).value = int(int(re.sub("[^0-9-,]","", line))/1024*100)
                    
                if "AE_TAG_NS_EVD" in line:
                    ws.cell(column=72, row=13).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_FLATTHD" in line:
                    ws.cell(column=72, row=14).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_NORMAL_Y" in line:
                    ws.cell(column=76, row=13).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_BT_Y" in line:
                    ws.cell(column=80, row=13).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_DT_Y" in line:
                    ws.cell(column=84, row=13).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_DT_THD" in line:
                    ws.cell(column=82, row=12).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_DT_LIMIT" in line:
                    ws.cell(column=84, row=12).value = int(re.sub("[^0-9-,]","", line))
                
                if "AE_TAG_NS_SKYENABLE" in line:
                    ws.cell(column=77, row=22).value = int(re.sub("[^0-9-,]","", line))
                if "AE_TAG_NS_SKYBV_X1" in line:
                    ws.cell(column=74, row=23).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_NS_SKYBV_Y1" in line:
                    ws.cell(column=77, row=23).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_NS_SKYBV_X2" in line:
                    ws.cell(column=74, row=24).value = int(re.sub("[^0-9-,]","", line)[1:])
                if "AE_TAG_NS_SKYBV_Y2" in line:
                    ws.cell(column=77, row=24).value = int(re.sub("[^0-9-,]","", line)[1:])
                
            ws.cell(column=48, row=3).value = int(AE_TAG_STABLE[2])
            ws.cell(column=53, row=13).value = int(AE_TAG_FACE_STS_SIZE[0])
            
            for j in range(0,(np.size(allFileList_jpg))):
                path_name_jpg = exif_path + "/" + allFileList_jpg[j]
                file_name_jpg = os.path.basename(path_name_jpg)
                base2 = os.path.splitext(file_name_jpg)[0]
                if file_name_jpg == base or base2 == base[0:-8]:
                    img = cv2.imdecode( np.fromfile( file = path_name_jpg, dtype = np.uint8 ), cv2.IMREAD_COLOR )
                    height, width = img.shape[0], img.shape[1]
                    
                    if height > width:
                        save_img = openpyxl.drawing.image.Image(path_name_jpg)
                        save_img.height = 176
                        save_img.width = 176 * width / height
                        save_img.anchor = 'Y2'
                        ws.add_image(save_img)
                    else:
                        save_img = openpyxl.drawing.image.Image(path_name_jpg)
                        save_img.height = 176
                        save_img.width = 176 * width / height
                        save_img.anchor = 'Y2'
                        ws.add_image(save_img)
                    
                    refer = 0
                    if j+1 < np.size(allFileList_jpg) and allFileList_jpg[j+1].split("_")[0] == path_name.split("_")[0]: 
                        path_name_jpg2 = code_path + "/" + allFileList_jpg[j+1]
                        refer = 1
                    if j-1>=0 and allFileList_jpg[j-1].split("_")[0] == path_name.split("_")[0]: 
                        path_name_jpg2 = code_path + "/" + allFileList_jpg[j-1]
                        refer = 1
                        
                    if refer == 1:
                        if j % 2 == 0:
                            path_name_jpg2 = exif_path + "/" + allFileList_jpg[j+1]
                        else:
                            path_name_jpg2 = exif_path + "/" + allFileList_jpg[j-1]
                        
                        file_name_jpg2 = os.path.basename(path_name_jpg2)
                        img2 = cv2.imdecode( np.fromfile( file = path_name_jpg2, dtype = np.uint8 ), cv2.IMREAD_COLOR )
                        height2, width2 = img2.shape[0], img2.shape[1]
                        
                        save_img2 = openpyxl.drawing.image.Image(path_name_jpg2)
                        if height > width and save_img2.height < save_img2.width:
                            rotate_name = exif_path + "/" + os.path.splitext(file_name_jpg2)[0] + "_rotate.png"
                            img2_rotate = Image.open(path_name_jpg2)
                            img2_rotate = img2_rotate.rotate(270, expand = True)
                            img2_rotate.save(rotate_name)
                            save_img2 = openpyxl.drawing.image.Image(rotate_name)
                            save_img2.height = 176
                            save_img2.width = 176 * width2 / height2 
                            save_img2.anchor = 'AD2'
                            ws.add_image(save_img2)
                        elif height > width and save_img2.height > save_img2.width:
                            save_img2.height = 176
                            save_img2.width = 176 * width2 / height2
                            save_img2.anchor = 'AD2'
                            ws.add_image(save_img2)
                        elif height < width and save_img2.height > save_img2.width:
                            rotate_name = exif_path + "/" + os.path.splitext(file_name_jpg2)[0] + "_rotate.png"
                            img2_rotate = Image.open(path_name_jpg2)
                            img2_rotate = img2_rotate.rotate(270, expand = True)
                            img2_rotate.save(rotate_name)
                            save_img2 = openpyxl.drawing.image.Image(rotate_name)
                            save_img2.height = 176
                            save_img2.width = 176 * width2 / height2 
                            save_img2.anchor = 'AH2'
                            ws.add_image(save_img2)
                        else:
                            save_img2.height = 176
                            save_img2.width = 176 * width2 / height2
                            save_img2.anchor = 'AH2'
                            ws.add_image(save_img2)
                            
                        gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)
                        plt.figure(figsize=(2.88,2.2))
                        plt.hist(gray2.ravel(), 255, [0, 255], color='powderblue',alpha=0.75)
                        plt.yticks(alpha=0)
                        # plt.xlim(0,255)
                        # plt.ylim(0,350000)
                        save_name2 = exif_path + "/" + os.path.splitext(file_name_jpg2)[0] + "_histogram.png"
                        plt.savefig(save_name2,dpi=100)
                        plt.close()
                        save_img_hist2 = openpyxl.drawing.image.Image(save_name2)
                        save_img_hist2.anchor = 'F24'
                        ws.add_image(save_img_hist2)
                        
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    plt.figure(figsize=(2.88,2.2))
                    plt.hist(gray.ravel(), 255, [0, 255], color='powderblue',alpha=0.75)
                    plt.axvline(AE_TAG_HSV6_BT_Final_Y/4096*255, color='r',linestyle=':')
                    plt.axvline(AE_TAG_HSV6_BT_THD/4096*255, color='r')
                    plt.axvline(AE_TAG_HSV6_MT_Final_Y/4096*255, color='b',linestyle=':')
                    plt.axvline(AE_TAG_HSV6_MT_THD/4096*255, color='b')
                    plt.axvline(AE_TAG_HSV6_DT_FINAL_Y/4096*255, color='g',linestyle=':')
                    plt.axvline(AE_TAG_HSV6_DT_THD/4096*255, color='g')
                    plt.yticks(alpha=0)
                    # plt.xlim(0,255)
                    # plt.ylim(0,350000)
                    save_name1 = exif_path + "/" + os.path.splitext(file_name_jpg)[0] + "_histogram.png"
                    plt.savefig(save_name1,dpi=100)
                    plt.close()
                    save_img_hist = openpyxl.drawing.image.Image(save_name1)
                    save_img_hist.anchor = 'B24'
                    ws.add_image(save_img_hist)
                    
                    faceCase = "faceCase.png"
                    if os.path.exists(faceCase):
                        face_img = openpyxl.drawing.image.Image(faceCase)
                        face_img_2 = cv2.imdecode( np.fromfile( file = faceCase, dtype = np.uint8 ), cv2.IMREAD_COLOR )
                        height3, width3 = face_img_2.shape[0], face_img_2.shape[1]
                        face_img.height = 176
                        face_img.width = 176 * width3 / height3
                        face_img.anchor = 'BE2'
                        ws.add_image(face_img)
                    else:
                        print("No faceCase.png in path.")
                    
            if i % 20 == 19 or i == len(allFileList_exif)-1:
                endNum = base.split("_")[0]
                file = "mtkAEanalysis_SX3_" + str(localtime[0]) + "_" + str(localtime[1]) + "_" + str(localtime[2]) + "_" + clock + "_" + startNum + "_" + endNum + ".xlsm"
                wb.active = 0
                wb.save(file)
        
        print("mtkAEanalysis is ok!")
        # os.system("pause")
        ##############################
        return file
        ##############################