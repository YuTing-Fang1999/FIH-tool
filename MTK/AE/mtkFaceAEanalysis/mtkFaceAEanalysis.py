import os
import cv2
import numpy as np
import openpyxl
import tkinter as tk
from tkinter import filedialog
import time
import re
import math
from PIL import Image, ImageStat

def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]

def file_filter(f):
    if f[-5:] in ['.exif'] or f[-4:] in ['.txt']:
        return True
    else:
        return False

def file_filter_jpg(f):
    if f[-4:] in ['.jpg', '.JPG']:
        return True
    else:
        return False
    
def faceDetect(img_path):
    # Read the input image
    img = cv2.imdecode( np.fromfile( file = img_path, dtype = np.uint8 ), cv2.IMREAD_COLOR )
     
    # Convert into grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Load the cascade
    face_cascade = cv2.CascadeClassifier('MTK/AE/mtkFaceAEanalysis/haarcascade_frontalface_alt2.xml')
    
    # Detect faces
    faces = face_cascade.detectMultiScale(gray, 1.1, 4)
     
    # Draw rectangle around the faces and crop the faces
    faceScale = 0
    for (x, y, w, h) in faces:
        w1 = w-int(w*0.8)
        w2 = int(w*0.8)
        h1 = h - int(h*0.9)
        h2 = int(h*0.9)
        scale = w*h
        cv2.rectangle(img, (x+w1, y+h1), (x+w2, y+h2), (0, 0, 255), 2)
        if scale > faceScale:
            faces = img[y+h1:y+h2, x+w1:x+w2]
            faceScale = scale
    return faces

def brightness(img_path):
    im = Image.open(img_path)
    h, w = im.size
    center_h = h / 2
    center_w = w / 2
    scale_h = h * 0.05
    scale_w = w * 0.05
    im_crop = im.crop((center_h-scale_h, center_w-scale_w, center_h+scale_h, center_w+scale_w))
    stat = ImageStat.Stat(im_crop)
    r,g,b = stat.rms
    return math.sqrt(0.241*(r**2)+0.691*(g**2)+0.068*(b**2))
    
def create_xls(code_path, base_excel_path):
    wb = openpyxl.load_workbook(base_excel_path, read_only=False, keep_vba=True)
    wb.active = 0
    ws = wb.active
    
    f = open(code_path, "r")
    
    fbt_bv = []
    fbt_dr = []
    fbt_ns_bv = []
    fbt_ns_dr = []
    TH_tbl_5 = []
    TH_tbl_6 = []
    TH_tbl_7 = []
    TH_tbl_8 = []
    
    ws.cell(column=2, row=6).value = code_path.split("/")[-3]
    for line in f:
        if "//u4_FD_TH: FD brightness target" in line:
            line2, line3, line4, line5, line6, line7, line8, line9, line10, line11 = [next(f) for _ in range(10)]
            if TH_tbl_5 == []:
                TH_tbl_5.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_5 != [] and TH_tbl_6 == []:
                TH_tbl_6.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_6 != [] and TH_tbl_7 == []:
                TH_tbl_7.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_7 != [] and TH_tbl_8 == []:
                TH_tbl_8.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            
        if "//int32_t  fbt_bv" in line:
            if fbt_bv != []:
                flt_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "//int32_t  fbt_dr" in line:
            if fbt_dr != []:
                flt_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "//int32_t  fbt_ns_bv" in line:
            if fbt_ns_bv != []:
                flt_ns_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_ns_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "//int32_t  fbt_ns_dr" in line:
            if fbt_ns_dr != []:
                flt_ns_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_ns_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            
    for j in range(0,np.size(flt_bv)):
        ws.cell(column=j+10, row=4).value = int(flt_bv[j])
    for j in range(0,np.size(flt_dr)):
        ws.cell(column=j+10, row=5).value = int(flt_dr[j])
    for j in range(0,np.size(flt_ns_bv)):
        ws.cell(column=j+22, row=4).value = int(flt_ns_bv[j])
    for j in range(0,np.size(flt_ns_dr)):
        ws.cell(column=j+22, row=5).value = int(flt_ns_dr[j])
    for j in range(0,np.size(flt_dr)):
        for i in range(0,np.size(flt_bv)):
            ws.cell(column=j+10, row=8+i).value = int(TH_tbl_5[i][j])
    for j in range(0,np.size(flt_ns_dr)):
        for i in range(0,np.size(flt_bv)):
            ws.cell(column=j+22, row=8+i).value = int(TH_tbl_7[i][j])
    
    sheet = wb[wb.sheetnames[0]]
    target = wb.copy_worksheet(sheet)
    target.title = "default"
    
    print("AE.cpp is ok!")
    return wb
    
if __name__ == "__main__":
    print("mtkAEfaceAuto is runing...")
        
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    print(file_path)

    localtime = time.localtime()
    clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])

    exif_path = "MTK/AE/mtkFaceAEanalysis/Exif"
    allFileList = os.listdir(exif_path)
    allFileList_exif = np.sort(allFileList,axis=0)
    allFileList_exif = list(filter(file_filter, allFileList_exif))
    allFileList_exif.sort(key=natural_keys)
    allFileList_jpg = np.sort(allFileList,axis=0)
    allFileList_jpg = list(filter(file_filter_jpg, allFileList_jpg))
    allFileList_jpg.sort(key=natural_keys)

    real_num = 0

    for i in range(0,(np.size(allFileList_exif))):
        path_name = exif_path + "/" + allFileList_exif[i]
        exifFile = open(path_name, "r")
        file_name = os.path.basename(path_name)
        base = os.path.splitext(file_name)[0]
        baseTag = base.split(".")[0]
        num = re.sub("[^0-9-,]","", base[0:2])
        
        if i == 0:
            startNum = re.sub("[^0-9-,]","", base[0:2])
            wb = create_xls(file_path)
            wb.active = 0
            ws = wb.active
        
        print(base)
        
        AE_TAG_FACE_20_FACE_AE_STABLE = []
        AE_TAG_CWV = []
        
        for line in exifFile:
            if "AE_TAG_FACE_20_FACE_AE_STABLE" in line:
                AE_TAG_FACE_20_FACE_AE_STABLE.append(re.sub("[^0-9-,]","", line)[2:])
            if "AE_TAG_FACE_20_FACE_AE_1STSTABLE" in line:
                ws.cell(column=7, row=22+i).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "AE_TAG_PURE_AE_CWR_STABLE" in line:
                ws.cell(column=9, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_LINK_AE_CWR_STABLE" in line:
                ws.cell(column=10, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_NS_PROB" in line:
                ws.cell(column=11, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_LINK_FACE_CWR_STABLE" in line:
                ws.cell(column=12, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_LINK_FACE_CWR_MIN" in line:
                ws.cell(column=14, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_REALBVX1000" in line:
                ws.cell(column=15, row=22+i).value = int(re.sub("[^0-9-,]","", line)[4:])
            if "AE_TAG_FLT_DR" in line:
                ws.cell(column=16, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_FLT_TARGET" in line:
                ws.cell(column=17, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_CWV" in line:
                AE_TAG_CWV.append(re.sub("[^0-9-,]","", line))
        
        ws.cell(column=2, row=22+i).value = int(num)
        ws.cell(column=6, row=22+i).value = int(AE_TAG_FACE_20_FACE_AE_STABLE[0])
        ws.cell(column=8, row=22+i).value = int(AE_TAG_CWV[0])
        
        for j in range(0,(np.size(allFileList_jpg))):
            path_name_jpg = exif_path + "/" + allFileList_jpg[j]
            file_name_jpg = os.path.basename(path_name_jpg)
            base2 = os.path.splitext(file_name_jpg)[0]
            
            if file_name_jpg == base or base2 == base[0:-8]:
                img = cv2.imdecode( np.fromfile( file = path_name_jpg, dtype = np.uint8 ), cv2.IMREAD_COLOR )
                height, width = img.shape[0], img.shape[1]
                
                if height > width:
                    save_img = openpyxl.drawing.image.Image(path_name_jpg)
                    save_img.height = 100
                    save_img.width = 100 * width / height
                    anchor_name = "C" + str(22+real_num)
                    save_img.anchor = anchor_name
                    ws.add_image(save_img)
                else:
                    save_img = openpyxl.drawing.image.Image(path_name_jpg)
                    save_img.height = 100
                    save_img.width = 100 * width / height
                    anchor_name = "C" + str(22+real_num)
                    save_img.anchor = anchor_name
                    ws.add_image(save_img)
                    
                save_name = exif_path + "/" + os.path.splitext(file_name_jpg)[0] + "_crop.png"
                img_crop = faceDetect(path_name_jpg)
                cv2.imwrite(save_name, img_crop)
                img_crop2 = openpyxl.drawing.image.Image(save_name)
                height_crop, width_crop = img_crop.shape[0], img_crop.shape[1]
                img_crop2.height = 100
                img_crop2.width = 100 * width_crop / height_crop
                anchor_name = "D" + str(22+real_num)
                img_crop2.anchor = anchor_name
                ws.add_image(img_crop2)
                ws.cell(column=18, row=22+real_num).value = brightness(path_name_jpg)
                
                real_num = real_num + 1
            
            elif re.sub("[^0-9-,]","", base[0:2]) == re.sub("[^0-9-,]","", base2[0:2]):
                save_name = exif_path + "/" + os.path.splitext(file_name_jpg)[0] + "_crop.png"
                img_crop = faceDetect(path_name_jpg)
                cv2.imwrite(save_name, img_crop)
                img_crop2 = openpyxl.drawing.image.Image(save_name)
                height_crop, width_crop = img_crop.shape[0], img_crop.shape[1]
                img_crop2.height = 100
                img_crop2.width = 100 * width_crop / height_crop
                anchor_name = "E" + str(22+real_num)
                img_crop2.anchor = anchor_name
                ws.add_image(img_crop2)
                ws.cell(column=19, row=22+real_num).value = brightness(path_name_jpg)

    endNum = re.sub("[^0-9-,]","", base[0:2])
    file = "mtkFaceAEanalysis_" + str(localtime[0]) + "_" + str(localtime[1]) + "_" + str(localtime[2]) + "_" + clock + "_" + startNum + "_" + endNum + ".xlsm"
    wb.active = 0
    wb.save(file)

    print("mtkAEfaceAuto is ok!")


def gen_excel(code_path, exif_path, base_excel_path):
    print("mtkAEfaceAuto is runing...")
    print(code_path)

    localtime = time.localtime()
    clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])

    allFileList = os.listdir(exif_path)
    allFileList_exif = np.sort(allFileList,axis=0)
    allFileList_exif = list(filter(file_filter, allFileList_exif))
    allFileList_exif.sort(key=natural_keys)
    allFileList_jpg = np.sort(allFileList,axis=0)
    allFileList_jpg = list(filter(file_filter_jpg, allFileList_jpg))
    allFileList_jpg.sort(key=natural_keys)

    real_num = 0
    Pic_path = []
    Crop_path = []
    ref_Crop_path = []

    for i in range(0,(np.size(allFileList_exif))):
        path_name = exif_path + "/" + allFileList_exif[i]
        exifFile = open(path_name, "r")
        file_name = os.path.basename(path_name)
        base = os.path.splitext(file_name)[0]
        baseTag = base.split(".")[0]
        num = re.sub("[^0-9-,]","", base[0:2])
        
        if i == 0:
            startNum = re.sub("[^0-9-,]","", base[0:2])
            wb = create_xls(code_path, base_excel_path)
            wb.active = 0
            ws = wb.active
        
        print(base)
        
        AE_TAG_FACE_20_FACE_AE_STABLE = []
        AE_TAG_CWV = []
        
        for line in exifFile:
            if "AE_TAG_FACE_20_FACE_AE_STABLE" in line:
                AE_TAG_FACE_20_FACE_AE_STABLE.append(re.sub("[^0-9-,]","", line)[2:])
            if "AE_TAG_FACE_20_FACE_AE_1STSTABLE" in line:
                ws.cell(column=7, row=22+i).value = int(re.sub("[^0-9-,]","", line)[3:])
            if "AE_TAG_PURE_AE_CWR_STABLE" in line:
                ws.cell(column=9, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_LINK_AE_CWR_STABLE" in line:
                ws.cell(column=10, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_NS_PROB" in line:
                ws.cell(column=11, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_LINK_FACE_CWR_STABLE" in line:
                ws.cell(column=12, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_LINK_FACE_CWR_MIN" in line:
                ws.cell(column=14, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_REALBVX1000" in line:
                ws.cell(column=15, row=22+i).value = int(re.sub("[^0-9-,]","", line)[4:])
            if "AE_TAG_FLT_DR" in line:
                ws.cell(column=16, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_FLT_TARGET" in line:
                ws.cell(column=17, row=22+i).value = int(re.sub("[^0-9-,]","", line))
            if "AE_TAG_CWV" in line:
                AE_TAG_CWV.append(re.sub("[^0-9-,]","", line))
        
        ws.cell(column=2, row=22+i).value = int(num)
        ws.cell(column=6, row=22+i).value = int(AE_TAG_FACE_20_FACE_AE_STABLE[0])
        ws.cell(column=8, row=22+i).value = int(AE_TAG_CWV[0])
        
        for j in range(0,(np.size(allFileList_jpg))):
            path_name_jpg = exif_path + "/" + allFileList_jpg[j]
            file_name_jpg = os.path.basename(path_name_jpg)
            base2 = os.path.splitext(file_name_jpg)[0]
            
            if file_name_jpg == base or base2 == base[0:-8]:
                img = cv2.imdecode( np.fromfile( file = path_name_jpg, dtype = np.uint8 ), cv2.IMREAD_COLOR )
                Pic_path.append(path_name_jpg)
                height, width = img.shape[0], img.shape[1]
                
                if height > width:
                    save_img = openpyxl.drawing.image.Image(path_name_jpg)
                    save_img.height = 100
                    save_img.width = 100 * width / height
                    anchor_name = "C" + str(22+real_num)
                    save_img.anchor = anchor_name
                    ws.add_image(save_img)
                else:
                    save_img = openpyxl.drawing.image.Image(path_name_jpg)
                    save_img.height = 100
                    save_img.width = 100 * width / height
                    anchor_name = "C" + str(22+real_num)
                    save_img.anchor = anchor_name
                    ws.add_image(save_img)
                    
                save_name = exif_path + "/" + os.path.splitext(file_name_jpg)[0] + "_crop.png"
                img_crop = faceDetect(path_name_jpg)
                if img_crop is None:                
                    print("圖片讀取失敗！")
                # 儲存含中文檔名的圖片
                cv2.imencode('.jpg', img_crop)[1].tofile(save_name)

                # cv2.imwrite(save_name, img_crop)
                Crop_path.append(save_name)
                img_crop2 = openpyxl.drawing.image.Image(save_name)
                height_crop, width_crop = img_crop.shape[0], img_crop.shape[1]
                img_crop2.height = 100
                img_crop2.width = 100 * width_crop / height_crop
                anchor_name = "D" + str(22+real_num)
                img_crop2.anchor = anchor_name
                ws.add_image(img_crop2)
                ws.cell(column=18, row=22+real_num).value = brightness(path_name_jpg)
                
                real_num = real_num + 1
            
            elif re.sub("[^0-9-,]","", base[0:2]) == re.sub("[^0-9-,]","", base2[0:2]):
                save_name = exif_path + "/" + os.path.splitext(file_name_jpg)[0] + "_crop.png"
                img_crop = faceDetect(path_name_jpg)
                if img_crop is None:                
                    print("圖片讀取失敗！")
                # 儲存含中文檔名的圖片
                cv2.imencode('.jpg', img_crop)[1].tofile(save_name)
    
                # cv2.imwrite(save_name, img_crop)
                ref_Crop_path.append(save_name)
                img_crop2 = openpyxl.drawing.image.Image(save_name)
                height_crop, width_crop = img_crop.shape[0], img_crop.shape[1]
                img_crop2.height = 100
                img_crop2.width = 100 * width_crop / height_crop
                anchor_name = "E" + str(22+real_num)
                img_crop2.anchor = anchor_name
                ws.add_image(img_crop2)
                ws.cell(column=19, row=22+real_num).value = brightness(path_name_jpg)

    endNum = re.sub("[^0-9-,]","", base[0:2])
    file = "mtkFaceAEanalysis_" + str(localtime[0]) + "_" + str(localtime[1]) + "_" + str(localtime[2]) + "_" + clock + "_" + startNum + "_" + endNum + ".xlsm"
    wb.active = 0
    wb.save(file)

    print("mtkAEfaceAuto is ok!")
    return file, len(allFileList_exif), {"Pic_path": Pic_path, "Crop_path":Crop_path, "ref_Crop_path":ref_Crop_path}
    
    
def parse_code(file_path):
    f = open(file_path, "r")
    
    fbt_bv = []
    fbt_dr = []
    fbt_ns_bv = []
    fbt_ns_dr = []
    TH_tbl_5 = []
    TH_tbl_6 = []
    TH_tbl_7 = []
    TH_tbl_8 = []
    bv_r_c = []
    dr_r_c = []

    def get_r_c(line):
        match = re.search(r'\b\d+\b', line)
        if match:
            return int(match.group())
        else:
            print("找不到code的範圍")
            return 10
    
    for line in f:
        if "/**************Face Link Target**************/" in line:
            line2, line3 = [next(f) for _ in range(2)]
            bv_r_c.append(get_r_c(line2))
            bv_r_c.append(get_r_c(line3))
            print(bv_r_c)
        elif "//u4_FD_TH: FD brightness target" in line:
            line2, line3, line4, line5, line6, line7, line8, line9, line10, line11 = [next(f) for _ in range(10)]
            if TH_tbl_5 == []:
                TH_tbl_5.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_5.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_5 != [] and TH_tbl_6 == []:
                TH_tbl_6.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_6.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])

                
                line2, line3, line4 = [next(f) for _ in range(3)]
                dr_r_c.append(get_r_c(line3))
                dr_r_c.append(get_r_c(line4))
                print(dr_r_c)
            elif TH_tbl_6 != [] and TH_tbl_7 == []:
                TH_tbl_7.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_7.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            elif TH_tbl_7 != [] and TH_tbl_8 == []:
                TH_tbl_8.append(re.sub("[^0-9-,]","", line2).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line3).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line4).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line5).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line6).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line7).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line8).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line9).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line10).split(",")[0:-1])
                TH_tbl_8.append(re.sub("[^0-9-,]","", line11).split(",")[0:-1])
            
        if "//int32_t  fbt_bv" in line:
            if fbt_bv != []:
                flt_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "//int32_t  fbt_dr" in line:
            if fbt_dr != []:
                flt_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "//int32_t  fbt_ns_bv" in line:
            if fbt_ns_bv != []:
                flt_ns_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_ns_bv = re.sub("[^0-9-,]","", line).split(",")[0:-2]
        if "//int32_t  fbt_ns_dr" in line:
            if fbt_ns_dr != []:
                flt_ns_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
            else:
                fbt_ns_dr = re.sub("[^0-9-,]","", line).split(",")[0:-2]
                
    return {
        "flt_bv": np.array(flt_bv).astype(np.int),
        "flt_dr": np.array(flt_dr).astype(np.int),
        "flt_ns_bv": np.array(flt_ns_bv).astype(np.int),
        "flt_ns_dr": np.array(flt_ns_dr).astype(np.int),
        "TH_tbl_5": np.array(TH_tbl_5).astype(np.int),
        "TH_tbl_6": np.array(TH_tbl_6).astype(np.int),
        "TH_tbl_7": np.array(TH_tbl_7).astype(np.int),
        "TH_tbl_8": np.array(TH_tbl_8).astype(np.int),
        "normal_light_r": bv_r_c[0]-1,
        "normal_light_c": bv_r_c[1]-1,
        "low_light_r": dr_r_c[0]-1,
        "low_light_c": dr_r_c[1]-1,
    }
