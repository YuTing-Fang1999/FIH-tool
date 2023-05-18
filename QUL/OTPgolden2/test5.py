from copy import deepcopy

import openpyxl
from openpyxl.chart import SurfaceChart, SurfaceChart3D, Reference
from openpyxl.drawing.image import Image
import numpy as np

# Load your workbook and sheet as you want, for example
wb = openpyxl.load_workbook('GM2_分析.xlsx', data_only=True)
ws = wb['Range1']

# 取得圖表資料
data = ws['G1':'W13']
# 將資料轉換為 NumPy 陣列
# data = np.array([[cell.value for cell in row] for row in data])

# 創建 SurfaceChart3D 圖表
chart = SurfaceChart3D()
chart.title = ws['F1'].value
data = Reference(ws, min_row=1, min_col=7, max_row=13, max_col=23)
chart.add_data(data)

# set the elevation to 90 degrees and the rotation to -90 degrees
chart.elevation = 90
chart.rotation = -90

# add the chart to the worksheet
ws.add_chart(chart, "X58")

# save the workbook to a file
wb.save("my_chart.xlsx")