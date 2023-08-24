from PyQt5.QtWidgets import QWidget, QApplication, QFileDialog, QStatusBar, QMessageBox, QLabel
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from .UI import Ui_Form
import win32com.client as win32
from myPackage.ParentWidget import ParentWidget
from myPackage.ImageViewer import ImageViewer
import os
import numpy as np
import re
import xml.etree.ElementTree as ET
import time
import cv2
import openpyxl
import xlwings as xw


class TxtWorkerThread(QThread):
        update_status_bar_signal = pyqtSignal(str)
        failed_signal = pyqtSignal(str)
        finish_signal = pyqtSignal()
        txt_path = ""

        def __init__(self, excel_template_path):
            super().__init__()
            self.excel_template_path = excel_template_path

        def run(self):
            print(f"Selected file: {self.txt_path}")
            try:
                self.update_status_bar_signal.emit("Loading txt...")
                self.load_txt(self.txt_path)
                self.finish_signal.emit()
            except Exception as error:
                print(error)
                self.update_status_bar_signal.emit("Failed to Load txt..."+str(error))
                self.failed_signal.emit("Failed to Load txt...\n"+str(error))

        def load_txt(self, txt_path):
            gain_title, gain_arr = self.parse_txt(txt_path)
            assert len(gain_title) == 4
            for i in range(4):
                gain_arr[i] = gain_arr[i].flatten()

            gain_arr = np.array(gain_arr).T
            # print(gain_arr.shape)

            # open excel
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False  # Set to True if you want to see the Excel application
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(self.excel_template_path)
            sheet = workbook.Worksheets('goldenOTP_check')
            sheet.Range('C3:F223').Value = gain_arr
            workbook.Save()
            workbook.Close()
            excel.DisplayAlerts = True

            self.update_status_bar_signal.emit("Load txt successfully")
            
        def parse_txt(self, fanme):
            
            with open(fanme) as f:
                text = f.read() + '\n\n'

                pattern = r"_gain:\n(.*?)\n\n"
                result = re.findall(pattern, text, re.DOTALL|re.MULTILINE)
                gain_arr = []
                for gain_txt in result:
                    # Split the string by the newline character
                    lines = gain_txt.split("\n")
                    # Split each line by whitespace and convert to floats
                    data = [[float(x) for x in line.split()] for line in lines if line.strip()]
                    gain_arr.append(np.array(data))
                pattern = r"\b\w+gain\b"
                gain_title = re.findall(pattern, text)
            
            return gain_title, gain_arr
        
class XmlWorkerThread(QThread):
        finish_signal = pyqtSignal(list)   
        failed_signal = pyqtSignal(str)
        update_status_bar_signal = pyqtSignal(str)
        xml_excel_path = None
        filepath = ""
        excel_path = ""

        def __init__(self):
            super().__init__()

        def run(self):
            try:
                self.update_status_bar_signal.emit("Loading txt...")
                print(f"Selected file: {self.filepath}")

                tree = ET.parse(self.filepath)
                root = tree.getroot()

                if 'lsc34' in root.tag:
                    lsc_type = 'lsc34'
                elif 'lsc35' in root.tag:
                    lsc_type = 'lsc35'
                else:
                    raise ValueError('Unsupported LSC type in XML')

                control_method = root.find('.//control_method')
                aec_exp_control = control_method.find('aec_exp_control').text    
                golden_data, data = self.find_lsc_data(root, lsc_type, aec_exp_control)

                localtime = time.localtime()
                clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])
            
                self.xml_excel_path = os.path.join(os.getcwd(), f'LSC_checkTool_{localtime[0]}_{localtime[1]}_{localtime[2]}_{clock}.xlsm')


                wb = self.create_xls(self.excel_path)
                wb.active = 0
                wb.save(self.xml_excel_path)

                app = xw.App(visible=False)
                wb = xw.Book(self.xml_excel_path)
                macro_vba = wb.app.macro('CopySheetWithChart')

                for i, item in enumerate(data, start=2):
                    if aec_exp_control == "control_lux_idx":
                        print(f'region {str(i).rjust(3)}: '
                            f'lux_idx_start: {str(item["lux_idx_start"]).rjust(5)}, '
                            f'lux_idx_end: {str(item["lux_idx_end"]).rjust(5)}, '
                            f'start: {str(item["cct_data"]["start"]).rjust(5)}, '
                            f'end: {str(item["cct_data"]["end"]).rjust(5)}')
                        sheet_name = f'lux_{item["lux_idx_start"]}_{item["lux_idx_end"]}_cct_{item["cct_data"]["start"]}_{item["cct_data"]["end"]}'
                    else:
                        print(f'region {str(i).rjust(3)}: '
                            f'gain_start: {str(item["gain_start"]).rjust(5)}, '
                            f'gain_end: {str(item["gain_end"]).rjust(5)}, '
                            f'start: {str(item["cct_data"]["start"]).rjust(5)}, '
                            f'end: {str(item["cct_data"]["end"]).rjust(5)}')
                        sheet_name = f'gain_{item["gain_start"]}_{item["gain_end"]}_cct_{item["cct_data"]["start"]}_{item["cct_data"]["end"]}'
                    macro_vba(sheet_name)
                    self.update_status_bar_signal.emit(sheet_name)

                wb.sheets[0].activate()
                wb.save()
                app.quit()

                wb = openpyxl.load_workbook(self.xml_excel_path, read_only=False, keep_vba=True)
                for i, item in enumerate(data, start=2):
                    wb.active = i
                    ws = wb.active
                    self.update_status_bar_signal.emit(ws.title)
                    r_gain_values = item["cct_data"]["r_gain"].split()
                    gr_gain_values = item["cct_data"]["gr_gain"].split()
                    gb_gain_values = item["cct_data"]["gb_gain"].split()
                    b_gain_values = item["cct_data"]["b_gain"].split()
                    for j, r_gain in enumerate(r_gain_values, start=3):
                        ws.cell(row=j, column=3).value = round(float(r_gain_values[j - 3]), 3)
                    for j, gr_gain in enumerate(gr_gain_values, start=3):
                        ws.cell(row=j, column=4).value = round(float(gr_gain_values[j - 3]), 3)
                    for j, gb_gain in enumerate(gb_gain_values, start=3):
                        ws.cell(row=j, column=5).value = round(float(gb_gain_values[j - 3]), 3)
                    for j, b_gain in enumerate(b_gain_values, start=3):
                        ws.cell(row=j, column=6).value = round(float(b_gain_values[j - 3]), 3)

                wb.active = 0
                ws = wb.active
                r_gain_values = golden_data["r_gain"]
                gr_gain_values = golden_data["gr_gain"]
                gb_gain_values = golden_data["gb_gain"]
                b_gain_values = golden_data["b_gain"]
                for j in range(3, 3 + len(r_gain_values)):
                    ws.cell(row=j, column=3).value = r_gain_values[j - 3]
                for j in range(3, 3 + len(gr_gain_values)):
                    ws.cell(row=j, column=4).value = gr_gain_values[j - 3]
                for j in range(3, 3 + len(gb_gain_values)):
                    ws.cell(row=j, column=5).value = gb_gain_values[j - 3]
                for j in range(3, 3 + len(b_gain_values)):
                    ws.cell(row=j, column=6).value = b_gain_values[j - 3]

                wb.save(self.xml_excel_path)

                self.update_status_bar_signal.emit("LSCcheck is ok!")
                print("LSCcheck is ok!")
                time.sleep(1)
                
                item_list = []
                if aec_exp_control == "control_lux_idx":
                    for i, item in enumerate(data, start=2):
                        item_name = f'lux_{item["lux_idx_start"]}_{item["lux_idx_end"]}_cct_{item["cct_data"]["start"]}_{item["cct_data"]["end"]}'
                        item_list.append(item_name)
                else:
                    for i, item in enumerate(data, start=2):
                        item_name = f'lux_{item["gain_start"]}_{item["gain_end"]}_cct_{item["cct_data"]["start"]}_{item["cct_data"]["end"]}'
                        item_list.append(item_name)
                item_list.append("LSC golden OTP(xml)")
                self.finish_signal.emit(item_list)
            except Exception as error:
                print(error)
                self.update_status_bar_signal.emit("Failed to Load xml..."+str(error))
                self.failed_signal.emit("Failed to Load xml...\n"+str(error))

        def create_xls(self, fn):
            wb = openpyxl.load_workbook(fn, read_only=False, keep_vba=True)
            wb.active = 0
            return wb
        
        def find_lsc_data(self, root, lsc_type, aec_exp_control):
            data = []
            golden_data = []
            
            lsc_golden_rgn_data = root.find(f'.//{lsc_type}_golden_rgn_data')
            r_gain_values = lsc_golden_rgn_data.find('r_gain_tab/r_gain').text
            r_gain_list = [float(value) for value in r_gain_values.split()]
            gr_gain_values = lsc_golden_rgn_data.find('gr_gain_tab/gr_gain').text
            gr_gain_list = [float(value) for value in gr_gain_values.split()]
            gb_gain_values = lsc_golden_rgn_data.find('gb_gain_tab/gb_gain').text
            gb_gain_list = [float(value) for value in gb_gain_values.split()]
            b_gain_values = lsc_golden_rgn_data.find('b_gain_tab/b_gain').text
            b_gain_list = [float(value) for value in b_gain_values.split()]
            
            golden_data = {
                "r_gain": r_gain_list,
                "gr_gain": gr_gain_list,
                "gb_gain": gb_gain_list,
                "b_gain": b_gain_list
            }
            
            if aec_exp_control == "control_lux_idx":
                for i, mod_lsc_aec_data in enumerate(root.findall(f'.//mod_{lsc_type}_aec_data')):
                    aec_trigger = mod_lsc_aec_data.find('aec_trigger')
                    start = aec_trigger.find('lux_idx_start').text
                    end = aec_trigger.find('lux_idx_end').text
                
                    for mod_lsc_cct_data in mod_lsc_aec_data.findall(f'.//mod_{lsc_type}_cct_data'):
                        cct_trigger = mod_lsc_cct_data.find('cct_trigger')
                        cct_start = cct_trigger.find('start').text
                        cct_end = cct_trigger.find('end').text
                        
                        lsc_rgn_data = mod_lsc_cct_data.find(f'{lsc_type}_rgn_data')
                        r_gain = lsc_rgn_data.find('r_gain_tab/r_gain').text
                        gr_gain = lsc_rgn_data.find('gr_gain_tab/gr_gain').text
                        gb_gain = lsc_rgn_data.find('gb_gain_tab/gb_gain').text
                        b_gain = lsc_rgn_data.find('b_gain_tab/b_gain').text
                
                        data.append({
                            "lux_idx_start": start,
                            "lux_idx_end": end,
                            "cct_data": {
                                "start": cct_start,
                                "end": cct_end,
                                "r_gain": r_gain,
                                "gr_gain": gr_gain,
                                "gb_gain": gb_gain,
                                "b_gain": b_gain,
                            }
                        })
            else:
                for i, mod_lsc_aec_data in enumerate(root.findall(f'.//mod_{lsc_type}_aec_data')):
                    aec_trigger = mod_lsc_aec_data.find('aec_trigger')
                    start = aec_trigger.find('gain_start').text
                    end = aec_trigger.find('gain_end').text
                
                    for mod_lsc_cct_data in mod_lsc_aec_data.findall(f'.//mod_{lsc_type}_cct_data'):
                        cct_trigger = mod_lsc_cct_data.find('cct_trigger')
                        cct_start = cct_trigger.find('start').text
                        cct_end = cct_trigger.find('end').text
                        
                        lsc_rgn_data = mod_lsc_cct_data.find(f'{lsc_type}_rgn_data')
                        r_gain = lsc_rgn_data.find('r_gain_tab/r_gain').text
                        gr_gain = lsc_rgn_data.find('gr_gain_tab/gr_gain').text
                        gb_gain = lsc_rgn_data.find('gb_gain_tab/gb_gain').text
                        b_gain = lsc_rgn_data.find('b_gain_tab/b_gain').text
                
                        data.append({
                            "gain_start": start,
                            "gain_end": end,
                            "cct_data": {
                                "start": cct_start,
                                "end": cct_end,
                                "r_gain": r_gain,
                                "gr_gain": gr_gain,
                                "gb_gain": gb_gain,
                                "b_gain": b_gain,
                            }
                        })
                    
            return golden_data, data

class SetChartWorkerThread(QThread):
        update_status_bar_signal = pyqtSignal(str)
        set_img_signal = pyqtSignal(str, int)
        update_grid_signal = pyqtSignal(list)
        failed_signal = pyqtSignal(str)

        excel_path = ""
        sheet_name = ""

        def __init__(self):
            super().__init__()

        def run(self):
            try:
                self.update_status_bar_signal.emit("load 資料中，請稍後")   
                excel = win32.Dispatch("Excel.Application")
                excel.Visible = False  # Set to True if you want to see the Excel application
                excel.DisplayAlerts = False
                workbook = excel.Workbooks.Open(self.excel_path)
                sheet = workbook.Worksheets(self.sheet_name)

                # Create the output folder if it doesn't exist
                output_folder = "charts"
                os.makedirs(output_folder, exist_ok=True)
                # Activate the sheet
                sheet.Activate()
                for i, chart in enumerate(sheet.ChartObjects()):
                    # print(chart.Chart.ChartTitle.Text)
                    # 要Activate才能存!!!
                    chart.Activate()
                    chart.Width = 400  
                    chart.Height = 250  
                    # Export each chart as .png
                    chart.Chart.Export(os.path.join(os.getcwd(), output_folder, chart.Chart.ChartTitle.Text)+".png")
                    self.set_img_signal.emit("charts/"+chart.Chart.ChartTitle.Text+".png", i)
                    # img = cv2.imdecode( np.fromfile( file = "charts/"+chart.Chart.ChartTitle.Text+".png", dtype = np.uint8 ), cv2.IMREAD_COLOR )
                    # self.img_viewer[i].setPhoto(img)

                # check NG
                # node = [(4, 7), (4, 23), (16, 7), (16, 23)]
                node = [(2, 9), (2, 25), (14, 9), (14, 25)]
                row_shift = -14
                data = []
                for i in range(4):
                    row_shift += 14
                    d = []
                    for j in range(4):
                        r1, c1 = node[j]
                        # r2, c2 = node[(j+1)%4]
                        Cells1 = str(round(sheet.Cells(r1+row_shift, c1).Value, 3))
                        d.append(Cells1)
                        # print(Cells1)
                        # Cells2 = float(sheet.Cells(r2+row_shift, c2).Value)
                        # if(abs(Cells1 - Cells2) > 0.2):
                        #     self.info_signal.emit("NG")
                        #     self.img_viewer[i].text = "NG"
                        #     self.img_viewer[i].setText()
                        #     break
                    data.append(d)


                workbook.Save()
                workbook.Close()
                excel.DisplayAlerts = True

                self.update_grid_signal.emit(data)
            except Exception as error:
                print(error)
                self.update_status_bar_signal.emit("Failed to Load txt..."+str(error))
                self.failed_signal.emit(str(error))
            

class MyWidget(ParentWidget):
    def __init__(self):
        super().__init__()  # in python3, super(Class, self).xxx = super().xxx
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.setupUi()

        self.excel_template_path = os.path.abspath("QUL/LSC/LSC_checkTool.xlsm")
        self.xml_worker = XmlWorkerThread()
        self.txt_worker = TxtWorkerThread(self.excel_template_path)
        self.set_chart_worker = SetChartWorkerThread()

        self.controller()
        
    def setupUi(self):
        # Create the status bar
        self.statusBar = QStatusBar()
        self.ui.verticalLayout_3.addWidget(self.statusBar)

        self.img_viewer = []
        for i in range(4):
            self.img_viewer.append(ImageViewer())
            self.ui.img_grid.addWidget(self.img_viewer[i], i//2, i%2)
        
    def controller(self):
        self.ui.load_and_export_txt_btn.clicked.connect(self.load_txt)
        self.ui.load_xml_btn.clicked.connect(self.load_xml)
        self.ui.open_excel_btn.clicked.connect(self.open_excel)
        self.ui.sheet_selecter.currentIndexChanged[str].connect(self.set_chart)

        self.xml_worker.finish_signal.connect(self.after_load_xml)
        self.xml_worker.failed_signal.connect(self.failed)
        self.xml_worker.update_status_bar_signal.connect(self.update_status_bar)

        self.txt_worker.finish_signal.connect(self.after_load_txt)
        self.txt_worker.update_status_bar_signal.connect(self.update_status_bar)
        self.txt_worker.failed_signal.connect(self.failed)

        self.set_chart_worker.update_grid_signal.connect(self.update_grid)
        self.set_chart_worker.update_status_bar_signal.connect(self.update_status_bar)
        self.set_chart_worker.failed_signal.connect(self.failed)
        self.set_chart_worker.set_img_signal.connect(self.set_img_viewer)

    def update_status_bar(self, text):
        self.statusBar.showMessage(text, 3000)

    def failed(self, text="Failed"):
        self.set_all_enable(True)
        QMessageBox.about(self, "Failed", text)
        
    def load_txt(self):
        filepath, filetype = QFileDialog.getOpenFileName(self,
                                                            "Open file",
                                                            self.get_path("QUL_LSC_filefolder"),  # start path
                                                            '*.txt')

        if filepath == '':
            return
        filefolder = '/'.join(filepath.split('/')[:-1])
        self.set_path("QUL_LSC_filefolder", filefolder)
        self.txt_worker.txt_path = filepath

        self.set_all_enable(False)
        self.txt_worker.start()

    def after_load_txt(self):
        index = self.ui.sheet_selecter.findText("LSC golden OTP(txt)")
        if index < 0:
            self.ui.sheet_selecter.addItem("LSC golden OTP(txt)")
    
        if self.ui.sheet_selecter.currentText() == "LSC golden OTP(txt)":
            self.set_chart("LSC golden OTP(txt)")
        else:
            self.ui.sheet_selecter.setCurrentText("LSC golden OTP(txt)")
        self.export_txt()
        self.set_all_enable(True)

    def export_txt(self):
            # open excel
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False  # Set to True if you want to see the Excel application
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(self.excel_template_path)
            sheet = workbook.Worksheets('goldenOTP_check')
            gain_arr = sheet.Range('C3:F223').Value
            workbook.Save()
            workbook.Close()
            excel.DisplayAlerts = True

            # localtime = time.localtime()
            filepath, filetype=QFileDialog.getSaveFileName(self,'save the transposed matrix',self.get_path("QUL_LSC_filefolder")+"/goldenOTP_check","*.txt")
            if filepath == '': return

            def formater(num):
                return "{:6d}".format(int(num))
            with open(filepath, 'w', newline='') as f:
                for row in gain_arr:
                    f.write(''.join(map(formater, row)))
                    f.write('\n')

    def load_xml(self):
        # Open file dialog
        filepath, filetype = QFileDialog.getOpenFileName(self,
                                                            "Open file",
                                                            self.get_path("QUL_LSC_filefolder"),  # start path
                                                            '*.xml')

        if filepath == '':
            return
        
        # filepath = "QUL/LSC/lsc34_bps.xml"
        filefolder = '/'.join(filepath.split('/')[:-1])
        self.set_path("QUL_LSC_filefolder", filefolder)

        self.set_all_enable(False)
        self.xml_worker.excel_path = self.excel_template_path
        self.xml_worker.filepath = filepath
        self.xml_worker.start()

    def after_load_xml(self, item_name):
        index = self.ui.sheet_selecter.findText("LSC golden OTP(txt)")
        self.ui.sheet_selecter.clear()
        if index >= 0:
            self.ui.sheet_selecter.addItem("LSC golden OTP(txt)")

        self.ui.sheet_selecter.addItems(item_name)
        self.ui.sheet_selecter.setCurrentText(item_name[0])
        self.set_all_enable(True)

    def set_chart(self, text):
        if text == "": return
        print("set_chart", text)
        if text == "LSC golden OTP(txt)":
            self.set_chart_worker.excel_path = self.excel_template_path
            self.set_chart_worker.sheet_name = "goldenOTP_check"
        else:
            if self.xml_worker.xml_excel_path == None: return
            if text == "LSC golden OTP(xml)": text = "goldenOTP_check"
            self.set_chart_worker.excel_path = self.xml_worker.xml_excel_path
            self.set_chart_worker.sheet_name = text
        self.set_all_enable(False)
        self.set_chart_worker.start()

    def after_set_chart(self):
        self.set_all_enable(True)

    
    def set_all_enable(self, enable):
        self.set_btn_enable(self.ui.load_and_export_txt_btn, enable)
        self.set_btn_enable(self.ui.load_xml_btn, enable)
        self.set_btn_enable(self.ui.open_excel_btn, enable)
        self.ui.sheet_selecter.setEnabled(enable)

    def set_img_viewer(self, path, i):
        img = cv2.imdecode( np.fromfile( file = path, dtype = np.uint8 ), cv2.IMREAD_COLOR )
        self.img_viewer[i].setPhoto(img)

    def update_grid(self, data):
        # Function to remove a widget from the grid layout by row and column
        def remove_widget(row, col):
            # Retrieve the widget at the specified row and column
            widget_item = self.ui.corner_grid.itemAtPosition(row, col)
            if widget_item:
                widget = widget_item.widget()
                
                # Remove the widget from the layout
                self.ui.corner_grid.removeWidget(widget)
                
                # Delete the widget
                widget.deleteLater()

        # check NG
        for i in range(4):
            for j in range(4):
                r, c = i*2+j//2+i, 1+j%2
                remove_widget(r, c)
                self.ui.corner_grid.addWidget(QLabel(data[i][j]), r, c)
                # Cells2 = float(sheet.Cells(r2+row_shift, c2).Value)
                
                # if(abs(Cells1 - Cells2) > 0.2):
                #     self.info_signal.emit("NG")
                #     self.img_viewer[i].text = "NG"
                #     self.img_viewer[i].setText()
                #     break
        self.statusBar.showMessage("成功load完資料", 3000)
        self.set_all_enable(True)

    def open_excel(self):
        if self.xml_worker.xml_excel_path == None:
            QMessageBox.about(self, "請先load xml", "請先load xml，才能打開所產生的excel")
            return
        self.statusBar.showMessage("開啟中，請稍後", 3000)
        import xlwings as xw
        app = xw.App(visible=True)
        app.books[0].close()
        # Maximize the Excel window
        app.api.WindowState = xw.constants.WindowState.xlMaximized
        wb = app.books.open(self.xml_worker.xml_excel_path)
        # Set the Excel window as the foreground window
        wb.app.activate(steal_focus=True)
        
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    Form = MyWidget()
    Form.show()
    sys.exit(app.exec_())