from PyQt5 import QtCore
from PyQt5.QtWidgets import QWidget, QApplication, QFileDialog, QMessageBox
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QUrl
from datetime import datetime
import shutil
from PyQt5.QtGui import QDesktopServices
from .newUI import Ui_Form

from time import sleep
from myPackage.ParentWidget import ParentWidget
from PyQt5.QtWebEngineWidgets import QWebEngineView
# from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from urllib.request import urlopen
# import requests
import configparser
import numpy as np
import subprocess
import sys
from .exceloutput_0215 import main as exceloutput_main
import tkinter as tk
from tkinter import filedialog
import os
import subprocess
from openpyxl import load_workbook
import urllib.request
from urllib.parse import urljoin

from tqdm import tqdm  # Import tqdm for terminal progress bar
import time


class SolverThread(QThread):
    update_status_bar_signal = pyqtSignal(str)
    failed_signal = pyqtSignal(str)
    finish_signal = pyqtSignal()
    data = None

    def __init__(self):
        super().__init__()

    def run(self):
        try:
            # . . . (要執行的程式)
            # This should be the total number of image pairs or tasks to process
            total_pairs = self.data['total_pairs']
            print()
            print("--------------------------------------------------")
            print(f"Found {total_pairs} pairs of images.")
            # # Create a tqdm progress bar in the terminal
            # with tqdm(total=total_pairs, desc="Processing", unit="pair") as pbar:
            #     for i in range(total_pairs):
            #         # Simulate task processing (e.g., image comparison)
            #         time.sleep(0.5)  # Simulate a time-consuming task

            #         # Update progress bar in the terminal
            #         pbar.update(1)  # Increment progress by 1
            exceloutput_main(self.data['folder_path_ours'], self.data['folder_path_ref'],
                             self.data['AWB_expected_path'], self.data['original_excel'],
                             self.data['new_excel'], self.data['EV_ratio'],
                             self.data['AWB_num_blocks_x'], self.data['AWB_num_blocks_y'],
                             )
            self.finish_signal.emit()
        except Exception as error:
            print(error)
            self.update_status_bar_signal.emit("Failed...")
            self.failed_signal.emit("Failed...\n"+str(error))


class MyWidget(ParentWidget):
    def __init__(self):
        super().__init__()  # in python3, super(Class, self).xxx = super().xxx
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.solver_thread = SolverThread()
        self.controller()

    def controller(self):
        self.ui.pushButton_photo_folder.clicked.connect(
            self.load_photo_folder)
        self.ui.pushButton_run.clicked.connect(self.do_run)
        self.solver_thread.failed_signal.connect(self.failed)
        self.solver_thread.finish_signal.connect(self.solver_finish)

    def load_photo_folder(self):
        # Open a dialog to select a folder
        folder_path = QFileDialog.getExistingDirectory(self, "選擇Data Path")

        if folder_path:
            # Show the folder path in the QLineEdit
            self.ui.lineEdit_photo_folder.setText(folder_path)

            # List all files in the selected folder
            files = os.listdir(folder_path)

            # Initialize a set to store unique task names
            tasks = set()

            # Loop through the files to identify valid image files with the specified naming convention
            for file in files:
                # Check for image files
                if file.endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.JPG', '.JPEG')):
                    parts = file.split('_')
                    # Ensure the filename follows the format 1_{task}_...
                    if len(parts) > 1:
                        # Extract the task text (e.g., 'ref', 'ours', etc.)
                        task = parts[1]
                        tasks.add(task)  # Add the task to the set

            if len(tasks) > 2:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Warning")
                msg.setText(
                    "Please ensure there are exactly 2 unique tasks.")
                msg.setInformativeText(
                    "Browse the folder again after cleaning.")
                msg.exec_()  # Show the message box
                return

            # Populate the combo box with unique task options
            self.ui.comboBox_ours.clear()  # Clear the combo box before adding new items
            for task in tasks:
                # Add each task to the combo box
                self.ui.comboBox_ours.addItem(task)

    def check_image_num(self, folder_path_ours, folder_path_ref):
        # List all files in the folders
        files_ours = [file for file in os.listdir(folder_path_ours) if file.endswith(
            ('.jpg', '.jpeg', '.png', '.bmp', '.gif', 'JPG', 'JPEG'))]
        files_ref = [file for file in os.listdir(folder_path_ref) if file.endswith(
            ('.jpg', '.jpeg', '.png', '.bmp', '.gif', 'JPG', 'JPEG'))]

        # If the numbers of images in both folders are not equal, show a message box
        if len(files_ours) != len(files_ref):
            # Determine which folder has more images
            if len(files_ours) > len(files_ref):
                diff = len(files_ours) - len(files_ref)
                message = f"There are {diff} more images in Ours."
            else:
                diff = len(files_ref) - len(files_ours)
                message = f"There are {diff} more images in Reference."

            if os.path.exists(folder_path_ours):
                shutil.rmtree(folder_path_ours)
            if os.path.exists(folder_path_ref):
                shutil.rmtree(folder_path_ref)

            # Create a message box to show the message
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Warning")
            msg.setText("The number of images in the two folders is unequal!")
            msg.setInformativeText(message)
            msg.exec_()  # Show the message box

            return False

        return True

    def do_run(self):
        photo_folder = self.ui.lineEdit_photo_folder.text()
        # folder_path_ref = self.ui.lineEdit_folder_path_ref.text()
        # Create AWB_expected_path by concatenating photo_folder and "/AWB"
        AWB_expected_path = os.path.join(photo_folder, "AWB")
        AE_expected_path = os.path.join(photo_folder, "AE")

        # If the directory does not exist, create it
        if not os.path.exists(AWB_expected_path):
            os.makedirs(AWB_expected_path)

        # If the directory does not exist, create it
        if not os.path.exists(AE_expected_path):
            os.makedirs(AE_expected_path)

        # Get the current time and format it as a string
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        root_folder = os.getcwd()  # Get the current working directory (root folder)

        original_excel = os.path.join(root_folder, "./NTU/Chou/example.xlsx")
        # original_excel = ".example.xlsx"
        new_excel = os.path.join(photo_folder, f"{current_time}.xlsx")
        EV_ratio = int(self.ui.lineEdit_EV_ratio.text())
        AWB_num_blocks_x = 40
        AWB_num_blocks_y = 40

        if (photo_folder == ''):
            # show the error message
            QMessageBox.about(
                self,  "ERROR", "Choose Folder.")
        # elif (not (original_excel.endswith('.xlsx'))):
        #     # show the error message
        #     QMessageBox.about(
        #         self,  "ERROR", "Choose an TONE.cpp file first.")

        else:

            # Get the selected task from the combo box
            selected_task = self.ui.comboBox_ours.currentText()

            # Create temporary folders for 'folder_path' and 'folder_path_ref'
            folder_path_ours = os.path.join(photo_folder, "folder_path_ours")
            folder_path_ref = os.path.join(photo_folder, "folder_path_ref")

            if not os.path.exists(folder_path_ours):
                os.makedirs(folder_path_ours)
            if not os.path.exists(folder_path_ref):
                os.makedirs(folder_path_ref)

            # Loop through the files again and move them to the appropriate folder
            total_pairs = 0
            files = os.listdir(photo_folder)
            for file in files:
                if file.endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.JPG', '.JPEG')):
                    parts = file.split('_')
                    if len(parts) > 1:
                        task = parts[1]
                        if task == selected_task:
                            # Move the image to folder_path
                            shutil.copy(os.path.join(
                                photo_folder, file), folder_path_ours)
                            total_pairs += 1
                        else:
                            # Move the image to folder_path_ref
                            shutil.copy(os.path.join(
                                photo_folder, file), folder_path_ref)

            if self.check_image_num(folder_path_ours, folder_path_ref):

                data = {
                    "folder_path_ours": folder_path_ours,
                    "folder_path_ref": folder_path_ref,
                    "AWB_expected_path": AWB_expected_path,
                    "original_excel": original_excel,
                    "new_excel": new_excel,
                    "EV_ratio": EV_ratio,
                    "AWB_num_blocks_x": AWB_num_blocks_x,
                    "AWB_num_blocks_y": AWB_num_blocks_y,
                    "total_pairs": total_pairs
                }
                self.solver_thread.data = data
                self.solver_thread.start()

            # Optionally: Delete the temporary folders after use (if you want to clear them after the operation)
            # shutil.rmtree(folder_path_ours)
            # shutil.rmtree(folder_path_ref)
            # mtk_main(your_path,file_path, LV_region, DR_region)
            # input()
            # sys.exit()

        pass

    # Win/Linux switch
    # def do_openfolder(self):
    #     # open folder
    #     your_path = self.ui.lineEdit_DataFolder.text()

    #     # For Windows
    #     os.startfile(your_path)

    #     # For Linux
    #     # opener = "open" if sys.platform == "darwin" else "xdg-open"
    #     # subprocess.call([opener, your_path])

    # def do_update(self):
    #     # Load Excel file
    #     def process_xlsm(file_path):
    #         workbook = load_workbook(file_path, keep_vba=True)

    #         for sheet_name in workbook.sheetnames:
    #             sheet = workbook[sheet_name]
    #             for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    #                 for cell in row:
    #                     if isinstance(cell.value, str) and "_xlfn.XLOOKUP" in cell.value:
    #                         cell.value = cell.value.replace(
    #                             "_xlfn.XLOOKUP", "XLOOKUPs")

    #         modified_file_path = os.path.join(os.path.dirname(
    #             file_path), "modified_" + os.path.basename(file_path))
    #         workbook.save(modified_file_path)

    #         print(
    #             f"{os.path.basename(file_path)} ----> modified_{os.path.basename(file_path)}")

    #     folder_path = self.ui.lineEdit_DataFolder.text()
    #     if (folder_path == ''):
    #         # show the error message
    #         QMessageBox.about(
    #             self,  "ERROR", "Choose Data Folder first.")
    #     else:
    #         for root, dirs, files in os.walk(folder_path):
    #             for file in files:
    #                 if file.endswith(".xlsm"):
    #                     process_xlsm(os.path.join(root, file))

    def failed(self, text="Failed"):
        self.set_all_enable(True)
        QMessageBox.about(self, "Failed", text)

    def solver_finish(self):
        self.set_all_enable(True)

        # Delete temp folder
        photo_folder = self.ui.lineEdit_photo_folder.text()
        folder_path_ours = os.path.join(photo_folder, "folder_path_ours")
        folder_path_ref = os.path.join(photo_folder, "folder_path_ref")
        if os.path.exists(folder_path_ours):
            shutil.rmtree(folder_path_ours)
        if os.path.exists(folder_path_ref):
            shutil.rmtree(folder_path_ref)
        # self.statusBar.hide()

    def set_all_enable(self, enable):
        self.ui.pushButton_photo_folder.setEnabled(enable)
        self.ui.pushButton_run.setEnabled(enable)


if __name__ == "__main__":
    import sys
    root = tk.Tk()
    root.withdraw()
    app = QApplication(sys.argv)
    Form = MyWidget()
    # Form.setFixedSize(590, 740)
    Form.show()
    sys.exit(app.exec_())
