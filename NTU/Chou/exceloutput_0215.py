#!/usr/bin/env python
# coding: utf-8

# Modification tag<br>
# 1108 add, 1108 delete<br>
# 1118 add, 1118 modified
# 0107 add: clustering_algorithm, cluster_image, single_image_EV_compair (previous change to ..._old)

# # Import and Function

# ## Import

# In[88]:


import numpy as np
from PIL import Image, ExifTags
import matplotlib.pyplot as plt
import cv2
import os
from skimage import io
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import argparse
import math
import tempfile
from tqdm import tqdm  # Import tqdm for terminal progress bar
import time
from openpyxl.styles import PatternFill


# ## Function

# ### Cluster

# In[89]:


def load_image(image_path):
    return io.imread(image_path)


# In[90]:


def calculate_mean_brightness(image):
    if len(image.shape) == 3:  # Color image
        gray_image = Image.fromarray(image).convert('L')
        gray_image = np.array(gray_image)
    else:  # Grayscale image
        gray_image = image
    return np.mean(gray_image)


# In[91]:


def calculate_exposure_difference(mean1, mean2):
    # Exposure difference in stops
    return np.log2(mean2 / mean1)


# In[92]:


def clustering_algorithm(histogram, N, w=8, max_iterations=500):
    # Initialize 5 clusters for EV-2, EV-1, 0EV, EV+1, EV+2
    mu1 = N / (6 + 4 * w)
    mu2 = N / (4 + 2 * w)
    mu3 = N / 2
    mu4 = N - N / (4 + 2 * w)
    mu5 = N - N / (6 + 4 * w)
    mu = [mu1, mu2, mu3, mu4, mu5]

    p = histogram / np.sum(histogram)  # Normalize histogram
    epsilon = 1e-5  # Convergence threshold
    converged = False
    iterations = 0

    while not converged and iterations < max_iterations:
        mu_old = mu.copy()
        d = np.zeros((N, 5))  # Distance matrix for 5 clusters
        for i in range(N):
            d[i, 0] = abs(i - mu[0])
            d[i, 1] = w * abs(i - mu[1])
            d[i, 2] = abs(i - mu[2])
            d[i, 3] = w * abs(i - mu[3])
            d[i, 4] = abs(i - mu[4])

        # Assign each point to the nearest cluster
        C = np.argmin(d, axis=1)

        # Update the cluster centers
        for j in range(5):
            numerator = np.sum([i * p[i] for i in range(N) if C[i] == j])
            denominator = np.sum([p[i] for i in range(N) if C[i] == j])
            if denominator != 0:
                mu[j] = numerator / denominator

        # Check for convergence
        if all(abs(mu[j] - mu_old[j]) < epsilon for j in range(5)):
            converged = True

        iterations += 1

    return mu, C


# In[93]:


def cluster_image(image):
    N = 256  # Number of possible pixel values for grayscale image

    # Compute the histogram
    histogram = cv2.calcHist([image], [0], None, [256], [0, 256]).flatten()
    mu, C = clustering_algorithm(histogram, N)
    # print("Cluster Centers:", mu)
    # print("Cluster Assignments:", C)

    # Create lists to store pixel values for each cluster
    cluster_0_pixels = []  # EV-2
    cluster_1_pixels = []  # EV-1
    cluster_2_pixels = []  # 0EV
    cluster_3_pixels = []  # EV+1
    cluster_4_pixels = []  # EV+2
    # cluster_4_OE_pixels = []  # +2EV Overexposed pixels

    # Separate the pixels into their respective clusters
    for i in range(N):
        if C[i] == 0:
            cluster_0_pixels.extend(image[image == i])
        elif C[i] == 1:
            cluster_1_pixels.extend(image[image == i])
        elif C[i] == 2:
            cluster_2_pixels.extend(image[image == i])
        elif C[i] == 3:
            cluster_3_pixels.extend(image[image == i])
        elif C[i] == 4:
            cluster_4_pixels.extend(image[image == i])
            # # Add overexposed pixels (values >= 230) to cluster_4_OE_pixels
            # if i >= 230:
            #     cluster_4_OE_pixels.extend(image[image == i])

    # Convert lists to arrays
    cluster_0_pixels = np.array(cluster_0_pixels)
    cluster_1_pixels = np.array(cluster_1_pixels)
    cluster_2_pixels = np.array(cluster_2_pixels)
    cluster_3_pixels = np.array(cluster_3_pixels)
    cluster_4_pixels = np.array(cluster_4_pixels)
    # cluster_4_OE_pixels = np.array(cluster_4_OE_pixels)

    # Group the clusters
    clusters = [
        cluster_0_pixels,
        cluster_1_pixels,
        cluster_2_pixels,
        cluster_3_pixels,
        cluster_4_pixels,
        # cluster_4_OE_pixels,  # Optional: Overexposed pixels as a separate output
    ]

    return clusters


# In[94]:


def calculate_cluster_brightness(image, clusters):
    cluster_brightness = []
    for cluster in clusters:
        cluster_pixels = image[np.isin(image, cluster)]
        cluster_brightness.append(calculate_mean_brightness(cluster_pixels))
    return cluster_brightness


# In[95]:


def showImage(image1, image2):
    # Plot images and histograms for each cluster
    fig, axs = plt.subplots(1, 2, figsize=(16, 12))

    # Display original images
    gray_image1 = PILImage.fromarray(image1).convert('L')
    axs[0].imshow(gray_image1, cmap='gray')
    axs[0].set_title("1 HDR")
    axs[0].axis('off')
    gray_image2 = PILImage.fromarray(image2).convert('L')
    axs[1].imshow(gray_image2, cmap='gray')
    axs[1].set_title("Ref")
    axs[1].axis('off')
    plt.tight_layout()
    plt.show()

    # Close the figure to free memory
    plt.close(fig)


# In[96]:


def single_pair_EV_compair(folder_path, folder_path_ref, img, img_ref):
    # Paths to the images
    image_path = os.path.join(folder_path, img)
    image_path_ref = os.path.join(folder_path_ref, img_ref)

    # Load images
    image1 = load_image(image_path)
    image2 = load_image(image_path_ref)

    # Cluster the images
    clusters1 = cluster_image(image1)
    clusters2 = cluster_image(image2)

    # print("Clusters1:", clusters1)
    # print("Clusters2:", clusters2)

    # Calculate mean brightness for each cluster
    brightness1 = calculate_cluster_brightness(image1, clusters1)
    brightness2 = calculate_cluster_brightness(image2, clusters2)

    # print("Brightness1:", brightness1)
    # print("Brightness2:", brightness2)
    # Ensure exposure differences align with 5 EV levels
    exposure_differences = [
        calculate_exposure_difference(b1, b2) for b1, b2 in zip(brightness1, brightness2)
    ]

    # Handle cases where exposure_differences might not have 5 values
    if len(exposure_differences) != 5:
        print(
            f"Warning: Expected 5 EV levels but got {len(exposure_differences)}. Adjusting...")
        while len(exposure_differences) < 5:
            exposure_differences.append(0)  # Pad with zeros if fewer clusters
        # Trim if too many clusters
        exposure_differences = exposure_differences[:5]

    # Print the exposure differences
    recommand_EV_changes = []
    EV_group = ['EV-2', 'EV-1', '0EV', 'EV+1', 'EV+2']
    for i, diff in enumerate(exposure_differences):
        if math.isnan(diff):  # Handle NaN values
            diff = 0
        # print(f'{image_path} - {EV_group[i]}: {diff:.2f} EV') #### 0421
        recommand_EV_changes.append(diff)

    return recommand_EV_changes


# ### Excel

# In[97]:


def resize_image_to_fit(image_path, max_size=(400, 400)):
    # Create a temporary file to store the resized image
    temp_file = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
    temp_file.close()  # Close the file so it can be opened by PIL
    with PILImage.open(image_path) as img:
        # Calculate aspect ratio
        aspect_ratio = img.width / img.height

        # Determine new dimensions while preserving aspect ratio
        if aspect_ratio > 1:  # Image is wider than it is tall
            new_width = max_size[0]
            new_height = int(new_width / aspect_ratio)
        else:  # Image is taller than it is wide
            new_height = max_size[1]
            new_width = int(new_height * aspect_ratio)

        # Resize the image
        img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)

        # Ensure the image is still within the max size, if not, downscale it
        # Ensure it fits within max_size bounds
        img.thumbnail(max_size, PILImage.Resampling.LANCZOS)

        # Save the resized image
        img.save(temp_file.name)
    return temp_file.name


# In[98]:

# Not used
def color_transfer(folder_path, folder_path_ref, img, img_ref):
    """
    Transfer the color distribution from the target image to the source image, while adjusting
    the Red and Blue channels relative to the Green channel. Also match the exposure based
    on the Green channels of the source and target images.
    Args:
    - source: Source image whose colors will be changed.
    - target: Target image whose colors will be replicated.
    Returns:
    # - Transferred image.
    - RGB gains.
    """
    # Load image
    source = cv2.imread(os.path.join(folder_path, img))
    target = cv2.imread(os.path.join(folder_path_ref, img_ref))

    # Resize source to match target size
    target_height, target_width = target.shape[:2]
    source_resized = cv2.resize(source, (target_width, target_height))

    # Convert the images from BGR to RGB color space
    source_rgb = cv2.cvtColor(
        source_resized, cv2.COLOR_BGR2RGB).astype("float32")
    target_rgb = cv2.cvtColor(target, cv2.COLOR_BGR2RGB).astype("float32")

    # Split the RGB channels
    r_s, g_s, b_s = cv2.split(source_rgb)
    r_t, g_t, b_t = cv2.split(target_rgb)

    # Calculate the ratio of R/G and B/G for the target image
    r_to_g_t = r_t / (g_t + 1e-5)  # Add small value to avoid division by zero
    b_to_g_t = b_t / (g_t + 1e-5)

    # Apply the same ratio to the source image (R and B channels adjusted relative to G)
    r_s_adjusted = r_to_g_t * g_s
    b_s_adjusted = b_to_g_t * g_s

    # Calculate the ratio of green channels for exposure adjustment
    green_exposure_ratio = np.mean(g_t) / (np.mean(g_s) + 1e-5)

    # Append return RGB gains
    r_to_g_t_flat = r_to_g_t.flatten()
    b_to_g_t_flat = b_to_g_t.flatten()
    r_mean = np.mean(r_to_g_t_flat)
    r_std_dev = np.std(r_to_g_t_flat)
    r_z_scores = (r_to_g_t_flat - r_mean) / r_std_dev
    b_mean = np.mean(b_to_g_t_flat)
    b_std_dev = np.std(b_to_g_t_flat)
    b_z_scores = (b_to_g_t_flat - b_mean) / b_std_dev
    threshold = 2

    # Remove outliers
    r_to_g_t_without_outliers = r_to_g_t_flat[np.abs(r_z_scores) <= threshold]
    b_to_g_t_without_outliers = b_to_g_t_flat[np.abs(b_z_scores) <= threshold]
    RBG_gains = [np.mean(r_to_g_t_without_outliers), np.mean(
        b_to_g_t_without_outliers), green_exposure_ratio]

    # Adjust the green channel of the source image for exposure matching
    g_s_adjusted = g_s * green_exposure_ratio
    r_s_adjusted *= green_exposure_ratio
    b_s_adjusted *= green_exposure_ratio

    # Clip the adjusted channels to valid range [0, 255]
    r_s_adjusted = np.clip(r_s_adjusted, 0, 255)
    g_s_adjusted = np.clip(g_s_adjusted, 0, 255)
    b_s_adjusted = np.clip(b_s_adjusted, 0, 255)

    # Merge the adjusted RGB channels back together
    rb_shifted_rgb = cv2.merge([r_s_adjusted, g_s, b_s_adjusted])
    transferred_rgb = cv2.merge([r_s_adjusted, g_s_adjusted, b_s_adjusted])

    # Convert back to BGR color space
    transferred_bgr = cv2.cvtColor(
        transferred_rgb.astype("uint8"), cv2.COLOR_RGB2BGR)
    return transferred_bgr, RBG_gains


# 1118 add

# In[99]:

# Not used
def color_transfer_for_AWB_expexted_output(folder_path, folder_path_ref, img, img_ref):
    """
    Transfer the color distribution from the target image to the source image.
    Args:
    - source: Source image whose colors will be changed.
    - target: Target image whose colors will be replicated.
    Returns:
    - Transferred image.
    """

    # Load image
    source = cv2.imread(os.path.join(folder_path, img))
    target = cv2.imread(os.path.join(folder_path_ref, img_ref))

    # Convert the images from BGR to LAB color space
    source_lab = cv2.cvtColor(source, cv2.COLOR_BGR2LAB).astype("float32")
    target_lab = cv2.cvtColor(target, cv2.COLOR_BGR2LAB).astype("float32")

    # Split the LAB channels
    l_s, a_s, b_s = cv2.split(source_lab)
    l_t, a_t, b_t = cv2.split(target_lab)

    # Function to match the means and standard deviations of each channel
    def match_histograms(source_channel, target_channel):
        source_mean, source_std = (
            np.mean(source_channel), np.std(source_channel))
        target_mean, target_std = (
            np.mean(target_channel), np.std(target_channel))

        # Scale the source image to match the target's mean and standard deviation
        result = (source_channel - source_mean) * \
            (target_std / source_std) + target_mean
        return np.clip(result, 0, 255)

    # Apply the histogram matching to each channel
    l_s = match_histograms(l_s, l_t)
    a_s = match_histograms(a_s, a_t)
    b_s = match_histograms(b_s, b_t)

    # Merge the LAB channels back together and convert to BGR
    transferred_lab = cv2.merge([l_s, a_s, b_s])
    transferred_bgr = cv2.cvtColor(
        transferred_lab.astype("uint8"), cv2.COLOR_LAB2BGR)
    return transferred_bgr


# In[110]:

# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image

def copy_data_to_ae_sheet(file_path):
    # Load workbook and sheets
    wb = load_workbook(file_path)
    sheet_src = wb['OverAll']
    sheet_ae = wb['AE']
    sheet_awb = wb['AWB']

    # Copy images from C, D, E
    for image in sheet_src._images:
        # Get the original image anchor (e.g., 'C5')
        anchor = image.anchor._from
        col_letter = anchor.col + 1  # openpyxl col is 0-based
        row = anchor.row + 1         # openpyxl row is 0-based

        if col_letter in [3, 4, 5]:  # C = 3, D = 4, E = 5
            # Clone the image and re-anchor to same row in AE sheet
            img_copy = Image(image.ref)
            img_copy.width = image.width
            img_copy.height = image.height
            cell = sheet_ae.cell(row=row, column=col_letter)
            img_copy.anchor = cell.coordinate
            sheet_ae.add_image(img_copy)
            sheet_awb.add_image(img_copy)

    # Copy numeric values from F (6) to J (10)
    for row in range(3, sheet_src.max_row + 1):
        for col in range(6, 11):  # F=6, G=7, H=8, I=9, J=10
            value = sheet_src.cell(row=row, column=col).value
            sheet_ae.cell(row=row, column=col, value=value)

    # === Copy numeric data from K-M to AWB sheet at F-H ===
    for row in range(3, sheet_src.max_row + 1):
        for offset in range(3):  # K=11, L=12, M=13 → F=6, G=7, H=8
            src_col = 11 + offset
            dst_col = 6 + offset
            value = sheet_src.cell(row=row, column=src_col).value
            sheet_awb.cell(row=row, column=dst_col, value=value)

    # Save workbook
    wb.save(file_path)


def start(folder_path, folder_path_ref, AE_expected_path, AWB_expected_path, original_excel, new_excel, EV_ratio, AWB_num_blocks_x=40, AWB_num_blocks_y=40):
    # if not os.path.exists(folder_path):
    #     print(f"Error: The folder '{folder_path}' does not exist.")
    #     return
    # if not os.path.exists(folder_path_ref):
    #     print(f"Error: The folder '{folder_path_ref}' does not exist.")
    #     return

    print("Success input. Start the process ...")
    print(f"AE path : {AE_expected_path}")
    print(f"AWB path : {AWB_expected_path}")

    # # Ensure the AWB expected path exists
    # os.makedirs(AWB_expected_path, exist_ok=True)

    # Get list of image files in each folder
    images = sorted(
        [f for f in os.listdir(folder_path) if os.path.isfile(
            os.path.join(folder_path, f)) and f.lower().endswith(('jpg', 'jpeg', 'JPG', 'JPEG'))]
    )
    images_ref = sorted(
        [f for f in os.listdir(folder_path_ref) if os.path.isfile(
            os.path.join(folder_path_ref, f)) and f.lower().endswith(('jpg', 'jpeg', 'JPG', 'JPEG'))]
    )

    # Ensure both folders have the same number of files and filenames match
    if len(images) != len(images_ref):
        print(images)
        print(images_ref)
        print("Error: The number of images in the two folders do not match.")
    else:
        # Step 1: Duplicate the original Excel file
        shutil.copy(original_excel, new_excel)

        # Step 2: Load the duplicated Excel file
        workbook = load_workbook(new_excel)

        # Step 3: Process images and calculate sum of squares
        results = []
        tmp_to_delete = []
        row = 3  # Start inserting from row 3 (assuming headers in row 1 and 2)
        with tqdm(total=len(images), desc="Processing Image Pairs", unit="pair") as pbar:
            for img, img_ref in zip(images, images_ref):
                if img.split('_')[0] == img_ref.split('_')[0]:
                    img_index = img.split('_')[0]

                    # Paths to the images
                    image_path = os.path.normpath(os.path.join(folder_path, img))
                    image_path_ref = os.path.normpath(os.path.join(folder_path_ref, img_ref))

                    # Load images
                    # image1 = load_image(image_path)
                    # image2 = load_image(image_path_ref)
                    image1 = cv2.imread(image_path)  # BGR
                    image2 = cv2.imread(image_path_ref)  # BGR

                    # Calculate the differences
                    recommand_EV_changes = single_pair_EV_compair(
                        folder_path, folder_path_ref, img, img_ref)

                    # Adjust weights if needed (example for neutral exposure)
                    if recommand_EV_changes[2] < 0:
                        recommand_EV_changes[2] = recommand_EV_changes[2] * 4
                   
                    # Calculate RGB gain respectively
                    # _, RBG_gains = color_transfer(folder_path, folder_path_ref, img, img_ref) ## 0122 delete old RGB
                    # 2) Align image1's color to match image2's color in the chosen block
                    roi_source, roi_target, annotated1, annotated2, modified_image, RBG_gains = align_image_by_block(
                        image1, image2,
                        AWB_num_blocks_x,
                        AWB_num_blocks_y
                    )

                    # Scale differences and calculate sum of squares
                    diff_group = [int(x * EV_ratio)
                                  for x in recommand_EV_changes]
                    ae_sum_of_squares = sum([x**2 for x in diff_group])


                    # Show the cropped block
                    # plt_awb_result(annotated1, annotated2, modified_image)

                    # RBG_gains = compute_rgb_gains_from_roi(roi_source, roi_target)

                    # Save expected AWB image
                    # result_image = color_transfer_for_AWB_expexted_output(folder_path, folder_path_ref, img, img_ref) #####0122 delete old
                    # result_image = modified_image  # 0122 add
                    awb_img = 'AWB_' + img
                    awb_path = os.path.join(AWB_expected_path, awb_img)
                    cv2.imwrite(awb_path, modified_image)
                    
                    tmp_to_delete.append(awb_path)

                    # Store the result as a tuple
                    results.append((img_index, ae_sum_of_squares, img, awb_img,
                                    img_ref, diff_group, RBG_gains, modified_image))
                else:
                    print(f"Error: Mismatch in filenames {img} and {img_ref}")

                # Update the progress bar after each pair is processed
                pbar.update(1)  # Increment progress bar by 1

        # Step 4: Sort results by sum of squares in descending order
        # results.sort(reverse=True, key=lambda x: x[0])
        # print('Sorting completed.')

        print("Start writing results into Excel ...")
        # Step 6: Insert sorted results into "AE" worksheet
        # We will create two separate lists for "AE" and "AWB" and then sort them accordingly

        # Temporary lists for the AE and AWB sorted data
        ae_results = []
        awb_results = []
        overall_results = []

        # Step 5: Collect results
        for img_index, ae_sum_of_squares, img, awb_img, img_ref, diff_group, RBG_gains, modified_image in results:
            ae_results.append(
                (img_index, ae_sum_of_squares, img, awb_img, img_ref, diff_group, modified_image))
            
            awb_dev = rgb_dev(RBG_gains)
            
            awb_results.append(
                (img_index, awb_dev, img, awb_img, img_ref, RBG_gains, modified_image))
            
      
        ae_vals = np.array([x[1] for x in ae_results])
        awb_vals = np.array([x[1] for x in awb_results])

        ae_min, ae_max = ae_vals.min(), ae_vals.max()
        awb_min, awb_max = awb_vals.min(), awb_vals.max()
            
        for img_index, ae_sum_of_squares, img, awb_img, img_ref, diff_group, RBG_gains, modified_image in results:
            awb_dev = rgb_dev(RBG_gains)
            ae_norm = (ae_sum_of_squares - ae_min) / (ae_max - ae_min + 1e-8)
            awb_norm = (awb_dev - awb_min) / (awb_max - awb_min + 1e-8)

            overall_spec = (ae_norm + awb_norm) / 2
            overall_results.append(
                (img_index, overall_spec, img, awb_img, img_ref, diff_group, RBG_gains, modified_image))

        # Step 6: Sort results for "AE" by the statistic difference of EV values
        ae_results_sorted = sorted(ae_results, key=lambda x: np.linalg.norm(
            x[1]), reverse=True)  # Sort by `diff_group`'s statistic difference

        # Step 7: Insert sorted results into "AE" worksheet
        ae_sheet = workbook['AE']
        row = 3  # Start inserting from row 3 (assuming headers in row 1 and 2)

        for img_index, ae_sum_of_squares, img, awb_img, img_ref, diff_group, modified_image in ae_results_sorted:
            # Insert img index
            ae_sheet[f'C{row}'] = img_index

            # Insert EV values (-2EV, -1EV, 0EV, +1EV, +2EV) into AE worksheet
            ae_sheet[f'G{row}'] = diff_group[0]  # -2EV
            ae_sheet[f'H{row}'] = diff_group[1]  # -1EV
            ae_sheet[f'I{row}'] = diff_group[2]  # 0EV
            ae_sheet[f'J{row}'] = diff_group[3]  # +1EV
            ae_sheet[f'K{row}'] = diff_group[4]  # +2EV

            # Insert ae_sum_of_squares and apply color
            cell = ae_sheet[f'B{row}']
            cell.value = ae_sum_of_squares
            
            # Add severity, Place result images in corresponding folder
            if ae_sum_of_squares > 60: 
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # critical
                severity_path = os.path.join(AE_expected_path, "critical")
                new_img_path = os.path.join(severity_path, img)
                new_img_ref_path = os.path.join(severity_path, img_ref)

                # Copy the image
                shutil.copy(image_path, new_img_path)
                shutil.copy(image_path_ref, new_img_ref_path)
            elif ae_sum_of_squares > 30:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # serious
                severity_path = os.path.join(AE_expected_path, "serious")
                new_img_path = os.path.join(severity_path, img)
                new_img_ref_path = os.path.join(severity_path, img_ref)

                # Copy the image
                shutil.copy(image_path, new_img_path)
                shutil.copy(image_path_ref, new_img_ref_path)
            else:
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # moderate
                severity_path = os.path.join(AE_expected_path, "moderate")
                new_img_path = os.path.join(severity_path, img)
                new_img_ref_path = os.path.join(severity_path, img_ref)

                # Copy the image
                shutil.copy(image_path, new_img_path)
                shutil.copy(image_path_ref, new_img_ref_path)

            
            temp_img_path = resize_image_to_fit(os.path.join(folder_path, img))
            temp_awb_img_path = resize_image_to_fit(
                os.path.join(AWB_expected_path, awb_img))
            temp_ref_img_path = resize_image_to_fit(
                os.path.join(folder_path_ref, img_ref))
            excel_img = Image(temp_img_path)
            excel_awb_img = Image(temp_awb_img_path)
            excel_ref_img = Image(temp_ref_img_path)

            image_path = os.path.normpath(os.path.join(folder_path, img))
            image_path_ref = os.path.normpath(os.path.join(folder_path_ref, img_ref))

            # Insert images into AE worksheet
            ae_sheet.add_image(excel_img, f'D{row}')
            ae_sheet.add_image(excel_awb_img, f'E{row}')
            ae_sheet.add_image(excel_ref_img, f'F{row}')

            row += 1
            # tmp_to_delete.append(temp_img_path)
            # tmp_to_delete.append(temp_ref_img_path)

        print("AE finished.")

        # Step 8: Sort results for "AWB" by the statistic difference of RGB gains
        awb_results_sorted = sorted(awb_results, key=lambda x: np.linalg.norm(
            x[1]), reverse=True)  # Sort by `RBG_gains`'s statistic difference

        # Step 9: Insert sorted results into "AWB" worksheet
        awb_sheet = workbook['AWB']
        row = 3  # Start inserting from row 3 (assuming headers in row 1 and 2)
        for img_index, awb_dev, img, awb_img, img_ref, RBG_gains, modified_image in awb_results_sorted:
            temp_img_path = resize_image_to_fit(os.path.join(folder_path, img))
            temp_awb_img_path = resize_image_to_fit(
                os.path.join(AWB_expected_path, awb_img))
            temp_ref_img_path = resize_image_to_fit(
                os.path.join(folder_path_ref, img_ref))
            excel_img = Image(temp_img_path)
            excel_awb_img = Image(temp_awb_img_path)
            excel_ref_img = Image(temp_ref_img_path)

            image_path = os.path.normpath(os.path.join(folder_path, img))
            image_path_ref = os.path.normpath(os.path.join(folder_path_ref, img_ref))

            # Insert img index
            awb_sheet[f'C{row}'] = img_index

            # Insert images into AWB worksheet
            awb_sheet.add_image(excel_img, f'D{row}')
            awb_sheet.add_image(excel_awb_img, f'E{row}')
            awb_sheet.add_image(excel_ref_img, f'F{row}')

            # Insert RGB gains into AWB worksheet
            awb_sheet[f'G{row}'] = RBG_gains[0]  # R gain
            awb_sheet[f'H{row}'] = RBG_gains[1]  # B gain
            awb_sheet[f'I{row}'] = RBG_gains[2]  # G gain

            cell = awb_sheet[f'B{row}']
            cell.value = f"{awb_dev:.2f}"  # or f"{awb_dev:.3f}" for formatting

            if awb_dev >= 0.5:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # critical
                severity_path = os.path.join(AWB_expected_path, "critical")
                new_img_path = os.path.join(severity_path, img)
                new_img_ref_path = os.path.join(severity_path, img_ref)

                # Copy the image
                shutil.copy(image_path, new_img_path)
                shutil.copy(image_path_ref, new_img_ref_path)

                # Add AWB modified image
                basename, ext = os.path.splitext(img_ref)
                modified_image_name = basename + "_Modified" + ext
                cv2.imwrite(os.path.join(severity_path, modified_image_name), modified_image)
            elif awb_dev >= 0.2:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # serious
                severity_path = os.path.join(AWB_expected_path, "serious")
                new_img_path = os.path.join(severity_path, img)
                new_img_ref_path = os.path.join(severity_path, img_ref)

                # Copy the image
                shutil.copy(image_path, new_img_path)
                shutil.copy(image_path_ref, new_img_ref_path)

                # Add AWB modified image
                basename, ext = os.path.splitext(img_ref)
                modified_image_name = basename + "_Modified" + ext
                cv2.imwrite(os.path.join(severity_path, modified_image_name), modified_image)
            else:
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # moderate
                severity_path = os.path.join(AWB_expected_path, "moderate")
                new_img_path = os.path.join(severity_path, img)
                new_img_ref_path = os.path.join(severity_path, img_ref)

                # Copy the image
                shutil.copy(image_path, new_img_path)
                shutil.copy(image_path_ref, new_img_ref_path)

                # Add AWB modified image
                basename, ext = os.path.splitext(img_ref)
                modified_image_name = basename + "_Modified" + ext
                cv2.imwrite(os.path.join(severity_path, modified_image_name), modified_image)

            row += 1
            # tmp_to_delete.append(temp_img_path)
            # tmp_to_delete.append(temp_ref_img_path)

        print("AWB finished.")

        overall_results_sorted = sorted(overall_results, key=lambda x: np.linalg.norm(
            x[1]), reverse=True)
        
        overall_sheet = workbook['OverAll']
        row = 3
        # Step 5: Insert sorted results into Excel
        for img_index, overall_spec, img, awb_img, img_ref, diff_group, RBG_gains, modified_image in overall_results_sorted:
            temp_img_path = resize_image_to_fit(os.path.join(folder_path, img))
            temp_awb_img_path = resize_image_to_fit(
                os.path.join(AWB_expected_path, awb_img))
            temp_ref_img_path = resize_image_to_fit(
                os.path.join(folder_path_ref, img_ref))
            excel_img = Image(temp_img_path)
            excel_awb_img = Image(temp_awb_img_path)
            excel_ref_img = Image(temp_ref_img_path)

            # Insert img index
            overall_sheet[f'C{row}'] = img_index

            # Insert images into Excel
            overall_sheet.add_image(excel_img, f'D{row}')
            overall_sheet.add_image(excel_awb_img, f'E{row}')
            overall_sheet.add_image(excel_ref_img, f'F{row}')

            # Insert EV values (-2EV, -1EV, 0EV, +1EV, +2EV) into Excel
            overall_sheet[f'G{row}'] = diff_group[0]  # -2EV
            overall_sheet[f'H{row}'] = diff_group[1]  # -1EV
            overall_sheet[f'I{row}'] = diff_group[2]  # 0EV
            overall_sheet[f'J{row}'] = diff_group[3]  # +1EV
            overall_sheet[f'K{row}'] = diff_group[4]  # +2EV

            # Insert RGB gains into Excel
            overall_sheet[f'L{row}'] = RBG_gains[0]  # R gain
            overall_sheet[f'M{row}'] = RBG_gains[1]  # B gain
            overall_sheet[f'N{row}'] = RBG_gains[2]  # G gain

            cell = overall_sheet[f'B{row}']
            if overall_spec >= 0.7:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif overall_spec >= 0.35:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")



            row += 1
            # tmp_to_delete.append(temp_img_path)
            # tmp_to_delete.append(temp_ref_img_path)

        # Clean up temporary files
        # for tmp_file in tmp_to_delete:
        #     os.remove(tmp_file)
        for filepath in tmp_to_delete:
            if os.path.exists(filepath):
                os.remove(filepath)

        # Step 6: Save the modified Excel file
        workbook.save(new_excel)
        print(f"Images successfully inserted and saved to {new_excel}")

    print("Complete!")
    print("--------------------------------------------------")

def rgb_dev(RBG_gains):
    r_gain, b_gain, _ = RBG_gains
    r_dev = abs(r_gain - 1)
    b_dev = abs(b_gain - 1)
    awb_dev = r_dev + b_dev

    return awb_dev

def find_common_gray_block_roi_relative(image1, image2, num_blocks_x=4, num_blocks_y=4):
    """
    1) Resizes both images to the same dimensions (based on image1 size).
    2) Divides [0..1, 0..1] into a (num_blocks_x x num_blocks_y) grid.
    3) For each grid cell, selects the relative block in each image.
    4) Computes how close each block pair is to (R/G=1, B/G=1).
    5) Return:
        - roi_source, roi_target: The chosen block region in image1, image2
        - annotated1, annotated2: The full images with the chosen block highlighted
    """

    # (A) Resize image2 to match the dimensions of image1
    h1, w1 = image1.shape[:2]
    h2, w2 = image2.shape[:2]
    if (h1, w1) != (h2, w2):  # Resize only if dimensions differ
        image2_resized = cv2.resize(
            image2, (w1, h1), interpolation=cv2.INTER_LINEAR)
    else:
        image2_resized = image2.copy()

    # Scaling factors for relative location
    width_factor = w2 / w1
    height_factor = h2 / h1

    best_distance = float("inf")
    best_coords = (0, 0, 0, 0)  # (x1, y1, x2, y2)

    # (B) Iterate over grid blocks
    for row in range(num_blocks_y):
        for col in range(num_blocks_x):
            # (B1) Normalized grid coordinates
            x1_norm = col / num_blocks_x
            x2_norm = (col + 1) / num_blocks_x
            y1_norm = row / num_blocks_y
            y2_norm = (row + 1) / num_blocks_y

            # (B2) Convert to pixel coordinates
            x1 = int(np.floor(x1_norm * w1))
            x2 = int(np.floor(x2_norm * w1))
            y1 = int(np.floor(y1_norm * h1))
            y2 = int(np.floor(y2_norm * h1))

            # (B3) Crop blocks from both images
            block1 = image1[y1:y2, x1:x2]
            block2 = image2_resized[y1:y2, x1:x2]

            # Skip empty blocks
            if block1.size == 0 or block2.size == 0:
                continue

            # (C) Distance from neutral gray for each block
            avg_bgr1 = np.mean(block1, axis=(0, 1))  # BGR
            G1 = max(avg_bgr1[1], 1e-6)
            r_g1 = avg_bgr1[2] / G1
            b_g1 = avg_bgr1[0] / G1
            dist1 = np.sqrt((r_g1 - 1)**2 + (b_g1 - 1)**2)

            avg_bgr2 = np.mean(block2, axis=(0, 1))  # BGR
            G2 = max(avg_bgr2[1], 1e-6)
            r_g2 = avg_bgr2[2] / G2
            b_g2 = avg_bgr2[0] / G2
            dist2 = np.sqrt((r_g2 - 1)**2 + (b_g2 - 1)**2)

            combined_dist = dist1 + dist2

            # (D) Update best block
            if combined_dist < best_distance:
                best_distance = combined_dist
                best_coords = (x1, y1, x2, y2)

    # (E) Crop the best block from both images
    x1, y1, x2, y2 = best_coords
    # print(f"W: {w1}, H: {h1}")
    # print(x1)
    # print(y1)
    # print(x2)
    # print(y2)
    # Resized
    x1_resized = int(x1 * width_factor)
    y1_resized = int(y1 * height_factor)
    x2_resized = int(x2 * width_factor)
    y2_resized = int(y2 * height_factor)
    # print(f"W: {w2}, H: {h2}")
    # print(x1_resized)
    # print(y1_resized)
    # print(x2_resized)
    # print(y2_resized)

    roi_source = image1[y1:y2, x1:x2]
    roi_target = image2[y1:y2, x1:x2]

    # (F) Annotate the rectangles
    annotated1 = image1.copy()
    annotated2 = image2.copy()
    cv2.rectangle(annotated1, (x1, y1), (x2, y2), (0, 255, 0), 10)
    cv2.rectangle(annotated2, (x1_resized, y1_resized),
                  (x2_resized, y2_resized), (0, 255, 0), 10)

    return roi_source, roi_target, annotated1, annotated2


def align_image_by_block(image1, image2, num_blocks_x=4, num_blocks_y=4):
    """
    1) Finds a 'best gray' block in each image using relative grids.
    2) Computes color scaling factors (for B, G, R) to align image1's block to image2's block.
    3) Applies those factors to the entire image1, producing 'modified_image'.
    4) Dynamically handles outdoor images to prevent green dominance.
    5) Returns:
        roi_source, roi_target      -> The chosen blocks
        annotated1, annotated2 -> Annotated images
        modified_image   -> The color-modified version of image1
    """
    # # Step 0: Handle BGR to RGB if needed
    # # If using OpenCV, images are in BGR format. Convert to RGB for processing.
    # image1 = image1[:, :, ::-1]  # Swap BGR to RGB
    # image2 = image2[:, :, ::-1]  # Swap BGR to RGB

    # # Ensure the image is scaled to uint8 (if it's float64 from skimage)
    # if image1.dtype == np.float64:  # Skimage loads as float64 by default
    #     image1 = (image1 * 255).clip(0, 255).astype(np.uint8)
    # if image2.dtype == np.float64:  # Skimage loads as float64 by default
    #     image2 = (image2 * 255).clip(0, 255).astype(np.uint8)

    # Step A: Find the common block
    roi_source, roi_target, annotated1, annotated2 = find_common_gray_block_roi_relative(
        image1,
        image2,
        num_blocks_x,
        num_blocks_y
    )

    # Step B: Compute R/G/B ratios for all three channels
    avg_bgr1 = np.mean(roi_source, axis=(0, 1))  # [B, G, R] in image1
    avg_bgr2 = np.mean(roi_target, axis=(0, 1))  # [B, G, R] in image2

    r_gain = avg_bgr2[2] / avg_bgr1[2] if avg_bgr1[2] != 0 else 1.0  # R ratio
    g_gain = avg_bgr2[1] / avg_bgr1[1] if avg_bgr1[1] != 0 else 1.0  # G ratio
    b_gain = avg_bgr2[0] / avg_bgr1[0] if avg_bgr1[0] != 0 else 1.0  # B ratio

    # Step C: Apply gains to the entire image1
    modified_image = image1.copy().astype(np.float32)
    modified_image[:, :, 2] *= r_gain  # Adjust R
    modified_image[:, :, 1] *= g_gain  # Adjust G
    modified_image[:, :, 0] *= b_gain  # Adjust B

    # Step D: Auto-exposure adjustment based on overall brightness
    brightness1 = np.mean(cv2.cvtColor(modified_image, cv2.COLOR_BGR2GRAY))
    brightness2 = np.mean(cv2.cvtColor(image2, cv2.COLOR_BGR2GRAY))

    exposure_adjustment = 1  # Default: no adjustment
    if brightness1 != 0:
        exposure_adjustment = brightness2 / brightness1
        modified_image *= exposure_adjustment

    # Final clip and conversion to uint8
    modified_image = np.clip(modified_image, 0, 255).astype(np.uint8)

    # Update gains with exposure adjustment
    r_gain *= exposure_adjustment
    g_gain *= exposure_adjustment
    b_gain *= exposure_adjustment
    
    return roi_source, roi_target, annotated1, annotated2, modified_image, [r_gain, b_gain, g_gain]

    # # Step B: Compute R/G and B/G ratios
    # avg_bgr1 = np.mean(roi_source, axis=(0, 1))  # [B, G, R] in image1
    # avg_bgr2 = np.mean(roi_target, axis=(0, 1))  # [B, G, R] in image2

    # r_gain = avg_bgr2[2] / avg_bgr1[2] if avg_bgr1[2] != 0 else 1.0  # R ratio
    # b_gain = avg_bgr2[0] / avg_bgr1[0] if avg_bgr1[0] != 0 else 1.0  # B ratio
    # # g_gain = avg_bgr2[1] / avg_bgr1[1] if avg_bgr1[1] != 0 else 1.0  # G

    # # Step C: Apply gains to the entire image1
    # modified_image = image1.copy().astype(np.float32)
    # modified_image[:, :, 2] *= r_gain  # Adjust R
    # modified_image[:, :, 0] *= b_gain  # Adjust B
    # # modified_image[:, :, 1] *= 1.0    # Keep G unchanged (exposure reference)

    # # Step D: Auto-exposure adjustment based on overall brightness
    # # Compute overall brightness as the average intensity across all channels
    # # Overall brightness of modified image
    # brightness1 = np.mean(cv2.cvtColor(modified_image, cv2.COLOR_BGR2GRAY))
    # # Overall brightness of target image
    # brightness2 = np.mean(cv2.cvtColor(image2, cv2.COLOR_BGR2GRAY))

    # exposure_adjustment = 1  # Default: no adjustment
    # if brightness1 != 0:
    #     # Compute exposure adjustment based on brightness
    #     exposure_adjustment = brightness2 / brightness1
    #     # Scale all channels equally
    #     modified_image *= exposure_adjustment

    # # Final clip and conversion to uint8
    # modified_image = np.clip(modified_image, 0, 255).astype(np.uint8)

    # # Update r_gain and b_gain with exposure adjustment
    # r_gain *= exposure_adjustment
    # b_gain *= exposure_adjustment

    # return roi_source, roi_target, annotated1, annotated2, modified_image, [r_gain, b_gain, 1]


def plt_awb_result(ann1, ann2, modified):
    # Instead of cv2.imshow(...), use Matplotlib
    # Convert from BGR (OpenCV) to RGB (Matplotlib) before showing
    ann1_rgb = cv2.cvtColor(ann1, cv2.COLOR_BGR2RGB)
    ann2_rgb = cv2.cvtColor(ann2, cv2.COLOR_BGR2RGB)
    modified_rgb = cv2.cvtColor(modified, cv2.COLOR_BGR2RGB)

    plt.figure(figsize=(15, 5))  # optional: adjust the figure size

    # Show annotated Image1
    plt.subplot(1, 3, 1)
    plt.imshow(ann1)
    plt.title("Chosen Block in Image1")
    plt.axis("off")

    # Show annotated Image2
    plt.subplot(1, 3, 2)
    plt.imshow(ann2)
    plt.title("Chosen Block in Image2")
    plt.axis("off")

    # Show modified Image1
    plt.subplot(1, 3, 3)
    plt.imshow(modified)
    plt.title("modified Image1")
    plt.axis("off")

    plt.tight_layout()
    plt.show()


# Not use
def compute_ycrcb_gains_from_roi(roi_source, roi_target):
    """
    Compute YCrCb gains from two ROIs:
      - The ratio Cr/Y and Cb/Y is taken from the target ROI (per-pixel),
        then averaged to get one factor for Cr and Cb channels.
      - The luminance_exposure_ratio is the ratio of average Y in target vs. source.

    Returns:
      [cr_gain, cb_gain, luminance_exposure_ratio]
    """

    # # Ensure input ROIs are the same size
    # # Assume target has larger shape
    # if roi_source.shape != roi_target.shape:
    #     roi_target = cv2.resize(
    #         roi_target, (roi_source.shape[1], roi_source.shape[0]))

    # Convert ROIs to YCrCb
    roi_source_ycrcb = cv2.cvtColor(roi_source, cv2.COLOR_BGR2YCrCb)
    roi_target_ycrcb = cv2.cvtColor(roi_target, cv2.COLOR_BGR2YCrCb)

    # 2) Split into (Y, Cr, Cb) channels
    y_s, cr_s, cb_s = cv2.split(roi_source_ycrcb)
    y_t, cr_t, cb_t = cv2.split(roi_target_ycrcb)

    # 3) Avoid dividing by zero
    eps = 1e-5
    cr_to_y_t = cr_t / (y_t + eps)  # ratio of Cr to Y in target
    cb_to_y_t = cb_t / (y_t + eps)  # ratio of Cb to Y in target

    # 4) Flatten & average
    cr_mean = np.mean(cr_to_y_t)
    cb_mean = np.mean(cb_to_y_t)

    # 5) Compare luminance means (for exposure matching)
    luminance_exposure_ratio = np.mean(y_t) / (np.mean(y_s) + eps)

    return [cr_mean, cb_mean, luminance_exposure_ratio]


# Not use,
def compute_rgb_gains_from_roi(roi_source, roi_target):
    """
    Compute RGB gains from two ROIs:
      - The ratio R/G and B/G is taken from the target ROI (per-pixel),
        then averaged to get one factor for R and B channels.
      - The green_exposure_ratio is the ratio of average G in target vs. source.

    Returns:
      [r_gain, b_gain, green_exposure_ratio]
    """
    # 1) Split into (R, G, B) or (B, G, R) depending on your data.
    #    Let's assume they are in *RGB* order for clarity:
    r_s, g_s, b_s = cv2.split(roi_source)
    r_t, g_t, b_t = cv2.split(roi_target)

    # 2) Avoid dividing by zero
    eps = 1e-5
    r_to_g_t = r_t / (g_t + eps)  # ratio of R to G in target
    b_to_g_t = b_t / (g_t + eps)  # ratio of B to G in target

    # 3) Flatten & average
    r_mean = np.mean(r_to_g_t)
    b_mean = np.mean(b_to_g_t)

    # 4) Compare green means (for exposure matching)
    green_exposure_ratio = np.mean(g_t) / (np.mean(g_s) + eps)

    return [r_mean, b_mean, green_exposure_ratio]
    # # Split channels
    # R_source, G_source, B_source = roi_source[:, :, 0], roi_source[:, :, 1], roi_source[:, :, 2]
    # R_target, G_target, B_target = roi_target[:, :, 0], roi_target[:, :, 1], roi_target[:, :, 2]

    # # Calculate mean of channels
    # mean_G_target = np.mean(G_target)
    # mean_R_source = np.mean(R_source)
    # mean_B_source = np.mean(B_source)

    # # Adjust R and B channels relative to G
    # R_modified = R_source * (mean_G_target / mean_R_source)
    # B_modified = B_source * (mean_G_target / mean_B_source)

    # # Stack adjusted channels
    # modified_image = np.stack([R_modified, G_source, B_modified], axis=-1)
    # return np.clip(modified_image, 0, 255).astype(np.uint8)


# # New Run

# In[116]:


# # Paths to the images
# folder_path = './test_data/20250106/LP3'
# folder_path_ref = './test_data/20250106/ref'
# original_excel = 'example5.xlsx'
# new_excel = 'output_with_images.xlsx'
# AWB_expected_path = './test_data/20250106/AWB_expected'
# EV_ratio = 6
# AWB_num_blocks_x = 6
# AWB_num_blocks_y = 6

# # Call the start function with updated paths
# start(folder_path, folder_path_ref, AWB_expected_path, original_excel, new_excel, EV_ratio, AWB_num_blocks_x, AWB_num_blocks_y)


# # Run

# In[ ]:


def main(folder_path, folder_path_ref, AE_expected_path, AWB_expected_path, original_excel, new_excel, EV_ratio, AWB_num_blocks_x, AWB_num_blocks_y):
    start(folder_path, folder_path_ref, AE_expected_path, AWB_expected_path, original_excel,
          new_excel, EV_ratio, AWB_num_blocks_x, AWB_num_blocks_y)


# python exceloutput.py --folder_path ./your/phone/image/path --folder_path_ref ./ref/phone/image/path --original_excel ./example.xlsx --new_excel ./output_excel_name.xlsx<br>
# %%
