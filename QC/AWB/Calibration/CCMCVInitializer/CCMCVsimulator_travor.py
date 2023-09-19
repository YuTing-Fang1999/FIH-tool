import cv2
import numpy as np
import os
import xlwings as xw
import time
import re
import openpyxl
from colour_checker_detection import detect_colour_checkers_segmentation


def create_xls():
    file_path = "CCMCVsimulator.xlsm"
    wb = openpyxl.load_workbook(file_path, read_only=False, keep_vba=True)
    wb.active = 0
    return wb

def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]

def file_filter_jpg(f):
    if f[-4:] in ['.jpg', '.JPG']:
        return True
    else:
        return False

def RGBtosRGB(rgb):
    srgb = []
    for i in range(0,3):
        if rgb[i] > 0.00304:
            V = (1+0.055)*((rgb[i])**(1/2.4))-0.055
            srgb.append(round(V*255,2))
        else:
            V = 12.92*rgb[i]
            srgb.append(round(V*255,2))
    return srgb

print("CCMCVsimulator is runing...")

dir_path = "Macbeth"
allFileList = os.listdir(dir_path)
allFileList_jpg = np.sort(allFileList,axis=0)
allFileList_jpg = list(filter(file_filter_jpg, allFileList_jpg))
allFileList_jpg.sort(key=natural_keys)
allFileList_jpg = [os.path.join(dir_path, f) for f in allFileList_jpg]

localtime = time.localtime()
clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])

wb = create_xls()
file = f"CCMCVsimulator_{localtime[0]}_{localtime[1]}_{localtime[2]}_{clock}.xlsm"
wb.active = 0
wb.save(file)

app = xw.App(visible=False)
wb = xw.Book(file)
macro_vba = wb.app.macro('CopySheetWithChart')

# fn = 'CCMCVsimulator.xlsm'
# app = xw.App(visible=False)
# wb = app.books.open(fn)
# ws = wb.sheets[0]

i = 0
while i < np.size(allFileList_jpg):
    macro_vba(os.path.basename(allFileList_jpg[i]).split('_')[0])
    i+=2
wb.sheets[0].activate()
wb.save()
app.quit()

wb = openpyxl.load_workbook(file, read_only=False, keep_vba=True)
i = 0
while i < np.size(allFileList_jpg):
    wb.active = i//2 + 4
    ws = wb.active
    print(os.path.basename(allFileList_jpg[i]))
    
    path_name1 = allFileList_jpg[i]
    path_name2 = allFileList_jpg[i+1]
    base = os.path.splitext(os.path.basename(path_name1))[0][:-2]
    ws.cell(range="B8").value = base
    # ws.cell(range='B8').value = base
    
    img1 = cv2.imread(path_name1, cv2.IMREAD_COLOR)
    img2 = cv2.imread(path_name2, cv2.IMREAD_COLOR)
    img1_crop = detect_colour_checkers_segmentation(img1, additional_data=True)[0]
    img2_crop = detect_colour_checkers_segmentation(img2, additional_data=True)[0]
    
    for j in range(0,24):
        ws.cell(range=f'B{j+15}').value = RGBtosRGB(img1_crop[0][j])[0]
        ws.cell(range=f'C{j+15}').value = RGBtosRGB(img1_crop[0][j])[1]
        ws.cell(range=f'D{j+15}').value = RGBtosRGB(img1_crop[0][j])[2]
        ws.cell(range=f'I{j+15}').value = RGBtosRGB(img2_crop[0][j])[0]
        ws.cell(range=f'J{j+15}').value = RGBtosRGB(img2_crop[0][j])[1]
        ws.cell(range=f'K{j+15}').value = RGBtosRGB(img2_crop[0][j])[2]
        # ws.cell(range=f'B{j+15}').value = RGBtosRGB(img1_crop[0][j])[0]
        # ws.cell(range=f'C{j+15}').value = RGBtosRGB(img1_crop[0][j])[1]
        # ws.cell(range=f'D{j+15}').value = RGBtosRGB(img1_crop[0][j])[2]
        # ws.cell(range=f'I{j+15}').value = RGBtosRGB(img2_crop[0][j])[0]
        # ws.cell(range=f'J{j+15}').value = RGBtosRGB(img2_crop[0][j])[1]
        # ws.cell(range=f'K{j+15}').value = RGBtosRGB(img2_crop[0][j])[2]
    
    i+=2
wb.save()
# else:
#     print("CCMCV simulator only supports one set of photos at a time!")

print("CCMCVsimulator is ok!")
# os.system("pause")