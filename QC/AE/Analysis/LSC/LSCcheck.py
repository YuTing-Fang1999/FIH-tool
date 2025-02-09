import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog
import openpyxl
import xlwings as xw
import time

def create_xls(file_path):
    fn = 'LSC_checkTool.xlsm'
    wb = openpyxl.load_workbook(fn, read_only=False, keep_vba=True)
    wb.active = 0
    return wb

print("LSC check is runing...")
print("Please scelect a lsc.xml.")

# create root window
root = tk.Tk()
root.withdraw()  # Hide the root window

# Open file dialog
file_path = filedialog.askopenfilename(filetypes=[("XML files", "*.xml")])
print(f"Selected file: {file_path}")

tree = ET.parse(file_path)
root = tree.getroot()

localtime = time.localtime()
clock = str(60*60*localtime[3] + 60*localtime[4] + localtime[5])

def find_lsc_data(root, lsc_type, aec_exp_control):
    data = []
    golden_data = []
    
    lsc_golden_rgn_data = root.find(f'.//{lsc_type}_golden_rgn_data')
    r_gain_values = lsc_golden_rgn_data.find('r_gain_tab/r_gain').text
    r_gain_list = [float(value) for value in r_gain_values.split()]
    gr_gain_values = lsc_golden_rgn_data.find('gr_gain_tab/gr_gain').text
    gr_gain_list = [float(value) for value in gr_gain_values.split()]
    gb_gain_values = lsc_golden_rgn_data.find('gb_gain_tab/gb_gain').text
    gb_gain_list = [float(value) for value in gb_gain_values.split()]
    b_gain_values = lsc_golden_rgn_data.find('b_gain_tab/b_gain').text
    b_gain_list = [float(value) for value in b_gain_values.split()]
    
    golden_data = {
        "r_gain": r_gain_list,
        "gr_gain": gr_gain_list,
        "gb_gain": gb_gain_list,
        "b_gain": b_gain_list
    }
    
    if aec_exp_control == "control_lux_idx":
        for i, mod_lsc_aec_data in enumerate(root.findall(f'.//mod_{lsc_type}_aec_data')):
            aec_trigger = mod_lsc_aec_data.find('aec_trigger')
            start = aec_trigger.find('lux_idx_start').text
            end = aec_trigger.find('lux_idx_end').text
        
            for mod_lsc_cct_data in mod_lsc_aec_data.findall(f'.//mod_{lsc_type}_cct_data'):
                cct_trigger = mod_lsc_cct_data.find('cct_trigger')
                cct_start = cct_trigger.find('start').text
                cct_end = cct_trigger.find('end').text
                
                lsc_rgn_data = mod_lsc_cct_data.find(f'{lsc_type}_rgn_data')
                r_gain = lsc_rgn_data.find('r_gain_tab/r_gain').text
                gr_gain = lsc_rgn_data.find('gr_gain_tab/gr_gain').text
                gb_gain = lsc_rgn_data.find('gb_gain_tab/gb_gain').text
                b_gain = lsc_rgn_data.find('b_gain_tab/b_gain').text
        
                data.append({
                    "lux_idx_start": start,
                    "lux_idx_end": end,
                    "cct_data": {
                        "start": cct_start,
                        "end": cct_end,
                        "r_gain": r_gain,
                        "gr_gain": gr_gain,
                        "gb_gain": gb_gain,
                        "b_gain": b_gain,
                    }
                })
    else:
        for i, mod_lsc_aec_data in enumerate(root.findall(f'.//mod_{lsc_type}_aec_data')):
            aec_trigger = mod_lsc_aec_data.find('aec_trigger')
            start = aec_trigger.find('gain_start').text
            end = aec_trigger.find('gain_end').text
        
            for mod_lsc_cct_data in mod_lsc_aec_data.findall(f'.//mod_{lsc_type}_cct_data'):
                cct_trigger = mod_lsc_cct_data.find('cct_trigger')
                cct_start = cct_trigger.find('start').text
                cct_end = cct_trigger.find('end').text
                
                lsc_rgn_data = mod_lsc_cct_data.find(f'{lsc_type}_rgn_data')
                r_gain = lsc_rgn_data.find('r_gain_tab/r_gain').text
                gr_gain = lsc_rgn_data.find('gr_gain_tab/gr_gain').text
                gb_gain = lsc_rgn_data.find('gb_gain_tab/gb_gain').text
                b_gain = lsc_rgn_data.find('b_gain_tab/b_gain').text
        
                data.append({
                    "gain_start": start,
                    "gain_end": end,
                    "cct_data": {
                        "start": cct_start,
                        "end": cct_end,
                        "r_gain": r_gain,
                        "gr_gain": gr_gain,
                        "gb_gain": gb_gain,
                        "b_gain": b_gain,
                    }
                })
            
    return golden_data, data

if 'lsc34' in root.tag:
    lsc_type = 'lsc34'
elif 'lsc35' in root.tag:
    lsc_type = 'lsc35'
else:
    raise ValueError('Unsupported LSC type in XML')

control_method = root.find('.//control_method')
aec_exp_control = control_method.find('aec_exp_control').text    
golden_data, data = find_lsc_data(root, lsc_type, aec_exp_control)

wb = create_xls(file_path)
    
file = f'LSC_checkTool_{localtime[0]}_{localtime[1]}_{localtime[2]}_{clock}.xlsm'
wb.active = 0
wb.save(file)

app = xw.App(visible=False)
wb = xw.Book(file)
macro_vba = wb.app.macro('CopySheetWithChart')

for i, item in enumerate(data, start=2):
    if aec_exp_control == "control_lux_idx":
        print(f'region {str(i).rjust(3)}: '
              f'lux_idx_start: {str(item["lux_idx_start"]).rjust(5)}, '
              f'lux_idx_end: {str(item["lux_idx_end"]).rjust(5)}, '
              f'start: {str(item["cct_data"]["start"]).rjust(5)}, '
              f'end: {str(item["cct_data"]["end"]).rjust(5)}')
        sheet_name = f'lux_{item["lux_idx_start"]}_{item["lux_idx_end"]}_cct_{item["cct_data"]["start"]}_{item["cct_data"]["end"]}'
    else:
        print(f'region {str(i).rjust(3)}: '
              f'gain_start: {str(item["gain_start"]).rjust(5)}, '
              f'gain_end: {str(item["gain_end"]).rjust(5)}, '
              f'start: {str(item["cct_data"]["start"]).rjust(5)}, '
              f'end: {str(item["cct_data"]["end"]).rjust(5)}')
        sheet_name = f'gain_{item["gain_start"]}_{item["gain_end"]}_cct_{item["cct_data"]["start"]}_{item["cct_data"]["end"]}'
    macro_vba(sheet_name)

wb.sheets[0].activate()
wb.save()
app.quit()

wb = openpyxl.load_workbook(file, read_only=False, keep_vba=True)
for i, item in enumerate(data, start=2):
    wb.active = i
    ws = wb.active
    r_gain_values = item["cct_data"]["r_gain"].split()
    gr_gain_values = item["cct_data"]["gr_gain"].split()
    gb_gain_values = item["cct_data"]["gb_gain"].split()
    b_gain_values = item["cct_data"]["b_gain"].split()
    for j, r_gain in enumerate(r_gain_values, start=3):
        ws.cell(row=j, column=3).value = round(float(r_gain_values[j - 3]), 3)
    for j, gr_gain in enumerate(gr_gain_values, start=3):
        ws.cell(row=j, column=4).value = round(float(gr_gain_values[j - 3]), 3)
    for j, gb_gain in enumerate(gb_gain_values, start=3):
        ws.cell(row=j, column=5).value = round(float(gb_gain_values[j - 3]), 3)
    for j, b_gain in enumerate(b_gain_values, start=3):
        ws.cell(row=j, column=6).value = round(float(b_gain_values[j - 3]), 3)


wb.active = 0
ws = wb.active
r_gain_values = golden_data["r_gain"]
gr_gain_values = golden_data["gr_gain"]
gb_gain_values = golden_data["gb_gain"]
b_gain_values = golden_data["b_gain"]
for j in range(3, 3 + len(r_gain_values)):
    ws.cell(row=j, column=3).value = r_gain_values[j - 3]
for j in range(3, 3 + len(gr_gain_values)):
    ws.cell(row=j, column=4).value = gr_gain_values[j - 3]
for j in range(3, 3 + len(gb_gain_values)):
    ws.cell(row=j, column=5).value = gb_gain_values[j - 3]
for j in range(3, 3 + len(b_gain_values)):
    ws.cell(row=j, column=6).value = b_gain_values[j - 3]

wb.save(file)
print("LSCcheck is ok!")
